"""
xtm_upload_translations.py  —  XTM Workbench: upload revised translations from Excel

Reads segment IDs (Column A) and translation text (Column C) from a
*_revised_translation_checks*.xlsx file and pushes each translation into the
matching XTM segment via the STOMP WebSocket protocol.

Rows 1–3 of the Excel are header rows and are skipped automatically.
Column A must contain the integer segment ID that matches the XTM unitId.
Report column written to Excel

Usage:
    - activate project manually in XTM,
    - set TEST_SEGMENT_LIMIT to 10-15 since session expire is not solved yet, and START_FROM_SEGMENT_ID to the desired segment ID, the server crashes frequently on hickups, saving and tag issues anyway, so better repeat the script.
    - then run this script with the project ID as argument:
    python xtm_upload_translations.py <project_id>
    e.g.  python xtm_upload_translations.py AIPX_2604_P0012
"""

import json
import os
import random
import re
import string
import sys
import time
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl
import requests
import websocket as _websocket
from dotenv import load_dotenv

from xtm_initial_download import (
    BASE_URL,
    WB_BASE,
    _find_pre_folder,
    _get_tasks,
    _init_workbench,
    _keepalive,
    _load_creds,
    _login,
    _find_task,
)

AUTO_CONFIRM_MATCHES    = True   # save ICE / 100% / internal-repetition segments using XTM pre-fill; fuzzy (<100%) always use Excel
KEEPALIVE_INTERVAL      = 25     # seconds between /sayHelloToServer.serv calls
TEST_SEGMENT_LIMIT: int | None = None
START_FROM_SEGMENT_ID: int     = 3
DEBUG_SOURCE_NODES_LIMIT       = 0
UPLOAD_BATCH_SIZE              = 15   # re-open editor every N segments to get a fresh session token
FILE_FILTER: str | None        = None  # set via --file; selects task by filename substring
BATCH_WAIT_SECONDS             = 120  # wait between batches for server to release doc lock


# ---------------------------------------------------------------------------
# Claim group task (USERGROUP → INTERNALLINGUIST)
# Doesnt work yet, activate project manually ---------------------------------------------------------------------------

def _claim_group_task(
    session: requests.Session, task: dict, uust: str, project_id: str
) -> dict:
    """If the task actor is USERGROUP, claim it so the actor becomes INTERNALLINGUIST.

    XTM's openEditor.serv will return 'There is no requested document' when called
    with actorType=INTERNALLINGUIST on an unclaimed USERGROUP task.  This step must
    run before _open_editor_write.  Returns the (possibly updated) task dict.
    """
    ad = task["additionalData"]
    actor_type = ad.get("actorType", "USERGROUP")

    if actor_type == "INTERNALLINGUIST":
        print("  Task already assigned to INTERNALLINGUIST, skipping claim step.")
        return task

    print(f"  Task actor is '{actor_type}', claiming as INTERNALLINGUIST...")

    # Log CONTEXT_MENU so the correct endpoint can be identified if this fails
    context_menu = task.get("CONTEXT_MENU", "")
    if context_menu:
        print(f"  CONTEXT_MENU: {str(context_menu)[:600]}")

    r = session.post(
        f"{BASE_URL}/myinbox/takeGroupTask.serv",
        data={
            "stepReferenceId": ad["stepReferenceId"],
            "fileId":          ad["fileId"],
            "uust":            uust,
        },
    )
    print(f"  Claim response: {r.status_code} — {r.text[:300]}")

    if not r.ok:
        print("  Warning: claim request returned an error status — actor may still be USERGROUP.")
        return task

    try:
        resp_json = r.json()
    except Exception:
        resp_json = {}

    if resp_json.get("session-expired"):
        print(
            "  Session expired during claim — the uust token is no longer valid.\n"
            "  Automatic claim did not complete."
        )
        return task

    # Re-fetch the task list so the updated actorType is visible
    time.sleep(4)
    updated_tasks = _get_tasks(session)
    updated_task = _find_task(updated_tasks, project_id, file_filter=FILE_FILTER)
    new_actor = updated_task["additionalData"].get("actorType", "?")
    print(f"  Updated actor type: {new_actor}")
    return updated_task


# ---------------------------------------------------------------------------
# Open editor for writing (readOnly=false)
# ---------------------------------------------------------------------------

def _open_editor_write(
    session: requests.Session, task: dict, uust: str
) -> tuple[str, str]:
    """POST to openEditor.serv with readOnly=false, return (wb_url, session_token)."""
    ad = task["additionalData"]
    payload = {
        "actorType":                  "INTERNALLINGUIST",
        "fileId":                     ad["fileId"],
        "groupTaskAction":            "undefined",
        "isGroupTaskWithPriorities":  "false",
        "readOnly":                   "false",
        "workflowStepName":           ad.get("stepName", ""),
        "stepReferenceId":            ad["stepReferenceId"],
        "taskType":                   "ACTIVE",
        "workflowReferenceStepName":  ad.get("stepReferenceName", ""),
        "uust":                       uust,
    }
    _MAX_ATTEMPTS = 3
    for attempt in range(1, _MAX_ATTEMPTS + 1):
        r = session.post(
            f"{BASE_URL}/openEditor.serv",
            data=payload,
            allow_redirects=False,
        )
        r.raise_for_status()
        xml = r.text
        if "<result>error</result>" not in xml:
            break
        msg_m = re.search(r"<msg>(.*?)</msg>", xml)
        msg = msg_m.group(1) if msg_m else xml[:200]
        if attempt == _MAX_ATTEMPTS:
            raise RuntimeError(f"openEditor error after {_MAX_ATTEMPTS} attempts: {msg}")
        print(f"  openEditor attempt {attempt} failed ({msg}), retrying in 5s...")
        time.sleep(5)

    url_m = (
        re.search(r"<url>(.*?)</url>", xml, re.DOTALL)
        or re.search(r"<msg>(https?://.*?)</msg>", xml, re.DOTALL)
    )
    if url_m:
        wb_url = url_m.group(1).strip().replace("&amp;", "&")
    elif r.headers.get("Location"):
        wb_url = r.headers["Location"]
    else:
        raise RuntimeError(f"Cannot parse openEditor response:\n{xml[:500]}")

    session_token = parse_qs(urlparse(wb_url).query).get("_s", [None])[0]
    if not session_token:
        raise RuntimeError(f"No _s token in workbench URL: {wb_url}")
    return wb_url, session_token


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _find_excel(project_id: str) -> Path:
    """Find *_revised_translation_checks*.xlsx in the project pre-processing folder."""
    pre = _find_pre_folder(project_id)
    matches = list(pre.glob("*_revised_translation_checks*.xlsx"))
    if not matches:
        raise RuntimeError(
            f"No *_revised_translation_checks*.xlsx found in {pre}\n"
            f"Run the archive step first."
        )
    if len(matches) > 1:
        print(f"  Warning: multiple Excel files found, using {matches[0].name}")
    return matches[0]


def _read_translations(path: Path) -> list[tuple[int, str]]:
    """Return list of (unitId, translation_text) from Column A and C.

    Rows 1–3 are headers and are skipped. Empty Column C cells are kept as
    empty strings so the position mapping stays intact.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Validate that Column C is the translation column by checking the header rows
    header_values = [
        str(ws.cell(row=r, column=3).value or "").strip()
        for r in range(1, 4)
    ]
    MARKERS = ("target", "german", "de")
    if not any(
        marker in cell.lower()
        for cell in header_values
        for marker in MARKERS
    ):
        raise ValueError(
            f"Column C in {path.name} does not look like a translation column.\n"
            f"  Header rows (C1–C3): {header_values}\n"
            f"  Expected one of {MARKERS} (case-insensitive) in the first 3 rows."
        )

    segments: list[tuple[int, str]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        unit_id = row[0]
        text = row[2]  # Column C
        if unit_id is None:
            continue
        segments.append((int(unit_id), str(text).strip() if text is not None else ""))
    return segments


def _write_results_to_excel(path: Path, results: dict[int, str]) -> None:
    """Append an 'Upload Status' column to the Excel file with per-segment outcomes."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Reuse existing Upload Status column if present, otherwise append one
    status_col = ws.max_column + 1
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=2, column=col).value or "").strip().lower() == "upload status":
            status_col = col
            break

    ws.cell(row=2, column=status_col).value = "Upload Status"

    for row in range(4, ws.max_row + 1):
        unit_id = ws.cell(row=row, column=1).value
        if unit_id is None:
            continue
        uid = int(unit_id)
        if uid in results:
            ws.cell(row=row, column=status_col).value = results[uid]

    wb.save(path)
    print(f"  Results written to column {status_col} of {path.name}")


# ---------------------------------------------------------------------------
# STOMP helpers
# ---------------------------------------------------------------------------

def _ts() -> str:
    """Current epoch milliseconds as a string, used as requestId."""
    return str(int(time.time() * 1000))


def _make_send_frame(destination: str, session_token: str, body: dict) -> str:
    """Wrap a STOMP SEND frame in the SockJS array envelope."""
    body_str = json.dumps(body, ensure_ascii=False)
    body_bytes = body_str.encode("utf-8")
    frame = (
        f"SEND\ndestination:{destination}\n"
        f"_s:{session_token}\ncontent-length:{len(body_bytes)}\n\n"
        f"{body_str}\x00"
    )
    return json.dumps([frame])


def _parse_stomp_messages(raw: str) -> list[dict]:
    """Extract parsed JSON payloads from a SockJS 'a[...]' frame."""
    if not raw.startswith("a"):
        return []
    results = []
    for frame in json.loads(raw[1:]):
        sep = frame.rfind("\n\n")
        if sep == -1:
            continue
        body_str = frame[sep + 2:].rstrip("\x00")
        if not body_str:
            continue
        try:
            msg = json.loads(body_str)
        except json.JSONDecodeError:
            continue
        results.append(msg)
    return results


# ---------------------------------------------------------------------------
# Tag-aware target node builder
# ---------------------------------------------------------------------------

def _clean_source_nodes(nodes: list[dict]) -> list[dict]:
    """Remove fuzzy-match diff markers from a source node list.

    matchesInfo.matches[n].source.nodes contains a DELETION/INSERTION diff when the
    match is a fuzzy TM hit.  DELETION nodes belong to the old TM entry, not the
    current segment — they must be removed.  INSERTION nodes are new in the current
    segment — keep them but strip the diff decoration so downstream code sees clean nodes.
    """
    result = []
    for node in nodes:
        decs = node.get("decorations", [])
        if any(d.get("type") == "DELETION" for d in decs):
            continue
        clean = dict(node)
        clean["decorations"] = [d for d in decs if d.get("type") not in ("DELETION", "INSERTION")]
        result.append(clean)
    return result



def _anchor_numeric_groups(
    text: str,
    groups: list[tuple[str, dict, str, dict]],
) -> list[dict]:
    """Insert (open_inline, digit_text, close_inline) triples at anchor positions in *text*.

    Each group carries left_ctx - the last 4 chars of the preceding source TEXT - used to
    find the right occurrence when the same digit appears multiple times in the sentence
    (e.g. two <g>3</g> for m³ and ft³).  Falls back to plain left-to-right substring search
    when the English context string is not present in the German translation.
    """
    result: list[dict] = []
    remaining = text
    for left_ctx, open_n, digit, close_n in groups:
        placed = False
        if left_ctx and left_ctx in remaining:
            ctx_end = remaining.index(left_ctx) + len(left_ctx)
            after_ctx = remaining[ctx_end:].lstrip()
            if after_ctx.startswith(digit):
                print(f"    [TAGS] anchored {digit!r} via left_ctx {left_ctx!r}")
                result.append({"type": "TEXT", "decorations": [], "content": remaining[:ctx_end]})
                result.append(open_n)
                result.append({"type": "TEXT", "decorations": [], "content": digit})
                result.append(close_n)
                remaining = after_ctx[len(digit):]
                placed = True
        if not placed and digit in remaining:
            idx = remaining.index(digit)
            print(f"    [TAGS] anchored {digit!r} via fallback substring at pos {idx}")
            result.append({"type": "TEXT", "decorations": [], "content": remaining[:idx]})
            result.append(open_n)
            result.append({"type": "TEXT", "decorations": [], "content": digit})
            result.append(close_n)
            remaining = remaining[idx + len(digit):]
            placed = True
        if not placed:
            print(f"    [TAGS] WARNING: could not place group digit={digit!r} - not found in text")
    if remaining:
        result.append({"type": "TEXT", "decorations": [], "content": remaining})
    return result


def _build_target_nodes(source_nodes: list[dict], excel_text: str) -> list[dict]:
    """Build target nodes preserving INLINE tags from the source in correct positions.

    Source nodes are first grouped into chunks: plain TEXT nodes and inline-groups
    (INLINE_open + TEXT + INLINE_close with the same inlineId).

    Inline-groups whose sandwiched TEXT is purely numeric AND that have TEXT chunks
    on both sides are anchored inside the translation using the tail of the preceding
    source TEXT as a locator (left_ctx).  All other inline-groups pile before the
    translation text (boundary pattern, e.g. paragraph numbers).
    """
    if not any(n.get("type") == "INLINE" for n in source_nodes):
        return [{"type": "TEXT", "decorations": [], "content": excel_text}]

    # --- Pass 1: parse source into chunks ---
    # chunk: ("text", node) | ("group", open_n, txt_n, close_n) | ("lone", node)
    chunks: list[tuple] = []
    i = 0
    while i < len(source_nodes):
        node = source_nodes[i]
        if node.get("type") == "TEXT":
            chunks.append(("text", node))
            i += 1
        elif node.get("type") == "INLINE":
            if (i + 2 < len(source_nodes)
                    and source_nodes[i + 1].get("type") == "TEXT"
                    and source_nodes[i + 2].get("type") == "INLINE"
                    and source_nodes[i + 2].get("inlineId") == node.get("inlineId")):
                chunks.append(("group", node, source_nodes[i + 1], source_nodes[i + 2]))
                i += 3
            else:
                chunks.append(("lone", node))
                i += 1
        else:
            i += 1

    # --- Pass 2: classify non-text chunks ---
    text_chunk_indices = [j for j, c in enumerate(chunks) if c[0] == "text"]
    last_text_ci = text_chunk_indices[-1] if text_chunk_indices else -1

    prefix = ""
    for c in chunks:
        if c[0] != "text":
            break
        prefix += c[1].get("content", "")

    numeric_mid: list[tuple[str, dict, str, dict]] = []
    boundary_pre: list[dict] = []
    boundary_post: list[dict] = []

    for j, chunk in enumerate(chunks):
        if chunk[0] == "text":
            continue
        after_last_text = (j > last_text_ci)

        if chunk[0] == "group":
            _, open_n, txt_n, close_n = chunk
            digit = txt_n.get("content", "")
            has_text_before = any(chunks[k][0] == "text" for k in range(j))
            has_text_after  = any(chunks[k][0] == "text" for k in range(j + 1, last_text_ci + 1))
            if has_text_before and has_text_after and digit.strip().isdigit():
                left_ctx = ""
                for prev in reversed(chunks[:j]):
                    if prev[0] == "text":
                        left_ctx = prev[1].get("content", "").rstrip()[-4:]
                        break
                print(f"    [TAGS] numeric group inlineId={open_n.get('inlineId')} digit={digit!r} left_ctx={left_ctx!r}")
                numeric_mid.append((left_ctx, {**open_n}, digit, {**close_n}))
            elif after_last_text:
                boundary_post.extend([{**open_n}, {**txt_n}, {**close_n}])
            else:
                boundary_pre.extend([{**open_n}, {**txt_n}, {**close_n}])
        elif chunk[0] == "lone":
            lone_n = chunk[1]
            if after_last_text:
                boundary_post.append({**lone_n})
            else:
                boundary_pre.append({**lone_n})

    if not boundary_pre and not numeric_mid and not boundary_post:
        return [{"type": "TEXT", "decorations": [], "content": excel_text}]

    # --- Pass 3: strip paragraph-number prefix from the Excel text ---
    # Source nodes may use NBSP ( ) inside the bracket while the Excel text
    # uses a regular space or no space - build a whitespace-flexible regex so the
    # comparison succeeds regardless of which Unicode space variant is present.
    text = excel_text
    prefix_stripped = prefix.strip()
    _prefix_matched = False
    if prefix_stripped:
        _ws_re = re.compile(r'[\s  ]+')
        _parts = [re.escape(p) for p in _ws_re.split(prefix_stripped) if p]
        if _parts:
            _prefix_pat = re.compile(r'[\s  ]*'.join(_parts))
            _m = _prefix_pat.match(text)
            if _m:
                text = text[_m.end():].lstrip()
                _prefix_matched = True

    # --- Pass 4: build translation body ---
    body: list[dict] = (
        _anchor_numeric_groups(text, numeric_mid) if numeric_mid
        else [{"type": "TEXT", "decorations": [], "content": text}]
    )
    if boundary_pre and body and body[0].get("type") == "TEXT":
        body[0] = {**body[0], "content": " " + body[0]["content"]}

    nodes: list[dict] = []
    if prefix_stripped and _prefix_matched:
        nodes.append({"type": "TEXT", "decorations": [], "content": prefix_stripped})
    nodes.extend(boundary_pre)
    nodes.extend(body)
    nodes.extend(boundary_post)
    return nodes


# ---------------------------------------------------------------------------
# Upload loop
# ---------------------------------------------------------------------------

def _upload_via_stomp(
    session: requests.Session,
    session_token: str,
    csrf_token: str,
    segments: list[tuple[int, str]],
) -> dict[int, str]:
    """Connect to the workbench WebSocket and upload all translations."""
    server_id = str(random.randint(0, 999)).zfill(3)
    session_id = "".join(random.choices(string.ascii_lowercase + string.digits, k=8))
    ws_url = (
        f"wss://word.welocalize.com/workbench/ws/{server_id}/{session_id}"
        f"/websocket?_s={session_token}"
    )
    cookie_str = "; ".join(f"{c.name}={c.value}" for c in session.cookies)

    ws = _websocket.WebSocket()
    ws.connect(ws_url, cookie=cookie_str)
    ws.settimeout(5)  # short per-recv timeout so wait_for checks deadline frequently

    last_keepalive = time.time()
    tu_updates: dict[int, dict] = {}  # unit_id → TRANS_UNIT_UPDATED payload cache

    def send(destination: str, body: dict) -> None:
        ws.send(_make_send_frame(destination, session_token, body))

    def maybe_keepalive() -> None:
        nonlocal last_keepalive
        if time.time() - last_keepalive >= KEEPALIVE_INTERVAL:
            _keepalive(session)
            last_keepalive = time.time()

    def wait_for(msg_type: str, timeout: float = 30.0, unit_id: int | None = None) -> dict:
        """Recv WebSocket frames until a matching message arrives; return its payload.

        All TRANS_UNIT_UPDATED messages seen along the way are cached in tu_updates.
        If unit_id is given, only a message whose payload.unitId matches will satisfy
        the wait (useful for TRANS_UNIT_UPDATED to avoid stopping on a stale broadcast
        for the segment we just saved rather than the one we just activated).
        """
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                raw = ws.recv()
            except _websocket.WebSocketTimeoutException:
                continue  # normal: no data in this 5 s window, check deadline and retry
            except (
                _websocket.WebSocketConnectionClosedException,
                _websocket.WebSocketProtocolException,
                OSError,
            ) as exc:
                raise RuntimeError(f"WebSocket connection lost while waiting for {msg_type}: {exc}") from exc
            if not raw or raw == "h":
                continue
            for msg in _parse_stomp_messages(raw):
                if msg.get("type") == "TRANS_UNIT_UPDATED":
                    p = msg.get("payload", {})
                    uid = p.get("unitId", p.get("id"))
                    if uid is not None:
                        tu_updates[uid] = p
                if msg.get("type") == msg_type:
                    p = msg.get("payload", {})
                    if unit_id is None or p.get("unitId", p.get("id")) == unit_id:
                        return p
        raise TimeoutError(f"Timed out after {timeout}s waiting for {msg_type} (unit_id={unit_id})")

    def _stomp_handshake() -> None:
        """Perform SockJS open + STOMP CONNECT + SUBSCRIBE on the current ws."""
        ws.recv()  # SockJS open frame 'o'
        connect_frame = (
            f"CONNECT\nX-CSRF-TOKEN:{csrf_token}\n"
            f"accept-version:1.0,1.1,1.2\nheart-beat:10000,10000\n\n\x00"
        )
        ws.send(json.dumps([connect_frame]))
        ws.recv()  # CONNECTED
        ws.send(json.dumps(["SUBSCRIBE\nid:sub-0\ndestination:/user/queue/main\n\n\x00"]))


    try:
        _stomp_handshake()

        # --- Init sequence ---
        send("/workbench/readChatInfo", {"requestId": _ts()})
        send("/workbench/document/init/rendered", {"requestId": _ts()})

        # --- Wait for first TRANS_UNIT_UPDATED; retry up to 3 times if server is slow ---
        _MAX_INIT_ATTEMPTS = 3
        _INIT_TIMEOUT = 15  # seconds per attempt
        for attempt in range(1, _MAX_INIT_ATTEMPTS + 1):
            send("/workbench/trans-unit/activate", {
                "requestId": _ts(),
                "activatedTransUnitId": segments[0][0],
                "forceTransUnitsUpdate": attempt > 1,  # force resend on retry
            })
            try:
                payload = wait_for("TRANS_UNIT_UPDATED", timeout=_INIT_TIMEOUT)
                tu_updates[segments[0][0]] = payload  # cache init segment
                break
            except TimeoutError:
                if attempt == _MAX_INIT_ATTEMPTS:
                    raise TimeoutError(
                        f"Server did not send TRANS_UNIT_UPDATED after "
                        f"{_MAX_INIT_ATTEMPTS} attempts ({_INIT_TIMEOUT}s each)"
                    )
                print(f"  No response from server (attempt {attempt}/{_MAX_INIT_ATTEMPTS}), retrying...")
        print(f"  Document: {payload.get('workbenchDocumentId', '?')}")
        # Print first segment state to help diagnose save errors
        debug_keys = ["status", "locked", "segmentType", "actorType", "workflowStepType",
                      "canEdit", "editable", "isLocked"]
        debug_info = {k: payload[k] for k in debug_keys if k in payload}
        if not debug_info:
            # Fall back to the full payload (truncated) so nothing is hidden
            debug_info = str(payload)[:600]
        print(f"  First segment state: {json.dumps(debug_info) if isinstance(debug_info, dict) else debug_info}")

        # --- Per-segment upload ---
        results: dict[int, str] = {}
        last_id = segments[-1][0]
        total = len(segments)
        consecutive_errors = 0
        _MAX_CONSECUTIVE_ERRORS = 2  # abort early if server keeps rejecting every save
        _debug_printed = 0
        _xtm_skipped: set[int] = set()  # uids XTM auto-confirmed (ICE); must not overwrite

        for i, (unit_id, text) in enumerate(segments):

            is_last = (i == total - 1)
            next_uid = segments[i + 1][0] if not is_last else None

            # --- Detect segments XTM already auto-confirmed as ICE matches ---
            if unit_id in _xtm_skipped:
                _xtm_skipped.discard(unit_id)
                results[unit_id] = "confirmed (ICE - auto-detected)"
                print(f"  [{unit_id}/{last_id}] confirmed (ICE - XTM skipped over it, not overwriting)")
                maybe_keepalive()
                continue

            # --- Skip empty segments ---
            if not text:
                results[unit_id] = "skipped"
                print(f"  [{unit_id}/{last_id}] skipped (empty)")
                if next_uid is not None:
                    send("/workbench/trans-unit/activate", {
                        "requestId": _ts(),
                        "activatedTransUnitId": next_uid,
                        "forceTransUnitsUpdate": True,
                        "deactivatedTransUnitId": unit_id,
                    })
                    try:
                        wait_for("TRANS_UNIT_UPDATED", timeout=15, unit_id=next_uid)
                    except TimeoutError:
                        print(f"  [{unit_id}/{last_id}] Warning: no TRANS_UNIT_UPDATED for segment {next_uid} after skip")
                maybe_keepalive()
                continue

            word_count = len(text.split())
            preview = text[:70] + ("…" if len(text) > 70 else "")

            try:
                # --- Classify match type (ICE / 100% / repetition / fuzzy) ---
                tu_payload = tu_updates.get(unit_id, {})
                _best_match = (tu_payload.get("matchesInfo", {}).get("matches") or [{}])[0]
                # `or ""` normalises JSON null → "" so the fuzzy guard works even
                # when XTM sends "matchQuality": null for untranslated repetitions.
                _match_quality = _best_match.get("matchQuality") or ""
                # A non-empty quality that is not "100%" means a genuine fuzzy match.
                # Guard every auto-confirm check so fuzzy quality always wins.
                _is_fuzzy = bool(_match_quality) and _match_quality != "100%"
                _is_ice = _best_match.get("iceMatch", False) and not _is_fuzzy
                _is_rep = (
                    _best_match.get("repetitionType") == "INTERNAL" and not _is_fuzzy
                )
                _is_100 = (
                    _match_quality == "100%"
                    and _best_match.get("matchType") != "MACHINE_TRANSLATION"
                )
                auto_confirm_label = (
                    "ICE" if _is_ice else
                    "repetition" if _is_rep else
                    "100%" if _is_100 else
                    None
                )

                # --- Build target nodes ---
                # ICE / 100% / repetition → use XTM's pre-built TM target directly.
                # Fuzzy or no match       → reconstruct target from source tags + Excel text.
                # If the TM target is empty (untranslated repetition backed by a fuzzy
                # TM hit), fall through to the Excel path rather than saving empty nodes.
                _use_tm = AUTO_CONFIRM_MATCHES and auto_confirm_label
                if _use_tm:
                    target_nodes = _best_match.get("target", {}).get("nodes", [])
                    if not target_nodes:
                        _use_tm = False
                        auto_confirm_label = None

                if _use_tm:
                    have_source = True
                    source_nodes = []
                else:
                    auto_confirm_label = None
                    source_nodes = tu_payload.get("source", {}).get("nodes", [])
                    if not source_nodes:
                        _matches = tu_payload.get("matchesInfo", {}).get("matches", [])
                        if _matches:
                            source_nodes = _clean_source_nodes(
                                _matches[0].get("source", {}).get("nodes", [])
                            )
                    have_source = bool(source_nodes)
                    target_nodes = _build_target_nodes(source_nodes, text)

                # --- Debug dump (controlled by DEBUG_SOURCE_NODES_LIMIT) ---
                if DEBUG_SOURCE_NODES_LIMIT and _debug_printed < DEBUG_SOURCE_NODES_LIMIT:
                    raw_tu = tu_updates.get(unit_id)
                    print(f"\n  [DEBUG unit {unit_id}] tu_updates entry: {json.dumps(raw_tu, ensure_ascii=False) if raw_tu is not None else 'NOT IN CACHE'}")
                    print(f"  [DEBUG unit {unit_id}] target nodes (sending): {json.dumps(target_nodes, ensure_ascii=False)}\n")
                    _debug_printed += 1
                elif not have_source:
                    print(f"  [{unit_id}/{last_id}] WARNING: source nodes not received — saving as plain text (tags may be lost)")

                # --- Send save-unit + activate next ---
                ts_ms = int(time.time() * 1000)
                send("/workbench/save-unit", {
                    "requestId": str(ts_ms),
                    "units": [{
                        "target": {"nodes": target_nodes},
                        "status": "DONE",
                        "statusOrigin": "M",
                        "timeTracking": {"timeSpent": min(20_000, max(2_000, word_count * 500))},
                        "unitId": unit_id,
                    }],
                })
                # XTM processes the save only after seeing activate
                # (mirrors browser: confirm button + click next = hourglass).
                if next_uid is not None:
                    send("/workbench/trans-unit/activate", {
                        "requestId": _ts(),
                        "activatedTransUnitId": next_uid,
                        "forceTransUnitsUpdate": True,
                        "deactivatedTransUnitId": unit_id,
                    })

                # --- Drain WebSocket responses ---
                # Wait for SAVE_RESPONSE (this segment) and TRANS_UNIT_UPDATED (next segment).
                save_resp = None
                next_tu_ready = (next_uid is None)
                _drain_seen_uids: set[int] = set()
                deadline = time.time() + 9.0
                while time.time() < deadline and not (save_resp is not None and next_tu_ready):
                    try:
                        raw = ws.recv()
                    except _websocket.WebSocketTimeoutException:
                        continue
                    except (
                        _websocket.WebSocketConnectionClosedException,
                        _websocket.WebSocketProtocolException,
                        OSError,
                    ) as exc:
                        raise RuntimeError(f"WebSocket connection lost: {exc}") from exc
                    if not raw or raw == "h":
                        continue
                    for msg in _parse_stomp_messages(raw):
                        mtype = msg.get("type")
                        p = msg.get("payload", {})
                        if mtype == "TRANS_UNIT_UPDATED":
                            uid = p.get("unitId", p.get("id"))
                            print(f"    [DRAIN] TU_UPDATED uid={uid} (want {next_uid}){' ←' if uid == next_uid else ''}")
                            if uid is not None:
                                tu_updates[uid] = p
                                _drain_seen_uids.add(uid)
                            if uid == next_uid:
                                next_tu_ready = True
                        elif mtype == "SAVE_RESPONSE":
                            res_type = p.get("result", {}).get("type", "?")
                            print(f"    [DRAIN] SAVE_RESPONSE result={res_type}")
                            save_resp = p
                        elif mtype and mtype != "PROGRESS_UPDATE":
                            print(f"    [DRAIN] other: {mtype}")

                if not next_tu_ready:
                    if next_uid is not None and any(uid > next_uid for uid in _drain_seen_uids):
                        # XTM advanced past next_uid — it auto-confirmed it as an ICE match.
                        # Saving Excel text into it would overwrite the correct auto-fill.
                        _xtm_skipped.add(next_uid)
                        print(f"  [{unit_id}/{last_id}] XTM auto-confirmed segment {next_uid} (ICE), will skip")
                    else:
                        print(f"  [{unit_id}/{last_id}] Warning: no TRANS_UNIT_UPDATED for segment {next_uid}, tags may be missing")

                # --- Record result ---
                result_type = save_resp.get("result", {}).get("type", "UNKNOWN") if save_resp else "TIMEOUT"
                if result_type != "SUCCESS":
                    reason = save_resp.get("result", {}).get("message", result_type) if save_resp else "timeout after 30s"
                    results[unit_id] = f"failed: {reason}"
                    consecutive_errors += 1
                    print(f"  [{unit_id}/{last_id}] FAILED: {reason}  ({word_count}w)")
                    print(f"    DEBUG full SAVE_RESPONSE: {json.dumps(save_resp, ensure_ascii=False)}")
                    if consecutive_errors >= _MAX_CONSECUTIVE_ERRORS:
                        print(
                            f"\n  Aborting: {consecutive_errors} consecutive save errors — "
                            "server is rejecting all writes.  Check step type and permissions.\n"
                            "  Mark remaining segments as 'not attempted'."
                        )
                        for uid, _ in segments[i + 1:]:
                            results[uid] = "not attempted"
                        break
                else:
                    consecutive_errors = 0
                    if auto_confirm_label:
                        results[unit_id] = f"confirmed ({auto_confirm_label})"
                        print(f"  [{unit_id}/{last_id}] confirmed ({auto_confirm_label})  ({word_count}w)")
                    elif have_source:
                        results[unit_id] = "saved"
                        print(f"  [{unit_id}/{last_id}] {preview}  ({word_count}w)")
                    else:
                        results[unit_id] = "saved (no source nodes — verify manually)"
                        print(f"  [{unit_id}/{last_id}] UNCERTAIN: {preview}  ({word_count}w) — saved without source nodes")

            except RuntimeError as exc:
                results[unit_id] = f"failed: {exc}"
                print(f"  [{unit_id}/{last_id}] FAILED: {exc}")
                for uid, _ in segments[i + 1:]:
                    results[uid] = "not attempted"
                break

            maybe_keepalive()
            time.sleep(4)

        return results

    finally:
        try:
            ws.send(json.dumps(["DISCONNECT\n\n\x00"]))
        except Exception:
            pass
        ws.close()


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def run(project_id: str) -> None:
    """Open XTM Workbench in write mode and upload revised translations."""
    # Fail fast before spending time on XTM login
    excel_path = _find_excel(project_id)

    username, password = _load_creds()

    session = requests.Session()
    session.headers.update({
        "Accept": "application/json, text/plain, */*",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:149.0) "
            "Gecko/20100101 Firefox/149.0"
        ),
    })

    print("Step 1 — Login to XTM Workbench...")
    uust = _login(session, username, password)
    session.headers.update({"uust": uust, "X-Requested-With": "XMLHttpRequest"})

    print("Step 2 — Fetching task list...")
    tasks = _get_tasks(session)
    print(f"  {len(tasks)} task(s) in progress")

    task = _find_task(tasks, project_id, file_filter=FILE_FILTER)
    ad = task["additionalData"]
    print(f"  Found: {ad.get('projectName', '?')}  (file {ad.get('fileId', '?')})")
    print(f"  Step: {task.get('STEP', '?')}  type: {task.get('STEP_TYPE', '?')}  role: {task.get('ROLE', '?')}  actorType: {ad.get('actorType', '?')}")

    print("Step 3 — Claiming task for INTERNALLINGUIST (if needed)...")
    task = _claim_group_task(session, task, uust, project_id)
    actor_after = task["additionalData"].get("actorType", "?")
    if actor_after != "INTERNALLINGUIST":
        raise RuntimeError(
            f"Task actor is still '{actor_after}' — automatic claim failed.\n"
            "  Please accept the task manually in XTM (open the project → accept/claim the step),\n"
            "  then re-run this script."
        )

    print("Step 5 — Reading translations from Excel...")
    segments = _read_translations(excel_path)
    if START_FROM_SEGMENT_ID > 1:
        segments = [(uid, t) for uid, t in segments if uid >= START_FROM_SEGMENT_ID]
        print(f"  Starting from segment ID {START_FROM_SEGMENT_ID} ({len(segments)} segments remaining)")
    if TEST_SEGMENT_LIMIT is not None:
        segments = segments[:TEST_SEGMENT_LIMIT]
        print(f"  TEST MODE: capped at {TEST_SEGMENT_LIMIT} segments")
    non_empty = sum(1 for _, t in segments if t)
    print(f"  {len(segments)} segments read ({non_empty} non-empty) from {excel_path.name}")

    def _hard_login():
        s = requests.Session()
        s.headers.update({
            "Accept": "application/json, text/plain, */*",
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:149.0) "
                "Gecko/20100101 Firefox/149.0"
            ),
        })
        uu = _login(s, username, password)
        s.headers.update({"uust": uu, "X-Requested-With": "XMLHttpRequest"})
        tsk = _find_task(_get_tasks(s), project_id, file_filter=FILE_FILTER)
        tsk = _claim_group_task(s, tsk, uu, project_id)
        if tsk["additionalData"].get("actorType") != "INTERNALLINGUIST":
            raise RuntimeError(
                f"Task actor is still '{tsk['additionalData'].get('actorType')}' — "
                "accept the task manually in XTM, then re-run."
            )
        url, tok = _open_editor_write(s, tsk, uu)
        csrf = _init_workbench(s, url, tok)
        time.sleep(3)
        _keepalive(s)
        return s, tsk, uu, tok, csrf

    def _reopen_editor(s, tsk, uu):
        url, tok = _open_editor_write(s, tsk, uu)
        csrf = _init_workbench(s, url, tok)
        time.sleep(3)
        _keepalive(s)
        return tok, csrf

    print("\nSteps 4 & 6 — Login, claim task, open editor, upload...")
    session, task, uust, session_token, csrf_token = _hard_login()
    print(f"  Session token: {session_token[:12]}...")

    todo          = list(segments)
    results: dict[int, str] = {}
    retry_counts: dict[int, int] = {}
    batch_num     = 0

    while todo:
        batch    = todo[:UPLOAD_BATCH_SIZE]
        first_id = batch[0][0]
        last_id  = batch[-1][0]
        batch_num += 1
        print(f"\nBatch {batch_num} — segments {first_id}–{last_id} ({len(batch)} segments)...")

        batch_results = _upload_via_stomp(session, session_token, csrf_token, batch)

        retry_ids = {uid for uid, st in batch_results.items()
                     if st == "not attempted" or st.startswith("failed")}
        for uid, st in batch_results.items():
            if uid not in retry_ids:
                results[uid] = st

        if retry_ids:
            permanent: set[int] = set()
            for uid in retry_ids:
                retry_counts[uid] = retry_counts.get(uid, 0) + 1
                if retry_counts[uid] >= 2:
                    permanent.add(uid)
                    results[uid] = "skipped (permanent server rejection — check manually in XTM)"
                    print(f"  Segment {uid}: giving up after {retry_counts[uid]} attempts.")

            retry_ids -= permanent
            if not retry_ids:
                todo = todo[UPLOAD_BATCH_SIZE:]
                continue

            print(f"  {len(retry_ids)} segment(s) failed/not-attempted — "
                  f"hard re-login in {BATCH_WAIT_SECONDS}s, resuming from segment {min(retry_ids)}...")
            time.sleep(BATCH_WAIT_SECONDS)
            session, task, uust, session_token, csrf_token = _hard_login()
            print(f"  New session token: {session_token[:12]}...")
            retry_map = {uid: txt for uid, txt in segments if uid in retry_ids}
            todo = [(uid, retry_map[uid]) for uid in sorted(retry_ids)] + todo[UPLOAD_BATCH_SIZE:]
        else:
            todo = todo[UPLOAD_BATCH_SIZE:]
            if todo:
                print(f"  Batch done — re-opening editor...")
                try:
                    session_token, csrf_token = _reopen_editor(session, task, uust)
                    print(f"  Session token: {session_token[:12]}...")
                except Exception as exc:
                    print(f"  Editor re-open failed ({exc}) — hard re-login in {BATCH_WAIT_SECONDS}s...")
                    time.sleep(BATCH_WAIT_SECONDS)
                    session, task, uust, session_token, csrf_token = _hard_login()
                    print(f"  New session token: {session_token[:12]}...")

    print("\nStep 7 — Writing results to Excel...")
    _write_results_to_excel(excel_path, results)

    saved         = sum(1 for s in results.values() if s == "saved")
    confirmed     = sum(1 for s in results.values() if s.startswith("confirmed"))
    skipped       = sum(1 for s in results.values() if s == "skipped")
    not_attempted = sum(1 for s in results.values() if s == "not attempted")
    failed        = {uid: s for uid, s in results.items() if s.startswith("failed")}

    print()
    print("=== Upload summary ===")
    print(f"  Saved:         {saved}")
    print(f"  Confirmed:     {confirmed}  (ICE / 100% / repetition)")
    print(f"  Skipped:       {skipped}  (empty translation)")
    if not_attempted:
        print(f"  Not attempted: {not_attempted}  (connection lost before attempt)")
    print(f"  Failed:        {len(failed)}")
    if failed:
        for uid, reason in sorted(failed.items()):
            print(f"    segment {uid}: {reason}")
    print(f"  Total:         {len(results)}")


def main() -> None:
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("project_id")
    ap.add_argument(
        "--file", default=None,
        help="Filename substring to select the right task when a project has multiple files, e.g. 'Anmeldefassung'.",
    )
    args = ap.parse_args()
    if args.file:
        global FILE_FILTER
        FILE_FILTER = args.file
    try:
        run(args.project_id)
    except (RuntimeError, TimeoutError) as e:
        print(f"\nError: {e}")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
