# ============================================================
# ipappify_translate.py
# ============================================================
# Calls the IP.appify translation API directly, segment by segment,
# replicating the context window the Word plugin sends.
#
# SETUP
#   1. Run mitmproxy while the Word plugin is open.
#   2. Copy the Bearer token from any captured request to
#      iptranslator.ipappify.de into .env as IPAPPIFY_TOKEN=<token>
#   3. The token is valid for ~1 hour. The script warns you when it expires.
#
# INPUT   projects/<project_id>/<any>.xlsx   — bilingual Excel from XTM
# OUTPUT  projects/<project_id>/<name>_translated.xlsx
# ============================================================

import os
import glob
import json
import base64
import struct
import time
import uuid
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import pandas as pd
import openpyxl
from datetime import datetime, timezone
from dotenv import load_dotenv

from project_log import project_dir as _pdir

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

API_URL    = "https://iptranslator.ipappify.de/translate"
MODEL      = "EN->DE v03"
SRC_LANG   = "en"
TGT_LANG   = "de"
CONTEXT_N  = 1   # segments of past/future context to include

# ============================================================
# Token management — auto-refresh via Azure B2C refresh token
# ============================================================

TOKEN_URL   = "https://ipappifyusers.b2clogin.com/tfp/ipappifyusers.onmicrosoft.com/B2C_1_susi2/oauth2/v2.0/token"
CLIENT_ID   = "4fb07bed-e217-4929-b283-f19c0c02621a"
SCOPE       = ("https://ipappifyusers.onmicrosoft.com/nmt-service-api/nmt.request "
               "https://ipappifyusers.onmicrosoft.com/nmt-service-api/user_impersonation "
               "offline_access openid profile")
ENV_PATH    = os.path.join(os.path.dirname(__file__), ".env")


def decode_exp(jwt_token):
    try:
        payload_b64 = jwt_token.split(".")[1]
        payload_b64 += "=" * (-len(payload_b64) % 4)
        payload = json.loads(base64.b64decode(payload_b64))
        return payload.get("exp", 0)
    except Exception:
        return 0


def refresh_access_token(refresh_token):
    resp = requests.post(
        TOKEN_URL,
        data={
            "client_id":     CLIENT_ID,
            "scope":         SCOPE,
            "client_info":   "1",
            "grant_type":    "refresh_token",
            "refresh_token": refresh_token,
        },
        verify=False,
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def update_env(key, value):
    """Update or insert a key=value line in the .env file."""
    lines = []
    found = False
    if os.path.exists(ENV_PATH):
        with open(ENV_PATH, "r", encoding="utf-8") as f:
            lines = f.readlines()
    new_lines = []
    for line in lines:
        if line.startswith(f"{key}="):
            new_lines.append(f"{key}={value}\n")
            found = True
        else:
            new_lines.append(line)
    if not found:
        new_lines.append(f"{key}={value}\n")
    with open(ENV_PATH, "w", encoding="utf-8") as f:
        f.writelines(new_lines)


# Try current access token first, refresh if expired or missing
token = os.environ.get("IPAPPIFY_TOKEN", "").strip()
refresh_token = os.environ.get("IPAPPIFY_REFRESH_TOKEN", "").strip()

if not refresh_token:
    print("ERROR: IPAPPIFY_REFRESH_TOKEN not set in .env.")
    print("  Capture it from mitmproxy: look for a POST to ipappifyusers.b2clogin.com")
    print("  and copy the refresh_token field from the request body.")
    exit()

exp = decode_exp(token) if token else 0
remaining = exp - time.time()

if remaining > 60:
    print(f"Token valid for {int(remaining // 60)} min {int(remaining % 60)} sec.")
else:
    print("Access token expired or missing — refreshing...")
    try:
        token_data = refresh_access_token(refresh_token)
        token = token_data["access_token"]
        new_refresh = token_data.get("refresh_token", refresh_token)
        update_env("IPAPPIFY_TOKEN", token)
        update_env("IPAPPIFY_REFRESH_TOKEN", new_refresh)
        remaining = decode_exp(token) - time.time()
        print(f"Token refreshed. Valid for {int(remaining // 60)} min {int(remaining % 60)} sec.")
    except Exception as e:
        print(f"ERROR: Could not refresh token — {e}")
        print("  Open Word with the IP.appify plugin to get a fresh session,")
        print("  then capture a new refresh_token from mitmproxy.")
        exit()

# ============================================================
# Load Excel
# ============================================================

proj_dir = _pdir()
xlsx_files = glob.glob(str(proj_dir / "*.xlsx"))
xlsx_files = [f for f in xlsx_files if "_translated" not in f and "_checks" not in f and not os.path.basename(f).startswith("~$")]
if not xlsx_files:
    print(f"ERROR: No .xlsx file found in '{proj_dir}'.")
    exit()
if len(xlsx_files) > 1:
    print(f"Multiple .xlsx files found, using: {xlsx_files[0]}")
input_path = xlsx_files[0]

raw_df = pd.read_excel(input_path, header=None, engine="openpyxl")
print(f"Processing: {raw_df.iloc[0, 0]}")

data_df = raw_df.iloc[3:].reset_index(drop=True)
data_df.columns = ["ID", "Source", "Target"] + list(data_df.columns[3:])
data_df = data_df[["ID", "Source", "Target"]].copy()
data_df.dropna(subset=["Source"], inplace=True)
data_df["ID"]     = data_df["ID"].astype(str).str.strip()
data_df["Source"] = data_df["Source"].astype(str).str.strip()
data_df["Target"] = data_df["Target"].fillna("").astype(str).str.strip()
segments = data_df.to_dict("records")
print(f"Loaded {len(segments)} segments.")

# ============================================================
# Output workbook — load for incremental writes; resume if exists
# ============================================================

out_path = input_path.replace(".xlsx", "_translated.xlsx")

resuming = os.path.exists(out_path)
if resuming:
    wb = openpyxl.load_workbook(out_path)
    print(f"Resuming from existing output: {os.path.basename(out_path)}")
else:
    wb = openpyxl.load_workbook(input_path)

ws = wb.active

# {seg_id: row_num} for fast cell lookup during incremental saves
row_map: dict[str, int] = {}
for row_num in range(4, ws.max_row + 1):
    seg_id = str(ws.cell(row=row_num, column=1).value).strip()
    row_map[seg_id] = row_num

# ============================================================
# Document session identifiers (generated client-side)
# ============================================================

doc_id  = str(uuid.uuid4())
doc_key = base64.b64encode(os.urandom(32)).decode()
doc_name = os.path.basename(input_path)

# ============================================================
# Translate segments
# ============================================================

headers = {
    "Authorization":          f"Bearer {token}",
    "Content-Type":           "text/json; charset=utf-8",
    "Connection":             "Keep-Alive",
    "Upgrade-Insecure-Requests": "1",
    "Cache-Control":          "max-age=0",
    "Host":                   "iptranslator.ipappify.de",
}

translations      = {}  # {segment_id: translated_text}
errors            = []
consecutive_nulls = 0
NULL_ABORT        = 10  # abort if this many null responses in a row

# Pre-populate translations from existing output (resume support).
# Only when loading from _translated.xlsx — the source file's column 3
# contains the original XTM translations and must not be treated as done.
if resuming:
    for seg_id, row_num in row_map.items():
        val = ws.cell(row=row_num, column=3).value
        if val and str(val).strip():
            translations[seg_id] = str(val).strip()
    if translations:
        print(f"Resuming: {len(translations)} segments already translated, skipping.")

print(f"\nTranslating {len(segments)} segments...")

for i, seg in enumerate(segments):
    if seg["ID"] in translations:
        continue
    past   = segments[max(0, i - CONTEXT_N):i]
    future = segments[i + 1:i + 1 + CONTEXT_N]

    body = {
        "CorrelationId": str(uuid.uuid4()),
        "Client":        "IPTranslator.Client",
        "ClientVersion": "2.0.31.0",
        "Translate": {
            "Options": {
                "CompletionUnit":       1,
                "BeamWidth":            4,
                "TopK":                 1,
                "MaxWordLength":        8,
                "NoAlign":              False,
                "DecoderOptions":       [],
                "ContrastiveAlpha":     0.0,
                "BackTranslationLambda":0.0,
                "Dictionary":           [],
                "DictionaryReward":     0.0,
                "ShortLength":          5,
            },
            "RequestId":   str(uuid.uuid4()),
            "DocumentRef": {
                "Id":        doc_id,
                "Key":       doc_key,
                "Name":      doc_name,
                "CustomRef": None,
                "Created":   datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.0000000+00:00"),
            },
            "Model":          MODEL,
            "SourceLanguage": SRC_LANG,
            "TargetLanguage": TGT_LANG,
            "Source":         seg["Source"],
            "Target":         "",
            "Job": {
                "SourceCurrent": seg["Source"],
                "TargetCurrent": "",
                "SourceSupport": [],
                "TargetSupport": [],
                "SourcePast":    [s["Source"] for s in past],
                "TargetPast":    [translations.get(s["ID"], "") for s in past],
                "SourceFuture":  [s["Source"] for s in future],
            },
        },
        "Align": None, "Evaluate": None, "Embed": None, "Replace": None,
        "Cancel": None, "GenAICheck": None, "Ping": None, "Usage": None,
        "Join": None, "Leave": None, "Update": None, "GetApiKey": None,
    }

    try:
        resp = requests.post(API_URL, headers=headers, data=json.dumps(body).encode("utf-8"), timeout=30, verify=False)
        resp.raise_for_status()
        data = resp.json()
        if data is None:
            consecutive_nulls += 1
            print(f"  [{i+1}/{len(segments)}] {seg['ID']}: null response (#{consecutive_nulls})")
            errors.append(seg["ID"])
            if consecutive_nulls >= NULL_ABORT:
                print(f"\nAborted: {NULL_ABORT} consecutive null responses — token likely expired mid-run.")
                print("Re-run the script to refresh the token and continue.")
                break
            continue
        consecutive_nulls = 0
        trans_list = (data.get("Translate") or {}).get("Translations", [])
        if trans_list:
            translated = trans_list[0].get("TargetText", "").strip()
            translations[seg["ID"]] = translated
            if seg["ID"] in row_map:
                ws.cell(row=row_map[seg["ID"]], column=3).value = translated
                try:
                    wb.save(out_path)
                except PermissionError:
                    print(f"    WARNING: Could not save — close {os.path.basename(out_path)} in Excel.")
            print(f"  [{i+1}/{len(segments)}] {seg['ID']}: {translated[:80]}")
        else:
            print(f"  [{i+1}/{len(segments)}] {seg['ID']}: empty response")
            errors.append(seg["ID"])

    except requests.exceptions.HTTPError as e:
        if resp.status_code == 401:
            print(f"\nERROR 401: Token rejected. Capture a fresh token from mitmproxy and re-run.")
            break
        print(f"  [{i+1}/{len(segments)}] {seg['ID']}: HTTP error — {e}")
        errors.append(seg["ID"])
    except Exception as e:
        print(f"  [{i+1}/{len(segments)}] {seg['ID']}: error — {e}")
        errors.append(seg["ID"])

    time.sleep(1)

print(f"\nDone. {len(translations)} translated, {len(errors)} errors.")
print(f'Output: "{out_path}".')

if errors:
    print(f"\nSegments with errors: {', '.join(errors)}")
