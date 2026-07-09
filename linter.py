"""
TODO: add check for "sowohl ... als auch" in target, wenn there is no "both" in source; comment "translate 'and' with 'und'"
add check for number of colons and semicolons matching between source and target
add check for finding "In einer Ausführung"/ "In Ausführungsformen" in target; comment "Bei einer Ausführungsform"/ "Bei Ausführungsformen" is preferred"
linter.py — Segment-level lint checks on *_revised_translation_checks.xlsx.
add check for ", was" in target when source does not contain "which" ; comment "Relativsatz nur mit 'was' einleiten, wenn 'which' in source, otherwise 'wobei'"
add check for finding "Methode" in target; comment "method = Verfahren"

Reads the active project's *_revised_translation_checks.xlsx, inserts a
"Linter" column as column D (shifting Glossary Checks to column E), and
annotates each segment with any issues found.

Checks:
  german_claim_no_article             — German article (Der/Die/Das/Ein/Eine) at start of claim text
  missing_leading_number              — leading [NNNN] / N.N / N. present in source but absent in target
  different_end_punctuation           — source and target end with different punctuation characters
  source_punctuation_where_none_in_target — source ends with punctuation but target ends with letter/digit
  target_punctuation_where_none_in_source — target ends with punctuation but source ends with letter/digit
  leading_trailing_spaces             — leading or trailing whitespace in the target cell value
  negation_not_transferred            — "not"/"no"/"none" in source but "nicht"/"kein" absent from target
  regular_space_before_unit           — regular space between number and unit should be non-breaking space (Alt+0160)
  hyphen_in_number_range              — hyphen between digits in target should be en dash (Alt+0150)
  in_response_to_mistranslated        — "in Reaktion" in target when source has "in response to" (→ "als Reaktion auf")
  plurality_not_transferred           — "plurality" count in source must match "Vielzahl" count in target
  beide_ambiguous                     — "beide*" in target flagged as ambiguous (zwei / either / both)
  welche_relativpronomen              — "welche*" in target flagged; use der/die/das instead
  werden_dynamic                      — "wird/wurde*/werden/geworden*" flagged; double-check static vs. dynamic
  preposition_contraction             — German im/vom/am/beim/zum/zur contractions in target
  betraegt_stative                    — "beträgt/betragen" in target flagged; use "ist/sind" instead
  same_selbe                          — "dieselbe*/derselbe*/dasselbe*/demselbe*/denselbe*/desselbe*" in target when source has "same"
  same_gleich_missing                 — "gleich*" absent from target when source has "same"
  comprise_umfassen                   — "compris*" count in source must match "umfass*" count in target
  vielzahl_plurality                  — "Vielzahl" count in target must match "plurality" count in source
  folgendes_umfasst                   — finite "umfasst:" before a list without "Folgendes" (→ "Folgendes umfasst:"); participial "umfassend:" is correct without "Folgendes"
  folgendes_konfiguriert              — "konfiguriert ist:" before a list without "zu Folgendem"
  dazu_konfiguriert                   — "konfiguriert" in target without "dazu" (→ "dazu konfiguriert")
  abbreviation_not_in_source          — d. h./z. B./bzw. in target when EN uses the spelled-out form
  jeweilig_not_respective             — "jeweilig*" in target when EN source has no "respective"
  german_quotation_marks              — straight "..." in target; use German „..." (Alt+0132/Alt+0147)
  patent_number_decimal               — patent Nr. decimal comma in target when source has decimal point
  acronym_in_compound                 — wrong acronym/hyphen placement in compound: (AKR-) or "(AKR) -Word"
  hyphen_in_long_compound             — hyphen between two words of 10+ chars — likely unnecessary in German compound
  durch_verwendung                    — "durch Verwendung" in target; should be "durch Verwenden" (nominalized verb)
  schritt_zum                         — "Schritt zum [Infinitiv]" in target; should be "Schritt eines [Genitiv-Infinitiv]"
  mindestens_at_least                 — "mindestens" in target but "at least" absent from source

Input / output: same *_revised_translation_checks.xlsx file (in-place,
collision-safe on PermissionError).

Usage: python linter.py [--pid <project_id>]
  --pid   project folder name under projects/; defaults to current project context
"""

import argparse
import glob
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, PatternFill

import project_log

_args = argparse.ArgumentParser()
_args.add_argument("--pid", default=None, help="Project ID (folder name under projects/). Defaults to current project context.")
_args = _args.parse_known_args()[0]

HEADER_ROWS = 3
_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# ── Checks ────────────────────────────────────────────────────────────────────
# Each check is a function (source: str, target: str) -> str | None.
# Return an annotation string when an issue is found, None otherwise.

_LEADING_PATTERNS = [          # order matters: more specific first
    re.compile(r"^\[\d+\]"),   # [0003]
    re.compile(r"^\d+\.\d+"),  # 1.2
    re.compile(r"^\d+\.(?!\d)"),  # 2.  (not followed by another digit)
]

_TRAILING_PUNCT_RE        = re.compile(r"[^\w\s]$")  # last non-word, non-space char
_TRAILING_LETTERDIGIT_RE  = re.compile(r"\w$")        # last letter or digit (no trailing punct)

_PRP_CONTRACTIONS = ["im", "vom", "am", "beim", "zum", "zur"]
_PRP_EXCEPTIONS   = ["zur Verwendung", "im Wesentlichen", "zur Verwendung", "zum Beispiel", "zum Zwecke", "im Gegensatz", "im Hinblick", "im Rahmen", "im Zusammenhang", "im Vergleich", "im Weiteren", "im Folgenden", "im Allgemeinen", "im Speziellen", "im Einzelnen", "zur Verfügung", "im Allgemeinen", "im Speziellen", "im Laufe der Zeit"]

_NEG_SOURCE_RE  = re.compile(r"\b(not|no|none)\b", re.IGNORECASE)
_BEIDE_RE           = re.compile(r"\bbeid[e]\w*\b", re.IGNORECASE)  # beide/beiden/beides/beider/beidem
_WELCHE_RE          = re.compile(r"\bwelch[e]\w*\b", re.IGNORECASE)  # welche/welcher/welches/welchem/welchen
_WERDEN_RE          = re.compile(r"\b(wird|wurde\w*|werden|geworden\w*)\b", re.IGNORECASE)
_PLURALITY_SRC_RE   = re.compile(r"\bpluralit\w*\b", re.IGNORECASE)
_PLURALITY_TGT_RE   = re.compile(r"\bVielzahl\b", re.IGNORECASE)
_COMPRISE_SRC_RE    = re.compile(r"\bcompris\w*\b", re.IGNORECASE)
_UMFASSEN_TGT_RE    = re.compile(r"\bumfass\w*\b", re.IGNORECASE)
_IN_RESPONSE_TO_RE  = re.compile(r"\bin response to\b", re.IGNORECASE)
_IN_REAKTION_RE     = re.compile(r"\bin Reaktion\b", re.IGNORECASE)
_RANGE_HYPHEN_RE    = re.compile(r"\d\s*-\s*\d")  # digit-hyphen-digit, with optional spaces

_BETRAEGT_RE        = re.compile(r"\b(beträgt|betragen)\b", re.IGNORECASE)

# ── New checks ────────────────────────────────────────────────────────────────
_FOLGENDES_UMFASST_OK_RE    = re.compile(r"\bFolgendes\s+umfass\w*\s*:", re.IGNORECASE)
_UMFASST_COLON_RE           = re.compile(r"\bumfass(?!end)\w*\s*:", re.IGNORECASE)  # excludes participial umfassend:
_FOLGENDES_KONFIG_OK_RE     = re.compile(r"\bzu\s+Folgendem\s+konfiguriert\s+ist\s*:", re.IGNORECASE)
_KONFIGURIERT_COLON_RE      = re.compile(r"\bkonfiguriert\s+ist\s*:", re.IGNORECASE)
_DAZU_KONFIGURIERT_RE       = re.compile(r"\bdazu\s+konfiguriert\b", re.IGNORECASE)
_KONFIGURIERT_IST_COMMA_RE  = re.compile(r"\bkonfiguriert\s+ist\s*,", re.IGNORECASE)  # "konfiguriert ist, zu [verb]" — valid relative clause
_KONFIGURIERT_RE            = re.compile(r"\bkonfiguriert\b", re.IGNORECASE)
_DE_ABBREV_PAIRS = [
    (re.compile(r"\bd\.\s*h\.", re.IGNORECASE), re.compile(r"\bi\.e\.", re.IGNORECASE), "d. h.", "i.e."),
    (re.compile(r"\bz\.\s*B\.", re.IGNORECASE), re.compile(r"\be\.g\.", re.IGNORECASE), "z. B.", "e.g."),
    (re.compile(r"\bbzw\.", re.IGNORECASE),      re.compile(r"\bresp\b|\brespectively\b", re.IGNORECASE), "bzw.", "resp./respectively"),
]
_JEWEILIG_RE        = re.compile(r"\bjeweilig\w*", re.IGNORECASE)
_RESPECTIVE_SRC_RE  = re.compile(r"\brespective\w*", re.IGNORECASE)
_STRAIGHT_QUOTE_RE  = re.compile(r'"[^"\n]{1,100}"')
_PATENT_NR_EN_RE    = re.compile(r"\b(?:No|Nr)\.\s*\d{5,}\.\d", re.IGNORECASE)
_PATENT_NR_DE_RE    = re.compile(r"\bNr\.\s*\d{5,},\d", re.IGNORECASE)
# hyphen inside parens: (AKR-)  |  space before hyphen: (AKR) -  |  space after hyphen: - (AKR…)
_ACRONYM_HYPHEN_RE      = re.compile(r"\([A-Z]\w*-\)|\([A-Z]\w*\)\s+-|-\s+\([A-Z]")
_LONG_COMPOUND_HYPHEN_RE = re.compile(r"\w{10,}-\w{10,}")
_DURCH_VERWENDUNG_RE    = re.compile(r"\bdurch\s+Verwendung\b", re.IGNORECASE)
_BY_GERUND_EN_RE        = re.compile(r"\bby\s+\w+ing\b", re.IGNORECASE)
_DURCH_UNG_TGT_RE       = re.compile(r"\bdurch\s+[A-Z]\w+ung\b")
_USING_EN_RE            = re.compile(r"\busing\b", re.IGNORECASE)
_UNTER_VERWENDUNG_RE    = re.compile(r"\bunter\s+Verwendung\b", re.IGNORECASE)
_SCHRITT_ZUM_RE         = re.compile(r"\bSchritt\w*\s+zum\s+[A-Z]\w+en\b")
_MINDESTENS_TGT_RE      = re.compile(r"\bmindestens\b", re.IGNORECASE)
_AT_LEAST_SRC_RE        = re.compile(r"\bat least\b", re.IGNORECASE)
_SAME_SRC_RE        = re.compile(r"\bsame\b", re.IGNORECASE)
_SELBE_TGT_RE       = re.compile(r"\b(die|der|das|dem|den|des)selb\w*\b", re.IGNORECASE)
_GLEICH_TGT_RE      = re.compile(r"\bgleich\w*\b", re.IGNORECASE)
_NUM_SPLIT_RE       = re.compile(r"[.,]")   # strip decimal/thousand separators before digit extraction
_NUM_EXTRACT_RE     = re.compile(r"\b\d+\b")

_UNIT_LIST = sorted([
    # Temperature
    "°C", "°F",
    # Frequency
    "THz", "GHz", "MHz", "kHz", "Hz",
    # Pressure
    "GPa", "MPa", "kPa", "mbar", "bar", "Pa",
    # Resistance
    "MΩ", "kΩ", "Ω",
    # Power
    "GW", "MW", "kW", "mW", "W",
    # Voltage
    "MV", "kV", "mV", "V",
    # Current
    "kA", "mA", "μA", "nA", "A",
    # Capacitance
    "mF", "μF", "nF", "pF", "F",
    # Inductance
    "mH", "μH", "nH", "H",
    # Energy
    "GJ", "MJ", "kJ", "mJ", "J", "MeV", "keV", "eV",
    # Time
    "min", "ms", "μs", "ns", "ps", "s", "h",
    # Length
    "km", "cm", "mm", "μm", "nm", "pm", "m",
    # Mass
    "kg", "mg", "μg", "ng", "g", "t",
    # Volume
    "mL", "μL", "ml", "μl", "L", "l",
    # Sound level
    "dBm", "dB",
    # Amount of substance
    "mmol", "μmol", "nmol", "mol",
    # Concentration
    "ppm", "ppb", "wt%", "vol%",
    # Rotation / data
    "rpm", "Gbit/s", "Mbit/s", "kbit/s",
    # Force
    "kN", "mN", "N",
    # Temperature / angle / percent
    "K", "%", "°",
], key=len, reverse=True)  # longest first so alternation matches greedily

_REGULAR_SPACE_UNIT_RE = re.compile(
    r"\d (?:" + "|".join(re.escape(u) for u in _UNIT_LIST) + r")(?!\w)"
)
_Z_B_REGULAR_SPACE_RE = re.compile(r"z\. B\.")  # regular space; NBSP form is z. B.
_NEG_TARGET_RE  = re.compile(r"\b(nicht|kein)", re.IGNORECASE)  # catches kein/keine/keinen/…

_CLAIM_MARKER_RE    = re.compile(r"^\d+\.(?!\d)\s*$")   # standalone "1."  (no following text)
_CLAIM_WITH_TEXT_RE = re.compile(r"^\d+\.(?!\d)\s+\S")  # "1. The device…"
_CLAIM_STRIP_RE     = re.compile(r"^\d+\.(?!\d)\s*")    # strip leading "1. " from target
_GERMAN_ARTICLE_RE  = re.compile(r"^(Der|Die|Das|Ein|Eine)\b", re.IGNORECASE)

_PRP_EXCEPTION_RES    = [re.compile(re.escape(e), re.IGNORECASE) for e in _PRP_EXCEPTIONS]
_PRP_CONTRACTION_RES  = {
    w: re.compile(r"\b" + re.escape(w) + r"\b", re.IGNORECASE)
    for w in _PRP_CONTRACTIONS
}


class GermanClaimNoArticle:
    """Flag a German article at the start of claim text.

    Fires for two segment types:
      • source starts with "N. some text"  (self-contained claim line)
      • source is preceded by a standalone claim marker "N." in the previous row

    State is carried across calls, so the singleton in CHECKS sees rows in order.
    Tests should create a fresh instance (GermanClaimNoArticle()) per test case.
    """

    def __init__(self) -> None:
        self._prev_was_marker = False

    def __call__(self, source: str, target: str) -> str | None:
        is_marker        = bool(_CLAIM_MARKER_RE.match(source))
        is_claim_segment = bool(_CLAIM_WITH_TEXT_RE.match(source)) or self._prev_was_marker
        self._prev_was_marker = is_marker

        if not is_claim_segment:
            return None

        tgt = target.strip()
        tgt_text = _CLAIM_STRIP_RE.sub("", tgt)  # remove leading "1. " if present
        m = _GERMAN_ARTICLE_RE.match(tgt_text)
        if m:
            return f'error: German claims must not start with article ("{m.group(1)}" found)'
        return None


german_claim_no_article = GermanClaimNoArticle()


def missing_leading_number(source: str, target: str) -> str | None:
    """Flag when a leading number or marker present in source is absent from target."""
    for pat in _LEADING_PATTERNS:
        m = pat.match(source.strip())
        if m:
            marker = m.group()
            if not target.strip().startswith(marker):
                return f"missing leading number: {marker} found in source but not in target"
    return None


def different_end_punctuation(source: str, target: str) -> str | None:
    """Flag when source and target end with different punctuation characters."""
    src_match = _TRAILING_PUNCT_RE.search(source.strip())
    if src_match is None:
        return None
    src_punct = src_match.group()
    tgt_match = _TRAILING_PUNCT_RE.search(target.strip())
    tgt_punct = tgt_match.group() if tgt_match else "(none)"
    if src_punct != tgt_punct:
        return f'error: different end punctuation: "{src_punct}" found in source but "{tgt_punct}" found in target'
    return None


def source_punctuation_where_none_in_target(source: str, target: str) -> str | None:
    """Flag when source ends with punctuation but target ends with a letter or digit."""
    src_match = _TRAILING_PUNCT_RE.search(source)
    if src_match is None:
        return None
    if _TRAILING_LETTERDIGIT_RE.search(target.strip()):
        return f'error: source end punctuation where none in target: "{src_match.group()}" missing in target'
    return None


def target_punctuation_where_none_in_source(source: str, target: str) -> str | None:
    """Flag when target ends with punctuation but source ends with a letter or digit."""
    if not _TRAILING_LETTERDIGIT_RE.search(source):
        return None
    tgt_match = _TRAILING_PUNCT_RE.search(target.strip())
    if tgt_match:
        return f'error: target punctuation where none in source: "{tgt_match.group()}" added in target'
    return None


def negation_not_transferred(source: str, target: str) -> str | None:
    """Flag when 'not'/'no'/'none' count in source exceeds 'nicht'/'kein' count in target."""
    src_count = len(_NEG_SOURCE_RE.findall(source))
    if src_count == 0:
        return None
    tgt_count = len(_NEG_TARGET_RE.findall(target))
    if tgt_count < src_count:
        tgt_info = f'only {tgt_count}x "nicht"/"kein"' if tgt_count > 0 else '"nicht"/"kein" not found'
        return f'error: {src_count}x "not"/"no" in source but {tgt_info} in target'
    return None


def leading_trailing_spaces(_: str, target: str) -> str | None:
    """Flag leading or trailing whitespace in the target cell value."""
    stripped = target.strip()
    if target == stripped:
        return None
    parts = []
    if target != target.lstrip():
        parts.append("leading")
    if target != target.rstrip():
        parts.append("trailing")
    return f"error: {' and '.join(parts)} spaces in target"


def regular_space_before_unit(_: str, target: str) -> str | None:
    """Flag a regular space between a number and a unit; should be non-breaking space (U+00A0)."""
    if _REGULAR_SPACE_UNIT_RE.search(target):
        return "error: use non-breaking space between number and unit, not regular space — NBSP: [ ]"
    return None


def z_b_nonbreaking_space(_: str, target: str) -> str | None:
    """Flag regular space in 'z. B.' — should be non-breaking space: z. B."""
    if _Z_B_REGULAR_SPACE_RE.search(target):
        return "error: use non-breaking space in \"z. B.\" not regular space — copy: [z. B.]"
    return None


def hyphen_in_number_range(_: str, target: str) -> str | None:
    """Flag a hyphen used between digits in target; number ranges require an en dash (–)."""
    if _RANGE_HYPHEN_RE.search(target):
        return 'error: use en dash for ranges, not hyphen — en dash: [–]'
    return None


def in_response_to_mistranslated(source: str, target: str) -> str | None:
    """Flag 'in Reaktion' in target when source contains 'in response to'."""
    if not _IN_RESPONSE_TO_RE.search(source):
        return None
    if _IN_REAKTION_RE.search(target):
        return 'error: "in response to" > "als Reaktion auf" (not "in Reaktion")'
    return None


def plurality_not_transferred(source: str, target: str) -> str | None:
    """Flag when 'plurality' count in source exceeds 'Vielzahl' count in target."""
    src_count = len(_PLURALITY_SRC_RE.findall(source))
    if src_count == 0:
        return None
    tgt_count = len(_PLURALITY_TGT_RE.findall(target))
    if tgt_count < src_count:
        tgt_info = f'only {tgt_count}x "Vielzahl"' if tgt_count > 0 else '"Vielzahl" not found'
        return f'error: {src_count}x "plurality" in source but {tgt_info} in target'
    return None


def vielzahl_plurality(source: str, target: str) -> str | None:
    """Flag when 'Vielzahl' count in target does not match 'plurality' count in source."""
    tgt_count = len(_PLURALITY_TGT_RE.findall(target))
    if tgt_count == 0:
        return None
    src_count = len(_PLURALITY_SRC_RE.findall(source))
    if src_count == tgt_count:
        return None
    src_info = f'only {src_count}x "plurality"' if src_count > 0 else '"plurality" not found'
    return f'error: {tgt_count}x "Vielzahl" in target but {src_info} in source'


def comprise_umfassen(source: str, target: str) -> str | None:
    """Flag when 'umfass*' count in target does not match 'compris*' count in source."""
    tgt_count = len(_UMFASSEN_TGT_RE.findall(target))
    if tgt_count == 0:
        return None
    src_count = len(_COMPRISE_SRC_RE.findall(source))
    if src_count == tgt_count:
        return None
    src_info = f'only {src_count}x "compris*"' if src_count > 0 else '"compris*" not found'
    return f'error: {tgt_count}x "umfass*" in target but {src_info} in source'


def werden_dynamic(_: str, target: str) -> str | None:
    """Flag 'wird/wurde*/werden/geworden*' in target; double-check static vs. dynamic phrasing."""
    m = _WERDEN_RE.search(target)
    if m:
        return f'error: double check static state/dynamic process ("{m.group()}" found)'
    return None


def welche_relativpronomen(_: str, target: str) -> str | None:
    """Flag 'welche*' in target; relative clauses should use der/die/das."""
    m = _WELCHE_RE.search(target)
    if m:
        return f'error: Relativpronomen der, die das (not "{m.group()}")'
    return None


def beide_ambiguous(_: str, target: str) -> str | None:
    """Flag 'beide*' in target as potentially ambiguous."""
    m = _BEIDE_RE.search(target)
    if m:
        return f'error: "{m.group()}" in target — beide = zwei | einer von (either) | sowohl als auch (both the X and Y)'
    return None


def same_selbe(source: str, target: str) -> str | None:
    """Flag 'dieselbe*/derselbe*/dasselbe*' in target when source contains 'same'."""
    if not _SAME_SRC_RE.search(source):
        return None
    m = _SELBE_TGT_RE.search(target)
    if m:
        return f'error: "{m.group()}" in target — same = gleiche, also check article'
    return None


def same_gleich_missing(source: str, target: str) -> str | None:
    """Flag when source contains 'same' but target contains no 'gleich*'."""
    if not _SAME_SRC_RE.search(source):
        return None
    if not _GLEICH_TGT_RE.search(target):
        return 'error: source: "same", target: "gleich" missing'
    return None


def betraegt_stative(_: str, target: str) -> str | None:
    """Flag 'beträgt/betragen' in target; static values should use 'ist/sind'."""
    m = _BETRAEGT_RE.search(target)
    if m:
        return f'error: "{m.group()}" in target — beträgt = ist'
    return None


def preposition_contraction(_: str, target: str) -> str | None:
    """Flag German preposition+article contractions (im/vom/am/beim/zum/zur) in target."""
    text = target.strip()
    for exc_re in _PRP_EXCEPTION_RES:
        text = exc_re.sub("", text)
    found = [w for w, pat in _PRP_CONTRACTION_RES.items() if pat.search(text)]
    if found:
        joined = ", ".join(f'"{w}"' for w in found)
        return f"error: Prp + article contraction {joined} used in target"
    return None


def folgendes_umfasst(_: str, target: str) -> str | None:
    """Flag finite 'umfasst:' before a list without 'Folgendes' — correct: 'Folgendes umfasst:'.
    Participial 'umfassend:' (preferred translation of 'comprising') is correct without 'Folgendes'."""
    stripped = _FOLGENDES_UMFASST_OK_RE.sub("", target)
    if _UMFASST_COLON_RE.search(stripped):
        return 'error: "umfasst:" before list — use "Folgendes umfasst:"'
    return None


def folgendes_konfiguriert(_: str, target: str) -> str | None:
    """Flag 'konfiguriert ist:' without 'zu Folgendem' — correct: 'zu Folgendem konfiguriert ist:'."""
    stripped = _FOLGENDES_KONFIG_OK_RE.sub("", target)
    if _KONFIGURIERT_COLON_RE.search(stripped):
        return 'error: "konfiguriert ist:" before list — use "zu Folgendem konfiguriert ist:"'
    return None


def dazu_konfiguriert(_: str, target: str) -> str | None:
    """Flag 'konfiguriert' in target without 'dazu' — correct: 'dazu konfiguriert'.
    Excludes 'konfiguriert ist, [zu-infinitive]' (valid relative-clause form)."""
    stripped = _DAZU_KONFIGURIERT_RE.sub("", target)
    stripped = _KONFIGURIERT_IST_COMMA_RE.sub("", stripped)
    if _KONFIGURIERT_RE.search(stripped):
        return "'konfiguriert' ohne 'dazu' verwenden. Nur 'dazu konfiguriert' verwenden, wenn auch 'zu Folgendem konfiguriert(ist):' verwendet wird."
    return None


def abbreviation_not_in_source(source: str, target: str) -> str | None:
    """Flag d. h./z. B./bzw. in target when EN source uses the spelled-out form."""
    for tgt_re, src_re, de_abbr, en_abbr in _DE_ABBREV_PAIRS:
        if tgt_re.search(target) and not src_re.search(source):
            return f'error: "{de_abbr}" in target but "{en_abbr}" not in source — spell out in full'
    return None


def jeweilig_not_respective(source: str, target: str) -> str | None:
    """Flag 'jeweilig*' in target when EN source has no 'respective/respectively'."""
    m = _JEWEILIG_RE.search(target)
    if m and not _RESPECTIVE_SRC_RE.search(source):
        return f'error: "{m.group()}" in target but "respective" not in source'
    return None


def german_quotation_marks(_: str, target: str) -> str | None:
    """Flag straight double quotes in target; use German „..." (Alt+0132/Alt+0147)."""
    if _STRAIGHT_QUOTE_RE.search(target):
        return 'error: use German quotation marks „..." (Alt+0132 open, Alt+0147 close)'
    return None


def durch_verwendung(source: str, target: str) -> str | None:
    """Flag 'durch Verwendung' (result noun) — should be 'durch Verwenden' (nominalized verb).
    Also flags 'durch [Capitalized-Noun-ung]' when EN source contains 'by [verb]ing'.
    """
    if _DURCH_VERWENDUNG_RE.search(target):
        return 'error: "durch Verwendung" — use "durch Verwenden" (nominalized verb, not result noun)'
    if _BY_GERUND_EN_RE.search(source):
        m = _DURCH_UNG_TGT_RE.search(target)
        if m:
            return f'error: "{m.group()}" — use "durch [Verb-en]" (nominalized verb, not result noun)'
    return None


def unter_verwendung(source: str, target: str) -> str | None:
    """Flag 'unter Verwendung' in target when EN source contains 'using'.
    Prefer 'unter Verwenden/mittels'
    """
    if _USING_EN_RE.search(source) and _UNTER_VERWENDUNG_RE.search(target):
        return 'warning: "unter Verwendung" — prefer "mithilfe" (unter Verwendung is allowed)'
    return None


def schritt_zum(_: str, target: str) -> str | None:
    """Flag 'Schritt zum [Infinitiv]' — should be 'Schritt eines [Genitiv-Infinitiv]'."""
    m = _SCHRITT_ZUM_RE.search(target)
    if m:
        return f'error: "{m.group()}" — use "Schritt eines [Verb-ens]" (genitive of nominalized verb)'
    return None


def mindestens_at_least(source: str, target: str) -> str | None:
    """Flag 'mindestens' in target when source does not contain 'at least'."""
    if not _MINDESTENS_TGT_RE.search(target):
        return None
    if not _AT_LEAST_SRC_RE.search(source):
        return 'error: "mindestens" in target but "at least" not in source'
    return None


def hyphen_in_long_compound(_: str, target: str) -> str | None:
    """Flag a hyphen between two words of 10+ characters — likely an unnecessary compound hyphen (Style Guide §3.10)."""
    m = _LONG_COMPOUND_HYPHEN_RE.search(target)
    if m:
        return f'error: unnecessary hyphen in compound "{m.group()}" — merge into one word (Style Guide §3.10)'
    return None


def acronym_in_compound(_: str, target: str) -> str | None:
    """Flag wrong acronym/hyphen placement in compound nouns.

    Wrong: (AKR-)  →  hyphen must follow the closing parenthesis: (AKR)-
    Wrong: (AKR) - or - (AKR)  →  no space between parenthesis and hyphen.
    """
    m = _ACRONYM_HYPHEN_RE.search(target)
    if m:
        return (
            f'error: wrong acronym/hyphen placement "{m.group().strip()}" — '
            'correct form is "(AKR)-Word" (hyphen after closing parenthesis, no spaces)'
        )
    return None


def patent_number_decimal(source: str, target: str) -> str | None:
    """Flag decimal comma in a patent application number when source has a decimal point."""
    if _PATENT_NR_EN_RE.search(source) and _PATENT_NR_DE_RE.search(target):
        return 'error: patent application number — keep decimal point as in source, not comma'
    return None


def _extract_numbers(text: str) -> list[str]:
    """Return sorted list of digit sequences, normalising decimal/thousand separators."""
    return sorted(_NUM_EXTRACT_RE.findall(_NUM_SPLIT_RE.sub(" ", text)))


def numeric_mismatch(source: str, target: str) -> str | None:
    """Flag when numbers present in source are absent from target or vice versa."""
    src_nums = _extract_numbers(source)
    tgt_nums = _extract_numbers(target)
    if src_nums == tgt_nums:
        return None
    from collections import Counter
    src_c, tgt_c = Counter(src_nums), Counter(tgt_nums)
    missing = sorted((src_c - tgt_c).elements())
    added   = sorted((tgt_c - src_c).elements())
    parts = []
    if missing:
        parts.append(f"missing in target: {missing}")
    if added:
        parts.append(f"added in target: {added}")
    return f'error: numeric mismatch — {"; ".join(parts)}'


CHECKS = [
    german_claim_no_article,
    missing_leading_number,
    different_end_punctuation,
    source_punctuation_where_none_in_target,
    target_punctuation_where_none_in_source,
    leading_trailing_spaces,
    negation_not_transferred,
    regular_space_before_unit,
    z_b_nonbreaking_space,
    hyphen_in_number_range,
    in_response_to_mistranslated,
    plurality_not_transferred,
    beide_ambiguous,
    welche_relativpronomen,
    werden_dynamic,
    preposition_contraction,
    betraegt_stative,
    same_selbe,
    same_gleich_missing,
    comprise_umfassen,
    vielzahl_plurality,
    folgendes_umfasst,
    folgendes_konfiguriert,
    dazu_konfiguriert,
    abbreviation_not_in_source,
    jeweilig_not_respective,
    german_quotation_marks,
    patent_number_decimal,
    acronym_in_compound,
    hyphen_in_long_compound,
    durch_verwendung,
    unter_verwendung,
    schritt_zum,
    mindestens_at_least,
    numeric_mismatch,
]


if __name__ == "__main__":
    # ── Locate input file ─────────────────────────────────────────────────────

    if _args.pid:
        proj_dir = project_log.find_project_dir(_args.pid)
    else:
        proj_dir = project_log.project_dir()

    src_files = [
        f for f in glob.glob(str(proj_dir / "*_revised_translation_checks.xlsx"))
        if not Path(f).name.startswith("~$")
    ]
    if not src_files:
        raise FileNotFoundError(f"No *_revised_translation_checks.xlsx found in {proj_dir}")

    src_path = Path(src_files[0])
    print(f"Input: {src_path.name}")

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active

    # ── Insert column D (idempotent: only skip if linter column already present)

    already_linted = (
        ws.cell(row=2, column=4).value == "Linter"
        and ws.cell(row=2, column=5).value == "Glossary Checks"
    )

    if already_linted:
        print("Linter column already present — overwriting annotations.")
        for row_num in range(HEADER_ROWS + 1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=4)
            cell.value = None
            cell.fill = PatternFill(fill_type="none")
    else:
        ws.insert_cols(4)

    ws.cell(row=2, column=4).value = "Linter"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35

    # ── Run checks over every data row ───────────────────────────────────────

    annotated = 0

    for row_num in range(HEADER_ROWS + 1, ws.max_row + 1):
        src_text = ws.cell(row=row_num, column=2).value
        tgt_text = ws.cell(row=row_num, column=3).value

        if not src_text:
            continue

        src_str = str(src_text).strip()
        tgt_str = str(tgt_text) if tgt_text else ""  # raw — checks do their own stripping

        issues = [
            result
            for check in CHECKS
            for result in (check(src_str, tgt_str),)
            if result is not None
        ]

        if issues:
            lint_cell = ws.cell(row=row_num, column=4)
            lint_cell.value = "\n".join(issues)
            lint_cell.alignment = Alignment(wrap_text=True)
            lint_cell.fill = _FILL
            annotated += 1

    # ── Save ─────────────────────────────────────────────────────────────────
    # If the file is open in Excel (Windows ~$ lock file exists), writing to it
    # while Excel has it open causes a "file modified externally" repair prompt
    # on reload.  Save to a timestamped copy instead.

    lock_file = src_path.parent / f"~${src_path.name}"
    if lock_file.exists():
        stamp = datetime.now().strftime("%H%M%S")
        out_path = src_path.parent / src_path.name.replace(".xlsx", f"_{stamp}.xlsx")
        print(f"File is open in Excel — saving to {out_path.name}")
    else:
        out_path = src_path

    try:
        wb.save(out_path)
    except PermissionError:
        stamp = datetime.now().strftime("%H%M%S")
        out_path = src_path.parent / src_path.name.replace(".xlsx", f"_{stamp}.xlsx")
        wb.save(out_path)

    print(f"Annotated {annotated} segment(s).")
    print(f"Saved: {out_path}")
