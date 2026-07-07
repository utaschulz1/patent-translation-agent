# ============================================================
# lara_translate.py
# ============================================================
# Translates bilingual XTM Excel files using the Lara Translate API.
#
# SETUP
#   pip install lara-sdk
#   Add to .env:
#     LARA_ACCESS_KEY_ID=your-access-key-id
#     LARA_ACCESS_KEY_SECRET=your-access-key-secret
#     LARA_GLOSSARY_IDS=gls_abc123,gls_def456   # optional, Pro plan
#     LARA_MEMORY_IDS=mem_abc123,mem_def456      # optional, Team plan only
#
# USAGE   python lara_translate.py [--pid <project_id>] [--file <filename>] [--seg-range START-END]
#           --pid    project folder name under projects/; defaults to current project context
#           --file   explicit xlsx filename to translate (overrides auto-detection)
#
# INPUT   projects/<project_id>/<project_id>*.xlsx   — bilingual Excel from XTM (auto-detected)
#         or any .xlsx specified via --file
# OUTPUT  projects/<project_id>/<name>_translated.xlsx
# ============================================================

import argparse
import os
import glob
import json
import time
import openpyxl
import pandas as pd
from dotenv import load_dotenv
from pathlib import Path

from lara_sdk import AccessKey, TextBlock, Translator

from project_log import project_dir as _pdir

_args = argparse.ArgumentParser()
_args.add_argument("--pid", default=None, help="Project ID (folder name under projects/). Defaults to current project context.")
_args.add_argument("--file", default=None, metavar="FILENAME",
                   help="Exact xlsx filename (or path) to translate. Overrides auto-detection.")
_args.add_argument("--seg-range", default=None, metavar="START-END",
                   help="Translate only this segment range, e.g. 421-488.")
_args = _args.parse_args()

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

SRC_LANG = "en"
TGT_LANG = "de"
INSTRUCTIONS = ["Use precise, formal language suitable for patent claims and descriptions."]
CONTEXT_BEFORE = 5   # preceding segments as context (grammar/clause structure)
CONTEXT_AFTER  = 1   # following segments as context; total window max 128

START_SEGMENT_ID  = None   # e.g. "668" — start from this segment; None = first
END_SEGMENT_ID    = None   # e.g. "700" — stop after this segment; None = last
FORCE_RETRANSLATE = True   # True = ignore existing output and retranslate everything

if _args.seg_range:
    _parts = _args.seg_range.split("-")
    START_SEGMENT_ID = _parts[0].strip()
    END_SEGMENT_ID   = _parts[1].strip() if len(_parts) > 1 else _parts[0].strip()

# ============================================================
# Credentials
# ============================================================

access_key_id     = os.environ.get("LARA_ACCESS_KEY_ID", "").strip()
access_key_secret = os.environ.get("LARA_ACCESS_KEY_SECRET", "").strip()

if not access_key_id or not access_key_secret:
    print("ERROR: LARA_ACCESS_KEY_ID and LARA_ACCESS_KEY_SECRET must be set in .env.")
    print("  Get them from: https://www.laratranslate.com/ → Account → API Credentials")
    exit()

lara = Translator(AccessKey(id=access_key_id, secret=access_key_secret))

# ============================================================
# Optional: memory IDs and glossary IDs from .env
# ============================================================

def _parse_ids(env_key: str) -> list[str]:
    raw = os.environ.get(env_key, "").strip()
    return [x.strip() for x in raw.split(",") if x.strip()] if raw else []

memory_ids = _parse_ids("LARA_MEMORY_IDS")   # Team plan required for adapt_to

_glossaries_file = Path(__file__).parent / "lara_glossaries.json"
_glossaries_registry: dict = json.loads(_glossaries_file.read_text(encoding="utf-8")) if _glossaries_file.exists() else {}

_memories_file = Path(__file__).parent / "lara_memories.json"
_memories_registry: dict = json.loads(_memories_file.read_text(encoding="utf-8")) if _memories_file.exists() else {}

# ============================================================
# Load Excel
# ============================================================

if _args.pid:
    proj_dir = Path(__file__).parent / "projects" / _args.pid
    if not proj_dir.exists():
        print(f"ERROR: Project folder not found: {proj_dir}")
        exit()
else:
    proj_dir = _pdir()

_glossary_key = f"glossary_{proj_dir.name}"
if _glossary_key in _glossaries_registry:
    glossary_ids = [_glossaries_registry[_glossary_key]]
    print(f"Glossary:      {glossary_ids[0]}  ({_glossary_key})")
elif _glossaries_registry:
    print(f"No glossary found for project {proj_dir.name!r} — none will be used.")
    glossary_ids = []
else:
    glossary_ids = []

_memory_key = f"memory_{proj_dir.name}"
if _memory_key in _memories_registry:
    memory_ids = [_memories_registry[_memory_key]] + memory_ids
    print(f"ICE TM:        {_memories_registry[_memory_key]}  ({_memory_key})")
if memory_ids:
    print(f"TM adaptation: {memory_ids}")
if _args.file:
    _p = Path(_args.file)
    input_path = str(_p if _p.is_absolute() else proj_dir / _p)
    if not os.path.exists(input_path):
        print(f"ERROR: File not found: {input_path}")
        exit()
else:
    xlsx_files = glob.glob(str(proj_dir / "*.xlsx"))
    xlsx_files = [f for f in xlsx_files
                  if not os.path.basename(f).startswith("~$")
                  and not f.endswith("_translated.xlsx")]

    # Prefer files whose name starts with the project ID (XTM export naming convention)
    project_files = [f for f in xlsx_files
                     if os.path.basename(f).lower().startswith(proj_dir.name.lower())]
    if project_files:
        xlsx_files = project_files

    if not xlsx_files:
        print(f"ERROR: No source .xlsx file found in '{proj_dir}'.")
        print(f"  Tip: use --file <filename> to specify it explicitly.")
        exit()
    if len(xlsx_files) > 1:
        print(f"Multiple .xlsx files found, using: {os.path.basename(xlsx_files[0])}")
        print(f"  Tip: use --file <filename> to select a specific file.")

    input_path = xlsx_files[0]

raw_df  = pd.read_excel(input_path, header=None, engine="openpyxl")
print(f"Processing: {raw_df.iloc[0, 0]}")

data_df = raw_df.iloc[3:].reset_index(drop=True)
data_df.columns = ["ID", "Source", "Target"] + list(data_df.columns[3:])
data_df = data_df[["ID", "Source", "Target"]].copy()
data_df.dropna(subset=["Source"], inplace=True)
data_df["ID"]     = data_df["ID"].astype(str).str.strip()
data_df["Source"] = data_df["Source"].astype(str).str.strip()
data_df["Target"] = data_df["Target"].fillna("").astype(str).str.strip()
segments = data_df.to_dict("records")
all_ids = [s["ID"] for s in segments]
print(f"Loaded {len(segments)} segments (IDs {all_ids[0]} – {all_ids[-1]}).")

if START_SEGMENT_ID or END_SEGMENT_ID:
    start_id = str(START_SEGMENT_ID) if START_SEGMENT_ID else all_ids[0]
    end_id   = str(END_SEGMENT_ID)   if END_SEGMENT_ID   else all_ids[-1]
    if start_id not in all_ids:
        print(f"ERROR: START_SEGMENT_ID {start_id!r} not found.")
        exit()
    if end_id not in all_ids:
        print(f"ERROR: END_SEGMENT_ID {end_id!r} not found.")
        exit()
    segments = segments[all_ids.index(start_id):all_ids.index(end_id) + 1]
    print(f"Range: {start_id} – {end_id} ({len(segments)} segments).")

# ============================================================
# Output workbook — incremental writes; resume if exists
# ============================================================

out_path = input_path.replace(".xlsx", "_translated.xlsx")

if os.path.exists(out_path):
    while True:
        try:
            with open(out_path, "a"):
                pass
            break
        except PermissionError:
            input(f"  {os.path.basename(out_path)} is open in Excel — close it, then press Enter...")

resuming = os.path.exists(out_path) and not FORCE_RETRANSLATE
if resuming:
    wb = openpyxl.load_workbook(out_path)
    print(f"Resuming from existing output: {os.path.basename(out_path)}")
else:
    wb = openpyxl.load_workbook(input_path)
    if FORCE_RETRANSLATE:
        print("FORCE_RETRANSLATE=True — ignoring existing output.")

ws = wb.active

row_map: dict[str, int] = {}
for row_num in range(4, ws.max_row + 1):
    seg_id = str(ws.cell(row=row_num, column=1).value).strip()
    row_map[seg_id] = row_num

# ============================================================
# Translate segments
# ============================================================

translations: dict[str, str] = {}
errors: list[str] = []

# Pre-populate from resumed output — only segments marked "lara" in column 4,
# which is written when Lara successfully translates a segment.  XTM TM pre-fills
# in column 3 are ignored because they never get the column-4 marker.
if resuming:
    for seg_id, row_num in row_map.items():
        marker = ws.cell(row=row_num, column=4).value
        if str(marker).strip().lower() == "lara":
            val = ws.cell(row=row_num, column=3).value
            if val and str(val).strip():
                translations[seg_id] = str(val).strip()
    if translations:
        print(f"Resuming: {len(translations)} Lara-translated segments found, skipping.")

print(f"\nTranslating {len(segments)} segments...")

for i, seg in enumerate(segments):
    if seg["ID"] in translations:
        continue

    kwargs: dict = {
        "source":       SRC_LANG,
        "target":       TGT_LANG,
        "instructions": INSTRUCTIONS,
        "no_trace":     True,   # patent content is confidential
    }
    if memory_ids:
        kwargs["adapt_to"] = memory_ids
    if glossary_ids:
        kwargs["glossaries"] = glossary_ids

    try:
        before = segments[max(0, i - CONTEXT_BEFORE):i]
        after  = segments[i + 1:i + 1 + CONTEXT_AFTER]
        window = (
            [TextBlock(text=s["Source"], translatable=False) for s in before]
            + [TextBlock(text=seg["Source"], translatable=True)]
            + [TextBlock(text=s["Source"], translatable=False) for s in after]
        )
        target_idx = len(before)  # position of the translatable block in the window

        result = lara.translate(window, **kwargs)
        translated = result.translation[target_idx].text.strip()
        translations[seg["ID"]] = translated

        if seg["ID"] in row_map:
            ws.cell(row=row_map[seg["ID"]], column=3).value = translated
            ws.cell(row=row_map[seg["ID"]], column=4).value = "lara"
            try:
                wb.save(out_path)
            except PermissionError:
                print(f"    WARNING: Could not save — close {os.path.basename(out_path)} in Excel.")

        print(f"  [{i+1}/{len(segments)}] {seg['ID']}: {translated[:80]}")

    except Exception as e:
        print(f"  [{i+1}/{len(segments)}] {seg['ID']}: error — {e}")
        errors.append(seg["ID"])
        if "401" in str(e) or "Unauthorized" in str(e):
            print("\nAborted: 401 Unauthorized — check LARA_ACCESS_KEY_ID and LARA_ACCESS_KEY_SECRET in .env.")
            break

    time.sleep(0.2)

print(f"\nDone. {len(translations)} translated, {len(errors)} errors.")

while True:
    try:
        wb.save(out_path)
        break
    except PermissionError:
        input(f"\n  Could not save — close {os.path.basename(out_path)} in Excel, then press Enter to retry...")

print(f'Output: "{out_path}".')

if errors:
    print(f"\nSegments with errors: {', '.join(errors)}")
