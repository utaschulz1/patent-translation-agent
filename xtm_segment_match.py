"""
xtm_segment_match.py — Issue Resolution step: find which XTM segments need correcting

Reads pre-processing/issue_resolution_xtm_corrections.json — a small
[{"old_text": "...", "new_text": "..."}] list authored by the human/skill
session after reviewing the docx (the judgment of "is this a text-content
fix that needs propagating to XTM, or an image/linebreak/DTP-only thing
that doesn't" stays human — this script only does the mechanical lookup).

Matches each old_text against the Target column of the top-level
Final_<stem>.xlsx (from XTM_SEGMENTS_DOWNLOAD) via exact substring search —
the same approach that found segment 11's 3-space typo by hand on
GRGM_2607_P0033. Never guesses: zero or multiple matches for a pair is a
hard failure, not a best guess.

Also copies the exact Final_<stem>.xlsx it read to
Final_<stem>_preupload_snapshot.xlsx — a plain local file copy, no extra XTM
round-trip — freezing the baseline xtm_verify_correction.py will later diff
against. (There is deliberately no separate snapshot-download step: with no
XTM edits happening between XTM_SEGMENTS_DOWNLOAD and this point, a fresh
download would just be the same state again, wasting a round-trip against a
server known to be laggy. If something *does* change in the gap, the
±2-context-segment check in xtm_verify_correction.py is exactly the
mechanism that catches that drift — a fresher snapshot wouldn't add safety,
only hide the drift verification exists to detect.)

Usage:
    python xtm_segment_match.py --pid <project_id>

Writes pre-processing/<stem>_revised_translation_checks_issue_resolution.xlsx
in the exact format xtm_upload_translations.py's _read_translations() already
expects. If the corrections list is empty, exits 0 printing "no XTM
correction needed" and writes nothing — that absence is the signal
downstream steps key off.
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

import project_log

CORRECTIONS_FILENAME = "issue_resolution_xtm_corrections.json"


def _find_final_xlsx(project_folder: Path) -> Path:
    """Finds the Final_*.xlsx written by XTM_SEGMENTS_DOWNLOAD, excluding any
    pre-upload snapshot copy.

    Raises:
        FileNotFoundError: none found.
        ValueError: more than one found.
    """
    matches = list(project_folder.glob("Final_*.xlsx"))
    matches = [p for p in matches if "_preupload_snapshot" not in p.stem]
    if not matches:
        raise FileNotFoundError(
            f"No Final_*.xlsx in {project_folder}. Run XTM_SEGMENTS_DOWNLOAD first."
        )
    if len(matches) > 1:
        raise ValueError(f"Multiple Final_*.xlsx found: {[p.name for p in matches]}")
    return matches[0]


def _load_rows(xlsx_path: Path) -> list[tuple]:
    """Return (id, source, target) for every data row (row 4+) of the Extended Table."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        seg_id = row[0]
        if seg_id is None:
            continue
        target = row[2] if len(row) > 2 else None
        rows.append((int(seg_id), row[1], target or ""))
    return rows


def match_corrections(rows: list[tuple], corrections: list[dict]) -> list[tuple[int, str]]:
    """Returns [(segment_id, new_target_text), ...]. Raises on 0 or >1 matches for any pair."""
    matched = []
    for correction in corrections:
        old_text = correction["old_text"]
        new_text = correction["new_text"]
        hits = [(seg_id, target) for seg_id, _source, target in rows if old_text in (target or "")]
        if not hits:
            raise ValueError(f"No segment found whose Target contains: {old_text!r}")
        if len(hits) > 1:
            ids = [h[0] for h in hits]
            raise ValueError(f"Multiple segments ({ids}) contain: {old_text!r} — need a more specific old_text")
        seg_id, target = hits[0]
        new_target = target.replace(old_text, new_text)
        matched.append((seg_id, new_target))
    return matched


def _write_checks_xlsx(path: Path, matched: list[tuple[int, str]], source_stem: str) -> None:
    """Writes the matched corrections in the exact format
    xtm_upload_translations.py's _read_translations() expects (3-row header,
    Column A = id, Column C = target)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Issue Resolution corrections"

    ws.cell(row=1, column=1, value=source_stem)
    header_font = Font(bold=True)
    for col, label in enumerate(["Id", "Source", "Target"], start=1):
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = header_font
    ws.cell(row=3, column=3, value="German (Germany)")

    for i, (seg_id, text) in enumerate(matched, start=4):
        ws.cell(row=i, column=1, value=seg_id)
        ws.cell(row=i, column=3, value=text)

    wb.save(path)


def run(project_id: str) -> Path | None:
    """Matches issue_resolution_xtm_corrections.json against XTM segments and
    writes the corrections checks xlsx, if any correction is actually needed.

    Short-circuits (no-op, returns None) if issue_resolution_status.json says
    any_needed_work is False — see the module docstring — without requiring
    the corrections file to exist in that case.

    Args:
        project_id: the project to match corrections for.

    Returns:
        Path to the written checks xlsx, or None if nothing needed matching
        (either because the job needed no resolution at all, or the human
        reviewer's corrections list was empty).

    Raises:
        FileNotFoundError: issue_resolution_xtm_corrections.json is missing
            (and any_needed_work wasn't False), or no Final_*.xlsx exists yet.
        ValueError: a correction's old_text matches zero or multiple segments.
    """
    pre_folder = project_log.find_project_dir(project_id)
    project_folder = pre_folder.parent

    status_path = pre_folder / "issue_resolution_status.json"
    if status_path.exists():
        status = json.loads(status_path.read_text(encoding="utf-8"))
        if not status.get("any_needed_work", True):
            print("issue_resolution_status.json: nothing needed resolving for this job — "
                  "no XTM correction possible, skipping (no corrections file required).")
            return None

    corrections_path = pre_folder / CORRECTIONS_FILENAME
    if not corrections_path.exists():
        raise FileNotFoundError(
            f"No {CORRECTIONS_FILENAME} in {pre_folder}. "
            "Author it after the interactive review — a JSON list of "
            '{"old_text": "...", "new_text": "..."} pairs for text-content fixes only.'
        )
    corrections = json.loads(corrections_path.read_text(encoding="utf-8"))

    final_xlsx = _find_final_xlsx(project_folder)
    print(f"Reading: {final_xlsx.name}")
    rows = _load_rows(final_xlsx)

    snapshot_path = project_folder / f"{final_xlsx.stem}_preupload_snapshot.xlsx"
    snapshot_path.write_bytes(final_xlsx.read_bytes())
    print(f"Snapshot frozen: {snapshot_path.name}")

    if not corrections:
        print("No XTM correction needed (empty corrections list).")
        return None

    matched = match_corrections(rows, corrections)
    print(f"Matched {len(matched)} correction(s):")
    for seg_id, text in matched:
        print(f"  segment {seg_id}: {text[:80]!r}")

    out_path = pre_folder / f"{final_xlsx.stem}_revised_translation_checks_issue_resolution.xlsx"
    _write_checks_xlsx(out_path, matched, final_xlsx.stem)
    print(f"Written: {out_path}")
    return out_path


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--pid", required=True)
    args = parser.parse_args()
    try:
        run(args.pid)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
