"""Build the annotated checks xlsx from the original checks file + verdicts.

Usage: python build_annotated.py <checks.xlsx> <verdicts.json> <output.xlsx>

verdicts.json: a list of objects, each keyed by the row's stable "id"
(the ID column value — NOT the spreadsheet row number, which shifts and
is easy to mis-key by hand):

    [
      {"id": 3, "linter_verdict": "...", "glossary_verdict": "...", "additional": "..."},
      {"id": 7, "additional": "..."},
      ...
    ]

Any of linter_verdict/glossary_verdict/additional may be omitted; rows
with no entry in verdicts.json get blank verdict/additional cells.

Verdict text should start with the capitalized marker (TRUE POSITIVE /
FALSE POSITIVE / MIXED / PARTIAL / SUMMARY) — this script does NOT
color-code cells (Excel handles fill colors badly in dark mode); the
marker text is the only signal, by design.
"""
import json
import sys

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

HEADERS = ["ID", "Source", "Target", "Linter", "Linter Verdict", "Glossary Checks", "Glossary Verdict", "Additional Issues Found"]
WIDTHS = [5, 45, 45, 22, 40, 22, 45, 45]


def build(checks_xlsx: str, verdicts_json: str, out_path: str) -> None:
    wb_in = openpyxl.load_workbook(checks_xlsx, data_only=True)
    ws_in = wb_in.active
    all_rows = list(ws_in.iter_rows(values_only=True))
    title = all_rows[0][0] if all_rows else None

    verdicts = json.loads(open(verdicts_json, encoding="utf-8").read())
    by_id = {v["id"]: v for v in verdicts}

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet"

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    if title is not None:
        ws_out["A1"] = title
        ws_out.merge_cells("A1:H1")
        ws_out["A1"].font = bold

    ws_out.append(HEADERS)
    for c in range(1, 9):
        ws_out.cell(row=2, column=c).font = bold

    ws_out.append(("", "EN", "DE", "", "", "", "", ""))

    for row in all_rows:
        id_, source, target, linter, glossary = (list(row) + [None] * 5)[:5]
        if id_ is None or not isinstance(id_, (int, float)):
            continue  # skip title/header/EN-DE rows from the source file
        v = by_id.get(int(id_), {})
        ws_out.append(
            (
                id_,
                source,
                target,
                linter,
                v.get("linter_verdict", ""),
                glossary,
                v.get("glossary_verdict", ""),
                v.get("additional", ""),
            )
        )
        out_row = ws_out.max_row
        for c in range(1, 9):
            ws_out.cell(row=out_row, column=c).alignment = wrap

    for idx, w in enumerate(WIDTHS, start=1):
        ws_out.column_dimensions[get_column_letter(idx)].width = w

    wb_out.save(out_path)


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)
    build(sys.argv[1], sys.argv[2], sys.argv[3])
    print(f"wrote {sys.argv[3]}")
