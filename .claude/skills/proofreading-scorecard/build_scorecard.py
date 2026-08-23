"""
build_scorecard.py — mechanical half of the proofreading-scorecard skill.

Given a project folder, finds the translator's XTM bilingual export(s) and
the proofreader's corrected version(s), diffs them segment-by-segment
(matched by Id), and writes a "Tracked Changes" worksheet into the project's
copy of the client's DQF-MQM scorecard template — with word-level tracked
changes rendered as xlsx rich text (strikethrough red = deleted, underline
blue = inserted). No Word round-trip involved.

Also fills the Task Details cells that don't require judgment: fixed values
(Copy Editor initials, locale, today's date), and — if --job-id is given —
the XTRF-derived fields (project number, weighted word count, content
specialty, budgeted time), by reusing xtrf_job_setup.py's session helpers.

This script does NOT touch Part Two (Errors) or Part Three (Overall
Summary) — those require reading the tracked changes and exercising
judgment, which is the SKILL.md reasoning stage's job.

Usage:
    python build_scorecard.py <project_dir> [--job-id 373610] [--output PATH]
"""

import argparse
import csv
import re
import shutil
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter

APP_DIR = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(APP_DIR / "agent"))

from config import extract_project_id  # noqa: E402

TEMPLATE_PATH = Path(__file__).parent / "Translator Scorecard 5.14.26.xlsx"

FONT_DEL = InlineFont(strike=True, color="FF0000")
FONT_INS = InlineFont(u="single", color="0000FF")

TASK_DETAIL_CELLS = {
    "project_number": "E5",
    "copy_editor": "E7",
    "weighted_word_count": "E8",
    "content_specialty": "E9",
    "target_locale": "E10",
    "source_locale": "E11",
    "time_spent_minutes": "E12",
    "date": "E13",
    "agency_or_fl": "F35",
}


# ── xlsx pair discovery ──────────────────────────────────────────────────────

def find_xlsx_pairs(project_dir: Path) -> list[tuple[str, Path, Path]]:
    """Find (doc_label, translated_path, proofread_path) triples in project_dir.

    Only looks at the top level of project_dir — pre-processing/ holds
    intermediate copies that would create false duplicates.
    """
    pairs = []
    for proofread in sorted(project_dir.glob("Proofread_*.xlsx")) + sorted(project_dir.glob("Final_*.xlsx")):
        prefix = "Proofread_" if proofread.name.startswith("Proofread_") else "Final_"
        base = proofread.name[len(prefix):]
        candidates = [
            project_dir / f"Translated_{base}",
            project_dir / base,
            project_dir / f"{Path(base).stem}_translated.xlsx",
        ]
        translated = next((c for c in candidates if c.exists()), None)
        if translated is None:
            print(f"  WARNING: no translated counterpart found for {proofread.name} — skipping.")
            continue
        doc_label = Path(base).stem
        pairs.append((doc_label, translated, proofread))
    return pairs


# ── word-level diff → rich text ──────────────────────────────────────────────

def _tokenize(text: str) -> list[str]:
    return re.findall(r"\S+|\s+", text or "")


def diff_rich_text(old: str, new: str) -> CellRichText:
    """xlsx rich text for the human-facing sheet: strikethrough=deleted, underline=inserted.

    Note: openpyxl 3.1.5 writes this correctly (verified via raw XML — proper
    multi-run inlineStr, renders fine in Excel/LibreOffice) but does NOT
    reliably reconstruct CellRichText when *reading a saved file back* — a
    round-trip re-read collapses runs to a single plain string. That's an
    openpyxl read-side limitation, not a write bug — don't rely on
    openpyxl to re-read this column; use diff_plain_text()/the CSV export
    for anything that needs to inspect what changed.
    """
    import difflib

    old_tok, new_tok = _tokenize(old), _tokenize(new)
    sm = difflib.SequenceMatcher(a=old_tok, b=new_tok, autojunk=False)
    parts: list = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            parts.append("".join(new_tok[j1:j2]))
        elif tag == "delete":
            parts.append(TextBlock(FONT_DEL, "".join(old_tok[i1:i2])))
        elif tag == "insert":
            parts.append(TextBlock(FONT_INS, "".join(new_tok[j1:j2])))
        elif tag == "replace":
            parts.append(TextBlock(FONT_DEL, "".join(old_tok[i1:i2])))
            parts.append(TextBlock(FONT_INS, "".join(new_tok[j1:j2])))
    return CellRichText(*parts)


def diff_plain_text(old: str, new: str) -> str:
    """Plain-text diff markup for the CSV: [-deleted-]{+inserted+}, readable by Claude."""
    import difflib

    old_tok, new_tok = _tokenize(old), _tokenize(new)
    sm = difflib.SequenceMatcher(a=old_tok, b=new_tok, autojunk=False)
    out = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            out.append("".join(new_tok[j1:j2]))
        elif tag == "delete":
            out.append(f"[-{''.join(old_tok[i1:i2])}-]")
        elif tag == "insert":
            out.append(f"{{+{''.join(new_tok[j1:j2])}+}}")
        elif tag == "replace":
            out.append(f"[-{''.join(old_tok[i1:i2])}-]")
            out.append(f"{{+{''.join(new_tok[j1:j2])}+}}")
    return "".join(out)


# ── segment-aligned diff of one xlsx pair ────────────────────────────────────

def diff_pair(doc_label: str, translated_path: Path, proofread_path: Path) -> list[dict]:
    wb_t = openpyxl.load_workbook(translated_path, data_only=True)
    wb_p = openpyxl.load_workbook(proofread_path, data_only=True)
    ws_t, ws_p = wb_t.active, wb_p.active

    p_by_id = {}
    for row in ws_p.iter_rows(min_row=1):
        seg_id = row[0].value
        if isinstance(seg_id, int):
            p_by_id[seg_id] = row

    changes = []
    for row in ws_t.iter_rows(min_row=1):
        seg_id = row[0].value
        if not isinstance(seg_id, int):
            continue
        p_row = p_by_id.get(seg_id)
        if p_row is None:
            continue
        source = row[1].value or ""
        target_t = row[2].value or ""
        target_p = p_row[2].value or ""
        if target_t != target_p:
            changes.append({
                "doc": doc_label,
                "id": seg_id,
                "source": source,
                "translated_target": target_t,
                "corrected_target": diff_rich_text(target_t, target_p),
                "corrected_target_plain": diff_plain_text(target_t, target_p),
            })
    return changes


# ── XTRF task-detail lookup ──────────────────────────────────────────────────

def fetch_xtrf_task_details(job_id: str) -> dict:
    import requests
    from xtrf_job_setup import _load_creds, _login, _get_job, BASE_URL  # noqa: E402

    creds = _load_creds()
    session = requests.Session()
    session.headers.update({
        "Accept": "application/json, text/plain",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:149.0) Gecko/20100101 Firefox/149.0",
    })
    _login(session, creds)
    job = _get_job(session, job_id)
    overview = job.get("overview", {})

    result = {}
    result["project_number"] = extract_project_id(overview.get("projectName", ""))

    weighted = overview.get("jobQuantities", {}).get("weightedQuantities", [])
    for q in weighted:
        if q.get("unit") == "source word":
            result["weighted_word_count"] = round(q["value"])
        elif q.get("unit") == "1h":
            result["time_spent_minutes"] = round(q["value"] * 60)

    m = re.search(
        r"Content Specialisation:.*?</strong>\s*([^<]+)",
        job.get("instructions", ""),
    )
    if m:
        result["content_specialty"] = m.group(1).strip()

    expected = {
        "weighted_word_count": "E8 (weighted word count) — no 'source word' entry in jobQuantities.weightedQuantities",
        "time_spent_minutes": "E12 (time spent) — no '1h' entry in jobQuantities.weightedQuantities",
        "content_specialty": "E9 (content specialty) — 'Content Specialisation:' not found in job instructions",
    }
    for key, msg in expected.items():
        if key not in result:
            print(f"  WARNING: XTRF response missing expected field for {msg}. Leaving blank for manual entry.")

    return result


# ── scorecard file discovery + writing ───────────────────────────────────────

def find_scorecard(project_dir: Path) -> Path | None:
    matches = [
        p for p in project_dir.rglob("Translator Scorecard*.xlsx")
        if "pre-processing" not in p.parts
    ]
    if not matches:
        return None
    if len(matches) > 1:
        print(f"  WARNING: multiple scorecard files found, using {matches[0]}: {matches}")
    return matches[0]


def ensure_scorecard(project_dir: Path, project_number: str) -> Path:
    """Find the project's scorecard, or create one from the blank template
    (colocated with this skill so it survives redeploy — see SKILL.md) if
    none exists yet."""
    existing = find_scorecard(project_dir)
    if existing is not None:
        lock_file = existing.parent / f".~lock.{existing.name}#"
        if lock_file.exists():
            print(f"  WARNING: {existing.name} appears to be open in LibreOffice/Word right now "
                  "(lock file present). Close it before running this script, or edits may be lost.")
        return existing

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"No scorecard found under {project_dir}, and the blank template is missing at {TEMPLATE_PATH}"
        )
    dest = project_dir / f"Translator Scorecard 5.14.26_{project_number}.xlsx"
    shutil.copy(TEMPLATE_PATH, dest)
    print(f"  No scorecard found under {project_dir} — created one from the template: {dest.name}")
    return dest


def add_tracked_changes_sheet(wb: openpyxl.Workbook, changes: list[dict]) -> None:
    if "Tracked Changes" in wb.sheetnames:
        del wb["Tracked Changes"]
    ws = wb.create_sheet("Tracked Changes")
    headers = ["Document", "Id", "Source", "Translated Target", "Corrected Target (tracked)"]
    ws.append(headers)
    for col, width in zip("ABCDE", (30, 6, 45, 45, 45)):
        ws.column_dimensions[col].width = width
    for row in changes:
        ws.append([
            row["doc"], row["id"], row["source"],
            row["translated_target"], row["corrected_target"],
        ])
    for r in range(2, ws.max_row + 1):
        for c in range(1, 6):
            ws.cell(row=r, column=c).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")


def write_tracked_changes_csv(changes: list[dict], path: Path) -> None:
    """Plain-text sibling of the xlsx sheet — for the SKILL.md reasoning stage
    to read with the Read tool (never re-read the xlsx rich text column, see
    diff_rich_text's docstring)."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["doc", "id", "source", "translated_target", "corrected_target_plain"])
        writer.writeheader()
        for row in changes:
            writer.writerow({k: row[k] for k in writer.fieldnames})


def fill_task_details(wb: openpyxl.Workbook, values: dict) -> None:
    ws = wb["Scorecard"]
    for key, cell in TASK_DETAIL_CELLS.items():
        if key in values:
            ws[cell] = values[key]


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("project_dir", type=Path)
    ap.add_argument("--job-id", default=None, help="XTRF job id, e.g. 373610")
    ap.add_argument("--output", type=Path, default=None,
                     help="Where to save the filled scorecard (default: overwrite in place)")
    args = ap.parse_args()

    project_dir = args.project_dir.resolve()
    pairs = find_xlsx_pairs(project_dir)
    if not pairs:
        print(f"No Translated/Proofread xlsx pairs found in {project_dir}. Nothing to do.")
        return

    all_changes = []
    for doc_label, translated_path, proofread_path in pairs:
        changes = diff_pair(doc_label, translated_path, proofread_path)
        print(f"  {doc_label}: {len(changes)} changed segments "
              f"({translated_path.name} vs {proofread_path.name})")
        all_changes.extend(changes)

    values = {
        "copy_editor": "US",
        "target_locale": "de-DE",
        "source_locale": "en-US",
        "date": date.today().strftime("%d/%m/%y"),
        "project_number": project_dir.name,
        "agency_or_fl": "FL",
    }
    if args.job_id:
        try:
            values.update(fetch_xtrf_task_details(args.job_id))
        except Exception as e:
            print(f"  WARNING: XTRF lookup failed ({e}); leaving those fields for manual entry.")

    scorecard_path = ensure_scorecard(project_dir, values["project_number"])
    wb = openpyxl.load_workbook(scorecard_path)
    add_tracked_changes_sheet(wb, all_changes)
    fill_task_details(wb, values)

    out_path = args.output or scorecard_path
    wb.save(out_path)

    csv_path = out_path.parent / "tracked_changes.csv"
    write_tracked_changes_csv(all_changes, csv_path)

    print(f"\nSaved: {out_path}")
    print(f"Diff CSV for the reasoning stage: {csv_path}")
    print(f"Total changed segments across {len(pairs)} document(s): {len(all_changes)}")
    print("Part Two (Errors) and Part Three (Overall Summary) still need the reasoning stage.")


if __name__ == "__main__":
    main()
