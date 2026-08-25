"""
finalize_scorecard.py — reasoning-stage write for the proofreading-scorecard skill.

Writes Part Two (Errors) and Part Three (Overall Summary) into the project's
scorecard, given Claude's classifications as a JSON file.

Why this exists instead of an inline openpyxl snippet: openpyxl reliably
WRITES CellRichText (strikethrough/underline runs) but does NOT reconstruct
it when READING a saved file back — a plain load_workbook()-then-save()
cycle silently collapses the Tracked Changes sheet's rich text to flat
strings, even though nothing about that sheet was touched. Confirmed live
2026-08-25: opening the build_scorecard.py output in real Excel showed
correct strikethrough/underline; after a script filled Part Two/Three via
load+save, the same cells were plain text in Excel.

The fix: never resave a workbook that was merely loaded with rich text
already in it. Instead, re-derive the diff from the source xlsx pair (cheap
and deterministic) and rewrite the Tracked Changes sheet fresh in the same
save that writes Part Two/Three, so the rich text is always generated
immediately before the one save call that persists it.

Usage:
    python finalize_scorecard.py <project_dir> <part_two_three.json>

part_two_three.json shape:
{
  "errors": [
    {"phrase": "...", "severity": "Major", "category": "Accuracy"},
    ...  (up to 5, in Title -> Claims -> Abstract -> Specifications order)
  ],
  "pass_fail": "Pass",
  "comments": "..."
}
"""

import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from build_scorecard import find_xlsx_pairs, diff_pair, find_scorecard, add_tracked_changes_sheet  # noqa: E402


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    project_dir = Path(sys.argv[1]).resolve()
    data = json.loads(Path(sys.argv[2]).read_text(encoding="utf-8"))

    errors = data.get("errors", [])
    if len(errors) > 5:
        raise ValueError(f"Template has exactly 5 error slots, got {len(errors)}.")

    scorecard_path = find_scorecard(project_dir)
    if scorecard_path is None:
        raise FileNotFoundError(
            f"No scorecard found under {project_dir} — run build_scorecard.py first."
        )

    pairs = find_xlsx_pairs(project_dir)
    if not pairs:
        raise FileNotFoundError(f"No Translated/Proofread xlsx pairs found in {project_dir}.")
    all_changes = []
    for doc_label, translated_path, proofread_path in pairs:
        all_changes.extend(diff_pair(doc_label, translated_path, proofread_path))

    wb = openpyxl.load_workbook(scorecard_path)

    # Rebuild Tracked Changes fresh from the source files, in this same save —
    # never rely on rich text surviving the load above.
    add_tracked_changes_sheet(wb, all_changes)

    ws = wb["Scorecard"]
    for i, err in enumerate(errors):
        base = 17 + 3 * i
        ws[f"F{base}"] = err["phrase"]
        ws[f"F{base + 1}"] = err["severity"]
        ws[f"F{base + 2}"] = err["category"]

    if "pass_fail" in data:
        ws["F33"] = data["pass_fail"]
    if "comments" in data:
        ws["F34"] = data["comments"]

    wb.save(scorecard_path)
    print(f"Saved Part Two/Three and refreshed Tracked Changes rich text: {scorecard_path}")


if __name__ == "__main__":
    main()
