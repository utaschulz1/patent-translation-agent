# ============================================================
# LLM_capability_comparison_xlsx.py
# ============================================================
# Uses an LLM to extract EN→DE capability-predicate pairs from
# bilingual patent segments.  Targets constructions of the form
# "is/are [past-participle] to/for …" that describe what a
# component is capable of or designed to do — e.g.:
#   "is configured to", "is adapted for", "is arranged to",
#   "is designed to", "is operable to", "is shaped to"
# These escape both the verb extractor (no active verbal pair)
# and the noun extractor (not a noun phrase), so they need their
# own pass.
#
# INPUT   projects/<project_id>/*_translated.xlsx
# OUTPUT  projects/<project_id>/capability_segment_pairs.csv
#                               capability_canonical_glossary.csv
#                               capability_flags.csv
#                               <name>_checks.xlsx  (adds column)
# ============================================================

import os
import sys
import glob
import json
import pandas as pd
import openpyxl
import spacy
from openai import OpenAI
from collections import defaultdict, Counter
from dotenv import load_dotenv
from datetime import datetime

from project_log import project_dir as _pdir

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

try:
    nlp_en = spacy.load("en_core_web_sm")
    nlp_de = spacy.load("de_core_news_sm")
    SPACY_AVAILABLE = True
except OSError:
    print("WARNING: spaCy models not found — lemmatization skipped. "
          "Run: python -m spacy download en_core_web_sm de_core_news_sm")
    SPACY_AVAILABLE = False


def _lemmatize(word, nlp):
    """Return lowercased lemma of a single-token word.
    Rejects a lemma longer than the original — sign of spaCy misanalysis on unknown forms."""
    word = word.strip()
    if not word:
        return word
    token = nlp(word)[0]
    lemma = token.lemma_
    return (lemma if len(lemma) <= len(word) else word).lower()


BATCH_SIZE = 15
MODEL = "deepseek/deepseek-chat-v3-0324"

# ============================================================
# STEP 0 — Load the Excel file
# ============================================================

proj_dir = _pdir()
xlsx_files = [f for f in glob.glob(str(proj_dir / "*_translated.xlsx")) if not os.path.basename(f).startswith("~$")]
if not xlsx_files:
    print(f"ERROR: No *_translated.xlsx file found in '{proj_dir}'. Run ipappify_translate.py first.")
    exit()
if len(xlsx_files) > 1:
    print(f"Multiple translated files found, using: {xlsx_files[0]}")
input_path = xlsx_files[0]

try:
    raw_df = pd.read_excel(input_path, header=None, engine="openpyxl")
except Exception as e:
    print(f"ERROR: Could not read Excel file: {e}")
    exit()

print(f"Processing file: {raw_df.iloc[0, 0]}")

data_df = raw_df.iloc[3:].reset_index(drop=True)
data_df.columns = ["ID", "Source", "Target"] + list(data_df.columns[3:])
data_df = data_df[["ID", "Source", "Target"]].copy()
data_df.dropna(subset=["Source"], inplace=True)
data_df["Target"] = data_df["Target"].fillna("")
print(f"Loaded {len(data_df)} segments.")

# ============================================================
# PHASE 1 — Extract capability-predicate pairs via batched LLM calls
# ============================================================

def build_batch_prompt(segments):
    lines = []
    for seg in segments:
        src = str(seg["source"]).replace("\n", " ").strip()
        tgt = str(seg["target"]).replace("\n", " ").strip() or "(no translation yet)"
        lines.append(f'[{seg["id"]}] EN: {src}\n[{seg["id"]}] DE: {tgt}')
    segment_block = "\n\n".join(lines)

    return (
        "You are a bilingual EN-DE patent translation analyst.\n"
        "For each numbered segment pair, find all capability predicates in the EN text.\n\n"
        "A capability predicate describes what a component IS CAPABLE OF or IS DESIGNED TO DO. "
        "It has the structure: [noun] + is/are/was/were + [past participle] + to/for\n\n"
        "Examples of capability predicates:\n"
        "  'is configured to'  →  en: configure\n"
        "  'is adapted for'    →  en: adapt\n"
        "  'is arranged to'    →  en: arrange\n"
        "  'is designed to'    →  en: design\n"
        "  'is operable to'    →  en: operate\n"
        "  'is shaped to'      →  en: shape\n"
        "  'adapted for supporting'  →  en: adapt\n\n"
        "Do NOT include:\n"
        "  - Pure descriptive adjectives without to/for: 'a bent rod', 'a connected device'\n"
        "  - Active verbs already handled separately: 'the device connects', 'it receives'\n\n"
        "For each capability predicate found, also find its German translation equivalent "
        "in the DE text and extract the German BASE INFINITIVE (not the inflected form).\n\n"
        "Return ONLY a JSON array — no markdown, no explanation. Each element:\n"
        '  {"id": "<segment_id>", "pairs": [{"en": "<infinitive>", "de": "<infinitive>"}, ...]}\n\n'
        "Rules:\n"
        "  - Use base/infinitive forms only (configure not configured; konfigurieren not konfiguriert)\n"
        "  - Include only pairs where you can identify a clear German equivalent in the DE text\n"
        "  - If a segment has no capability predicates, return an empty pairs array\n"
        "  - Return ONLY the raw JSON array\n\n"
        f"Segments:\n\n{segment_block}"
    )


api_key = os.environ.get("OPENROUTER_API_KEY")
if not api_key:
    print("ERROR: OPENROUTER_API_KEY not found in .env file.")
    exit()

client = OpenAI(api_key=api_key, base_url="https://openrouter.ai/api/v1")

segments = [
    {"id": str(row["ID"]), "source": row["Source"], "target": row["Target"]}
    for _, row in data_df.iterrows()
]

all_pairs = []

_args = sys.argv[1:]
try:
    seg_start = int(_args[0]) if len(_args) >= 1 else 1
    seg_end   = int(_args[1]) if len(_args) >= 2 else min(100, len(segments))
except ValueError:
    print(f"ERROR: invalid arguments {_args!r}.")
    print(f"  Usage: python {os.path.basename(__file__)} [start] [end]")
    print(f"  Example: python {os.path.basename(__file__)} 1 100")
    print(f"  File has {len(segments)} segments. Arguments are 1-based and inclusive.")
    exit(1)
if seg_start < 1 or seg_end < seg_start or seg_end > len(segments):
    print(f"ERROR: segment range {seg_start}–{seg_end} is out of bounds (file has {len(segments)} segments).")
    print(f"  Usage: python {os.path.basename(__file__)} [start] [end]")
    print(f"  Example: python {os.path.basename(__file__)} 1 100")
    exit(1)
segments = segments[seg_start - 1:seg_end]
print(f"Segments {seg_start}–{seg_end} selected ({len(segments)} segments, total in file: {len(data_df)}).")

batches = [segments[i:i + BATCH_SIZE] for i in range(0, len(segments), BATCH_SIZE)]
print(f"\nPhase 1: extracting capability predicate pairs — {len(batches)} batches of up to {BATCH_SIZE} segments...")

for batch_num, batch in enumerate(batches, 1):
    prompt = build_batch_prompt(batch)
    try:
        response = client.chat.completions.create(
            model=MODEL,
            max_tokens=2048,
            temperature=0,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.choices[0].message.content.strip()

        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1]
            raw = raw.rsplit("```", 1)[0].strip()

        parsed = json.loads(raw)

        seg_text = {str(s["id"]): s for s in batch}
        batch_count = 0
        for seg_result in parsed:
            seg_id = str(seg_result.get("id", ""))
            for pair in seg_result.get("pairs", []):
                en = pair.get("en", "").strip().lower()
                de = pair.get("de", "").strip().lower()
                if en and de:
                    all_pairs.append({
                        "segment_id":  seg_id,
                        "en_verb":     en,
                        "de_verb":     de,
                        "source_text": seg_text.get(seg_id, {}).get("source", ""),
                        "target_text": seg_text.get(seg_id, {}).get("target", ""),
                    })
                    batch_count += 1

        print(f"  Batch {batch_num}/{len(batches)}: {batch_count} pairs extracted.")

    except json.JSONDecodeError as e:
        print(f"  Batch {batch_num}: JSON parse error — {e}")
        print(f"    Raw response (first 300 chars): {raw[:300]}")
    except Exception as e:
        print(f"  Batch {batch_num}: API error — {e}")

print(f"\nTotal capability predicate pairs extracted across all segments: {len(all_pairs)}")

if not all_pairs:
    print("No capability predicate pairs extracted. Check the API connection and model response above.")
    exit()

# ============================================================
# PHASE 1b — Lemmatize extracted verb forms
# ============================================================

if SPACY_AVAILABLE:
    print("\nPhase 1b: lemmatizing extracted verb forms...")
    for p in all_pairs:
        p["en_verb"] = _lemmatize(p["en_verb"], nlp_en)
        p["de_verb"] = _lemmatize(p["de_verb"], nlp_de)
    print(f"  Done ({len(all_pairs)} pairs normalized).")
else:
    print("\nPhase 1b: skipped (spaCy models unavailable).")

pairs_csv = str(proj_dir / "capability_segment_pairs.csv")
current_ids = {s["id"] for s in segments}
if os.path.exists(pairs_csv):
    existing_df = pd.read_csv(pairs_csv, encoding="utf-8-sig")
    kept = existing_df[~existing_df["segment_id"].astype(str).isin(current_ids)]
    print(f"Merging {len(kept)} existing pairs + {len(all_pairs)} new pairs.")
    all_pairs = kept.to_dict("records") + all_pairs
else:
    print(f"No existing pairs file — starting fresh.")

pairs_df = pd.DataFrame(all_pairs)
pairs_df.to_csv(pairs_csv, index=False, encoding="utf-8-sig")
print('Saved "capability_segment_pairs.csv".')

# ============================================================
# PHASE 2 — Build canonical glossary (majority-vote per EN verb)
# ============================================================

en_to_de = defaultdict(Counter)
for p in all_pairs:
    en_to_de[p["en_verb"]][p["de_verb"]] += 1

canonical = {}
glossary_rows = []

for en_verb, de_counter in sorted(en_to_de.items()):
    total = sum(de_counter.values())
    canonical_de, canonical_count = de_counter.most_common(1)[0]
    canonical[en_verb] = canonical_de
    for de_verb, count in de_counter.most_common():
        glossary_rows.append({
            "EN Verb":              en_verb,
            "DE Verb":              de_verb,
            "Count":                count,
            "Total EN Occurrences": total,
            "Canonical":            "yes" if de_verb == canonical_de else "no",
        })

glossary_df = pd.DataFrame(glossary_rows)
glossary_df.to_csv(str(proj_dir / "capability_canonical_glossary.csv"), index=False, encoding="utf-8-sig")
print('\nSaved "capability_canonical_glossary.csv".')

print("\nCanonical glossary (most common translation per EN capability verb):")
for en_verb, de_verb in sorted(canonical.items()):
    count = en_to_de[en_verb][de_verb]
    total = sum(en_to_de[en_verb].values())
    alternatives = [
        f"{d}({c})" for d, c in en_to_de[en_verb].most_common() if d != de_verb
    ]
    alt_str = f"  alternatives: {', '.join(alternatives)}" if alternatives else ""
    print(f"  {en_verb} → {de_verb}  ({count}/{total}){alt_str}")

# ============================================================
# PHASE 3 — Flag segments from the canonical glossary
# ============================================================

DOMINANCE_THRESHOLD = 0.60

flags = []

print("\nPhase 3 – Case 1: checking for inconsistent capability-predicate translations...")
case1_count = 0

for en_verb, de_counter in en_to_de.items():
    if len(de_counter) == 1:
        continue

    total = sum(de_counter.values())
    canonical_de, canonical_count = de_counter.most_common(1)[0]
    ratio = canonical_count / total
    clear_winner = ratio >= DOMINANCE_THRESHOLD

    for p in all_pairs:
        if p["en_verb"] != en_verb:
            continue
        if clear_winner and p["de_verb"] == canonical_de:
            continue

        alts = ", ".join(
            f"{d}({c})" for d, c in de_counter.most_common() if d != canonical_de
        )
        if clear_winner:
            note = (
                f"[Cap 1] {en_verb}: minority translation \"{p['de_verb']}\" "
                f"(canonical: {canonical_de} {canonical_count}/{total}; also seen: {alts})"
            )
        else:
            note = (
                f"[Cap 1] {en_verb}: no dominant translation "
                f"({', '.join(f'{d}({c})' for d, c in de_counter.most_common())})"
            )

        flags.append({
            "Segment ID":  p["segment_id"],
            "Flag Type":   "Cap 1 – minority" if clear_winner else "Cap 1 – no winner",
            "Note":        note,
            "Source Text": p["source_text"],
            "Target Text": p["target_text"],
        })
        case1_count += 1

print(f"  Case 1 flags: {case1_count}")

print("Phase 3 – Case 2: checking for same DE verb used for multiple EN capability verbs...")
case2_count = 0

de_to_en_sources = defaultdict(list)
for en_verb, de_verb in canonical.items():
    de_to_en_sources[de_verb].append(en_verb)

shared_de_verbs = {de: en_list for de, en_list in de_to_en_sources.items() if len(en_list) > 1}

for p in all_pairs:
    if p["de_verb"] not in shared_de_verbs:
        continue
    en_sources = shared_de_verbs[p["de_verb"]]
    note = (
        f"[Cap 2] \"{p['de_verb']}\": same translation for different EN capability verbs: "
        + ", ".join(en_sources)
    )
    flags.append({
        "Segment ID":  p["segment_id"],
        "Flag Type":   "Cap 2 – shared DE verb",
        "Note":        note,
        "Source Text": p["source_text"],
        "Target Text": p["target_text"],
    })
    case2_count += 1

print(f"  Case 2 flags: {case2_count}")
print(f"\nTotal flags: {len(flags)}")

flags_df = pd.DataFrame(flags)
flags_df.to_csv(str(proj_dir / "capability_flags.csv"), index=False, encoding="utf-8-sig")
print('Saved "capability_flags.csv".')

if not flags_df.empty:
    print(flags_df[["Segment ID", "Flag Type", "Note"]].to_string(index=False))

# ============================================================
# PHASE 4 — Annotate the original Excel file
# ============================================================

seg_annotations = defaultdict(list)
for f in flags:
    seg_annotations[str(f["Segment ID"])].append(f["Note"])

out_path = input_path.replace(".xlsx", "_checks.xlsx")
wb = openpyxl.load_workbook(out_path if os.path.exists(out_path) else input_path)
ws = wb.active

# Find existing "Capability Flags" column or insert one after "Verb Flags" (col 4)
cap_col = next(
    (c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value == "Capability Flags"),
    None
)
if cap_col is None:
    ws.insert_cols(5)
    cap_col = 5
    ws.cell(row=1, column=cap_col).value = "Capability Flags"

for row_num in range(4, ws.max_row + 1):
    seg_id = str(ws.cell(row=row_num, column=1).value)
    ws.cell(row=row_num, column=cap_col).value = (
        "\n".join(seg_annotations[seg_id]) if seg_id in seg_annotations else None
    )
    ws.cell(row=row_num, column=cap_col).alignment = openpyxl.styles.Alignment(wrap_text=True)
try:
    wb.save(out_path)
    print(f'\nSaved annotated Excel: "{out_path}".')
except PermissionError:
    stamp = datetime.now().strftime("%H%M%S")
    out_path = input_path.replace(".xlsx", f"_checks_{stamp}.xlsx")
    wb.save(out_path)
    print(f'\nFile was open in Excel — saved as "{out_path}" instead.')
