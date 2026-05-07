# ============================================================
# LLM_noun_comparison_xlsx.py
# ============================================================
# Checks translation consistency of noun phrases across a bilingual
# patent Excel file (EN source / DE target).
#
# INPUT   projects/<project_id>/*_translated.xlsx
# OUTPUT  projects/<project_id>/noun_segment_pairs.csv
#                               noun_canonical_glossary.csv
#                               noun_inconsistency_table.csv
#                               <name>_checks.xlsx  (noun column appended)
# ============================================================

import os
import sys
import glob
import json
import pandas as pd
import openpyxl
import spacy
from openai import OpenAI
from collections import defaultdict, Counter
from dotenv import load_dotenv
from datetime import datetime

from project_log import project_dir as _pdir

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

try:
    nlp_en = spacy.load("en_core_web_sm")
    nlp_de = spacy.load("de_core_news_sm")
    SPACY_AVAILABLE = True
except OSError:
    print("WARNING: spaCy models not found — lemmatization skipped. "
          "Run: python -m spacy download en_core_web_sm de_core_news_sm")
    SPACY_AVAILABLE = False


def _lemmatize_phrase(phrase, nlp, keep_case=False):
    """Lemmatize each token in a phrase and rejoin. keep_case preserves German capitalisation.
    Rejects a lemma if it is longer than the original token — sign of spaCy misanalysis on
    unknown compounds."""
    phrase = phrase.strip()
    if not phrase:
        return phrase
    tokens = []
    for token in nlp(phrase):
        lemma = token.lemma_
        tokens.append(lemma if len(lemma) <= len(token.text) else token.text)
    result = " ".join(tokens)
    return result if keep_case else result.lower()


BATCH_SIZE = 10
MODEL = "deepseek/deepseek-chat-v3-0324"
RUN_EVALUATOR = False

# ============================================================
# STEP 0 — Load the Excel file
# ============================================================

proj_dir = _pdir()
source_files = [f for f in glob.glob(str(proj_dir / "*_translated.xlsx")) if not os.path.basename(f).startswith("~$")]
if not source_files:
    print(f"ERROR: No *_translated.xlsx file found in '{proj_dir}'. Run ipappify_translate.py first.")
    exit()
if len(source_files) > 1:
    print(f"Multiple translated files found, using: {source_files[0]}")
input_path = source_files[0]

try:
    raw_df = pd.read_excel(input_path, header=None, engine="openpyxl")
except Exception as e:
    print(f"ERROR: Could not read Excel file: {e}")
    exit()

print(f"Processing file: {raw_df.iloc[0, 0]}")

data_df = raw_df.iloc[3:].reset_index(drop=True)
data_df.columns = ["ID", "Source", "Target"] + list(data_df.columns[3:])
data_df = data_df[["ID", "Source", "Target"]].copy()
data_df.dropna(subset=["Source"], inplace=True)
data_df["Target"] = data_df["Target"].fillna("")
print(f"Loaded {len(data_df)} segments.")

# ============================================================
# PHASE 1 — Extract noun phrase pairs per segment (batched LLM calls)
# ============================================================

def build_batch_prompt(segments):
    lines = []
    for seg in segments:
        src = str(seg["source"]).replace("\n", " ").strip()
        tgt = str(seg["target"]).replace("\n", " ").strip() or "(no translation yet)"
        lines.append(f'[{seg["id"]}] EN: {src}\n[{seg["id"]}] DE: {tgt}')
    segment_block = "\n\n".join(lines)

    return (
        "You are a bilingual EN-DE patent translation analyst.\n"
        "For each numbered segment pair, identify technical noun phrases in the EN text "
        "and their German translation equivalents in the DE text.\n\n"
        "Extraction rules:\n"
        "  - Extract every noun phrase: technical terms, device names, method names, "
        "component names, parameter names, and general structural terms (e.g. 'method', 'device', 'system').\n"
        "  - Always take the LONGEST enclosing phrase per concept. "
        "If the text contains 'grain drying device', extract 'grain drying device', not 'device'.\n"
        "  - Do NOT list sub-phrases of a phrase you already extracted "
        "(e.g. if you extract 'base station antenna', skip 'base station' and 'antenna' separately).\n"
        "  - German noun phrases: include the full compound or phrase as written "
        "(e.g. 'Basisstationsantenne', 'drahtloses Kommunikationssystem'). "
        "Strip articles (der/die/das/ein) from the DE side.\n"
        "  - EN side: lowercase, no articles (a/an/the), singular preferred.\n"
        "  - DE side: keep German capitalisation, no articles, nominative singular preferred.\n"
        "  - Only include pairs where the DE equivalent is clearly present in the DE text.\n"
        "  - Skip pronouns, pure adjectives, and verb phrases.\n\n"
        "Return ONLY a JSON array — no markdown, no explanation. Each element:\n"
        '  {"id": "<segment_id>", "pairs": [{"en": "<noun phrase>", "de": "<Nominalphrase>"}, ...]}\n\n'
        "If a segment has no extractable noun phrase pairs, return an empty pairs array.\n\n"
        f"Segments:\n\n{segment_block}"
    )


api_key = os.environ.get("OPENROUTER_API_KEY")
if not api_key:
    print("ERROR: OPENROUTER_API_KEY not found in .env file.")
    exit()

client = OpenAI(api_key=api_key, base_url="https://openrouter.ai/api/v1")

segments = [
    {"id": str(row["ID"]), "source": row["Source"], "target": row["Target"]}
    for _, row in data_df.iterrows()
]

all_pairs = []

_args = sys.argv[1:]
try:
    seg_start = int(_args[0]) if len(_args) >= 1 else 1
    seg_end   = int(_args[1]) if len(_args) >= 2 else min(100, len(segments))
except ValueError:
    print(f"ERROR: invalid arguments {_args!r}.")
    print(f"  Usage: python {os.path.basename(__file__)} [start] [end]")
    print(f"  Example: python {os.path.basename(__file__)} 1 100")
    print(f"  File has {len(segments)} segments. Arguments are 1-based and inclusive.")
    exit(1)
if seg_start < 1 or seg_end < seg_start or seg_end > len(segments):
    print(f"ERROR: segment range {seg_start}–{seg_end} is out of bounds (file has {len(segments)} segments).")
    print(f"  Usage: python {os.path.basename(__file__)} [start] [end]")
    print(f"  Example: python {os.path.basename(__file__)} 1 100")
    exit(1)
segments  = segments[seg_start - 1:seg_end]
print(f"Segments {seg_start}–{seg_end} selected ({len(segments)} segments, total in file: {len(data_df)}).")

batches = [segments[i:i + BATCH_SIZE] for i in range(0, len(segments), BATCH_SIZE)]
print(f"\nPhase 1: extracting noun phrase pairs — {len(batches)} batches of up to {BATCH_SIZE} segments...")

for batch_num, batch in enumerate(batches, 1):
    prompt = build_batch_prompt(batch)
    try:
        response = client.chat.completions.create(
            model=MODEL,
            max_tokens=2048,
            temperature=0,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.choices[0].message.content.strip()

        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1]
            raw = raw.rsplit("```", 1)[0].strip()

        parsed = json.loads(raw)

        seg_text = {str(s["id"]): s for s in batch}
        batch_count = 0
        for seg_result in parsed:
            seg_id = str(seg_result.get("id", ""))
            for pair in seg_result.get("pairs", []):
                en = pair.get("en", "").strip().lower()
                de = pair.get("de", "").strip()
                if en and de:
                    all_pairs.append({
                        "segment_id":  seg_id,
                        "en_phrase":   en,
                        "de_phrase":   de,
                        "source_text": seg_text.get(seg_id, {}).get("source", ""),
                        "target_text": seg_text.get(seg_id, {}).get("target", ""),
                    })
                    batch_count += 1

        print(f"  Batch {batch_num}/{len(batches)}: {batch_count} pairs extracted.")

    except json.JSONDecodeError as e:
        print(f"  Batch {batch_num}: JSON parse error — {e}")
        print(f"    Raw response (first 300 chars): {raw[:300]}")
    except Exception as e:
        print(f"  Batch {batch_num}: API error — {e}")

print(f"\nTotal noun phrase pairs extracted: {len(all_pairs)}")

if not all_pairs:
    print("No noun phrase pairs extracted. Check API connection and model response above.")
    exit()

# ============================================================
# PHASE 1b — Lemmatize extracted noun phrase forms
# ============================================================

if SPACY_AVAILABLE:
    print("\nPhase 1b: lemmatizing extracted noun phrase forms...")
    for p in all_pairs:
        p["en_phrase"] = _lemmatize_phrase(p["en_phrase"], nlp_en, keep_case=False)
        p["de_phrase"] = _lemmatize_phrase(p["de_phrase"], nlp_de, keep_case=True)
    print(f"  Done ({len(all_pairs)} pairs normalized).")
else:
    print("\nPhase 1b: skipped (spaCy models unavailable).")

pairs_csv = str(proj_dir / "noun_segment_pairs.csv")
current_ids = {s["id"] for s in segments}
if os.path.exists(pairs_csv):
    existing_df = pd.read_csv(pairs_csv, encoding="utf-8-sig")
    kept = existing_df[~existing_df["segment_id"].astype(str).isin(current_ids)]
    print(f"Merging {len(kept)} existing pairs + {len(all_pairs)} new pairs.")
    all_pairs = kept.to_dict("records") + all_pairs
else:
    print(f"No existing pairs file — starting fresh.")

pairs_df = pd.DataFrame(all_pairs)
pairs_df.to_csv(pairs_csv, index=False, encoding="utf-8-sig")
print('Saved "noun_segment_pairs.csv".')

# ============================================================
# PHASE 2 — Build canonical glossary (majority-vote per EN phrase)
# ============================================================

en_to_de = defaultdict(Counter)
for p in all_pairs:
    en_to_de[p["en_phrase"]][p["de_phrase"]] += 1

canonical = {}
glossary_rows = []

for en_phrase, de_counter in sorted(en_to_de.items()):
    total = sum(de_counter.values())
    if total < 2:
        continue
    canonical_de, canonical_count = de_counter.most_common(1)[0]
    canonical[en_phrase] = canonical_de
    for de_phrase, count in de_counter.most_common():
        glossary_rows.append({
            "EN Phrase":            en_phrase,
            "DE Phrase":            de_phrase,
            "Count":                count,
            "Total EN Occurrences": total,
            "Canonical":            "yes" if de_phrase == canonical_de else "no",
        })

glossary_df = pd.DataFrame(glossary_rows)
glossary_df.to_csv(str(proj_dir / "noun_canonical_glossary.csv"), index=False, encoding="utf-8-sig")
print('\nSaved "noun_canonical_glossary.csv".')

print(f"\nCanonical glossary ({len(canonical)} phrases with ≥2 occurrences):")
for en_phrase, de_phrase in sorted(canonical.items()):
    count = en_to_de[en_phrase][de_phrase]
    total = sum(en_to_de[en_phrase].values())
    alts = [f"{d}({c})" for d, c in en_to_de[en_phrase].most_common() if d != de_phrase]
    alt_str = f"  alternatives: {', '.join(alts)}" if alts else ""
    print(f"  {en_phrase} → {de_phrase}  ({count}/{total}){alt_str}")

# ============================================================
# PHASE 3 — Inconsistency check
# ============================================================

seg_pairs_by_id = defaultdict(list)
for p in all_pairs:
    seg_pairs_by_id[p["segment_id"]].append(p)

inconsistencies = []
for seg_id, pairs in seg_pairs_by_id.items():
    for pair in pairs:
        en = pair["en_phrase"]
        de_actual = pair["de_phrase"]
        de_expected = canonical.get(en)
        if de_expected and de_actual != de_expected:
            inconsistencies.append({
                "Segment ID":        seg_id,
                "EN Phrase":         en,
                "Expected DE":       de_expected,
                "Actual DE":         de_actual,
                "Expected Count":    en_to_de[en][de_expected],
                "Actual Count":      en_to_de[en][de_actual],
                "Total Occurrences": sum(en_to_de[en].values()),
                "Source Text":       pair["source_text"],
                "Target Text":       pair["target_text"],
            })

incons_df = pd.DataFrame(inconsistencies)
print(f"\nInconsistencies found: {len(inconsistencies)}")
if not incons_df.empty:
    display_cols = ["Segment ID", "EN Phrase", "Expected DE", "Actual DE",
                    "Expected Count", "Actual Count", "Total Occurrences"]
    print(incons_df[display_cols].to_string(index=False))
else:
    print("No inconsistencies found.")

incons_df.to_csv(str(proj_dir / "noun_inconsistency_table.csv"), index=False, encoding="utf-8-sig")
print('Saved "noun_inconsistency_table.csv".')

# ============================================================
# PHASE 5 — LLM evaluator: identify false positives
# ============================================================

evaluator_ran = False

if not RUN_EVALUATOR:
    print("\nPhase 5: skipped (RUN_EVALUATOR = False).")
elif not inconsistencies:
    print("\nPhase 5: No inconsistencies to evaluate — skipping evaluator.")
else:
    def build_evaluator_prompt(canonical, inconsistencies):
        glossary_lines = []
        for en_phrase, de_phrase in sorted(canonical.items()):
            total = sum(en_to_de[en_phrase].values())
            count = en_to_de[en_phrase][de_phrase]
            alts = [(d, c) for d, c in en_to_de[en_phrase].most_common() if d != de_phrase]
            alt_str = (", alternatives: " + ", ".join(f"{d}({c}x)" for d, c in alts)) if alts else ""
            glossary_lines.append(f"  {en_phrase} → {de_phrase} ({count}/{total}x){alt_str}")
        glossary_block = "\n".join(glossary_lines)

        entry_lines = []
        for i, inc in enumerate(inconsistencies, 1):
            entry_lines.append(
                f"[{i}] Segment {inc['Segment ID']} | EN: \"{inc['EN Phrase']}\" "
                f"| canonical DE: \"{inc['Expected DE']}\" | found DE: \"{inc['Actual DE']}\"\n"
                f"    EN: {str(inc['Source Text'])[:200]}\n"
                f"    DE: {str(inc['Target Text'])[:200]}"
            )
        entries_block = "\n\n".join(entry_lines)

        return (
            "You are a bilingual EN-DE patent translation quality evaluator.\n\n"
            "CONTEXT\n"
            "A script analysed a patent translation by extracting EN-DE noun phrase pairs from every segment. "
            "It determined a canonical (most frequent) DE translation per EN phrase across the document. "
            "Any segment using a different DE phrase was flagged as an inconsistency.\n\n"
            "CANONICAL NOUN PHRASE GLOSSARY:\n"
            f"{glossary_block}\n\n"
            "FLAGGED INCONSISTENCIES:\n"
            "Each entry shows the segment text, the EN phrase, the canonical DE translation, "
            "and what was actually found in that segment.\n\n"
            f"{entries_block}\n\n"
            "YOUR TASK\n"
            "For each flagged entry decide: TRUE inconsistency (wrong or inconsistent translation) "
            "or FALSE POSITIVE (acceptable despite differing from the canonical)?\n\n"
            "Reasons a flag may be a false positive:\n"
            "  - The context genuinely requires a different form (e.g. plural, genitive construction)\n\n"
            "Return ONLY a JSON array — no markdown, no explanation. Each element:\n"
            '  {"index": <number>, "false_positive": true or false, "reason": "<one sentence>"}\n\n'
            "Cover every entry in the list."
        )

    print(f"\nPhase 5: sending {len(inconsistencies)} flagged entries to evaluator LLM...")
    eval_results = []

    try:
        response = client.chat.completions.create(
            model=MODEL,
            max_tokens=4096,
            temperature=0,
            messages=[{"role": "user", "content": build_evaluator_prompt(canonical, inconsistencies)}],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()

        eval_results = json.loads(raw)
        print(f"Evaluator returned judgements for {len(eval_results)} entries.")

    except json.JSONDecodeError as e:
        print(f"Evaluator JSON parse error — {e}. Raw (first 300 chars):\n{raw[:300]}")
    except Exception as e:
        print(f"Evaluator API error — {e}")

    if eval_results:
        index_map = {r["index"]: r for r in eval_results}
        for i, inc in enumerate(inconsistencies, 1):
            judgement = index_map.get(i, {})
            inc["False Positive"] = judgement.get("false_positive", None)
            inc["Reason"] = judgement.get("reason", "")

        incons_df = pd.DataFrame(inconsistencies)
        incons_df.to_csv(str(proj_dir / "noun_inconsistency_table.csv"), index=False, encoding="utf-8-sig")
        print('Updated "noun_inconsistency_table.csv" with evaluator judgements.')

        confirmed = incons_df[incons_df["False Positive"] == False]
        false_pos  = incons_df[incons_df["False Positive"] == True]
        print(f"\nConfirmed inconsistencies: {len(confirmed)}  |  False positives: {len(false_pos)}")
        if not confirmed.empty:
            print(confirmed[["Segment ID", "EN Phrase", "Expected DE", "Actual DE", "Reason"]].to_string(index=False))

        evaluator_ran = True

# ============================================================
# PHASE 4 — Annotate the _checks.xlsx file
# ============================================================

if evaluator_ran:
    to_annotate = confirmed.to_dict("records")
    label = "confirmed inconsistencies"
else:
    to_annotate = inconsistencies
    label = "all inconsistencies (evaluator did not run)"

seg_annotations = defaultdict(list)
for inc in to_annotate:
    seg_id = str(inc["Segment ID"])
    note = f"{inc['EN Phrase']}, canonical: {inc['Expected DE']}, here: {inc['Actual DE']}"
    seg_annotations[seg_id].append(note)

checks_path = input_path.replace(".xlsx", "_checks.xlsx")
wb_path = checks_path if os.path.exists(checks_path) else input_path

wb = openpyxl.load_workbook(wb_path)
ws = wb.active

noun_col = next(
    (c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value == "Noun Inconsistencies"),
    None
)
if noun_col is None:
    noun_col = ws.max_column + 1
    ws.cell(row=1, column=noun_col).value = "Noun Inconsistencies"
insert_col = noun_col

for row_num in range(4, ws.max_row + 1):
    seg_id = str(ws.cell(row=row_num, column=1).value)
    if seg_id in seg_annotations:
        ws.cell(row=row_num, column=insert_col).value = "\n".join(seg_annotations[seg_id])
        ws.cell(row=row_num, column=insert_col).alignment = openpyxl.styles.Alignment(wrap_text=True)

try:
    wb.save(checks_path)
    print(f'\nSaved annotated Excel ({label}): "{checks_path}".')
except PermissionError:
    stamp = datetime.now().strftime("%H%M%S")
    fallback = checks_path.replace("_checks.xlsx", f"_checks_{stamp}.xlsx")
    wb.save(fallback)
    print(f'\nCould not overwrite — saved as "{fallback}" instead.')
