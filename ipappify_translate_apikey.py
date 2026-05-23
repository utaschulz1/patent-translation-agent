# ============================================================
# ipappify_translate_apikey.py
# ============================================================
# Like ipappify_translate.py but authenticates with a static API key (Authorization: ApiKey ...) instead of OAuth2 Bearer tokens. Set IPAPPIFY_API_KEY in .env.
#
# If a clean_glossary_*.csv is present in the project folder it is loaded and injected per segment via the Dictionary field.
# Correct field names: SourceText / TargetText / IsLiteral.
#
# INPUT   projects/<project_id>/<any>.xlsx   — bilingual Excel from XTM
#         projects/<project_id>/clean_glossary_*.csv  (optional)
# OUTPUT  projects/<project_id>/<name>_iptranslated.xlsx
# ============================================================

import os
import sys
import glob
from pathlib import Path
import json
import base64
import time
import uuid
import argparse
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import pandas as pd
import openpyxl
from datetime import datetime, timezone
from dotenv import load_dotenv

from project_log import project_dir as _pdir

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

API_URL  = "https://iptranslator.ipappify.de/translate"
MODEL    = "EN->DE v03"
SRC_LANG = "en"
TGT_LANG = "de"
CONTEXT_N = 1

api_key = os.environ.get("IPAPPIFY_API_KEY", "").strip()
if not api_key:
    print("ERROR: IPAPPIFY_API_KEY not set in .env.")
    exit()

parser = argparse.ArgumentParser()
parser.add_argument("--file", help="Path to the bilingual Excel file (overrides auto-discovery)")
args = parser.parse_args()

# ============================================================
# Load Excel
# ============================================================

if args.file:
    input_path = os.path.abspath(args.file)
    if not os.path.isfile(input_path):
        print(f"ERROR: File not found: {input_path}")
        exit()
    proj_dir = Path(os.path.dirname(input_path))
else:
    proj_dir = _pdir()
    xlsx_files = glob.glob(str(proj_dir / "*.xlsx"))
    xlsx_files = [f for f in xlsx_files if "_translated" not in f and "_iptranslated" not in f and "_checks" not in f and not os.path.basename(f).startswith("~$") and not os.path.basename(f).startswith("clean_glossary_")]
    if not xlsx_files:
        print(f"ERROR: No .xlsx file found in '{proj_dir}'.")
        exit()
    if len(xlsx_files) > 1:
        print(f"Multiple .xlsx files found, using: {xlsx_files[0]}")
    input_path = xlsx_files[0]

raw_df = pd.read_excel(input_path, header=None, engine="openpyxl")
print(f"Processing: {raw_df.iloc[0, 0]}")

data_df = raw_df.iloc[3:].reset_index(drop=True)
data_df.columns = ["ID", "Source", "Target"] + list(data_df.columns[3:])
data_df = data_df[["ID", "Source", "Target"]].copy()
data_df.dropna(subset=["Source"], inplace=True)
data_df["ID"]     = data_df["ID"].astype(str).str.strip()
data_df["Source"] = data_df["Source"].astype(str).str.strip()
data_df["Target"] = data_df["Target"].fillna("").astype(str).str.strip()
segments = data_df.to_dict("records")
print(f"Loaded {len(segments)} segments.")

# ============================================================
# Glossary (optional) — loaded from clean_glossary_*.csv
# ============================================================

glossary = []  # list of (en_term, de_term)
glossary_files = glob.glob(str(proj_dir / "clean_glossary_*.csv"))
if glossary_files:
    if len(glossary_files) > 1:
        print(f"Multiple glossary files found, using: {glossary_files[0]}")
    gloss_df = pd.read_csv(glossary_files[0], encoding="utf-8-sig")
    gloss_df.columns = [c.strip() for c in gloss_df.columns]
    try:
        en_col = next(c for c in gloss_df.columns if c.upper() in ("EN", "ENGLISH", "SOURCE"))
        de_col = next(c for c in gloss_df.columns if c.upper() in ("DE", "GERMAN", "TARGET"))
        glossary = [
            (str(r[en_col]).strip(), str(r[de_col]).strip())
            for _, r in gloss_df.iterrows()
            if pd.notna(r[en_col]) and pd.notna(r[de_col])
            and str(r[en_col]).strip() and str(r[de_col]).strip()
        ]
        print(f"Loaded {len(glossary)} glossary terms from '{glossary_files[0]}'.")
    except StopIteration:
        print(f"WARNING: Glossary has no EN/DE columns (found: {list(gloss_df.columns)}) — skipping.")
else:
    print("No glossary_*.csv found — translating without dictionary.")


def build_dictionary(source_text):
    """Return glossary entries whose EN term has at least one word in source_text."""
    src_lower = source_text.lower()
    return [
        {"SourceText": en, "TargetText": de, "IsLiteral": False}
        for en, de in glossary
        if any(w in src_lower for w in en.lower().split())
    ]


# ============================================================
# Output workbook — load for incremental writes; resume if exists
# ============================================================

out_path = input_path.replace(".xlsx", "_iptranslated.xlsx")

resuming = os.path.exists(out_path)
if resuming:
    wb = openpyxl.load_workbook(out_path)
    print(f"Resuming from existing output: {os.path.basename(out_path)}")
else:
    wb = openpyxl.load_workbook(input_path)

ws = wb.active

row_map: dict[str, int] = {}
for row_num in range(4, ws.max_row + 1):
    seg_id = str(ws.cell(row=row_num, column=1).value).strip()
    row_map[seg_id] = row_num

# ============================================================
# Document session identifiers (generated client-side)
# ============================================================

doc_id   = str(uuid.uuid4())
doc_key  = base64.b64encode(os.urandom(32)).decode()
doc_name = os.path.basename(input_path)

# ============================================================
# Translate segments
# ============================================================

headers = {
    "Authorization":             f"ApiKey {api_key}",
    "Content-Type":              "text/json; charset=utf-8",
    "Connection":                "Keep-Alive",
    "Upgrade-Insecure-Requests": "1",
    "Cache-Control":             "max-age=0",
    "Host":                      "iptranslator.ipappify.de",
}

translations      = {}
errors            = []
consecutive_nulls = 0
NULL_ABORT        = 10

if resuming:
    for seg_id, row_num in row_map.items():
        val = ws.cell(row=row_num, column=3).value
        if val and str(val).strip():
            translations[seg_id] = str(val).strip()
    if translations:
        print(f"Resuming: {len(translations)} segments already translated, skipping.")

print(f"\nTranslating {len(segments)} segments...")

for i, seg in enumerate(segments):
    if seg["ID"] in translations:
        continue
    past   = segments[max(0, i - CONTEXT_N):i]
    future = segments[i + 1:i + 1 + CONTEXT_N]

    dictionary = build_dictionary(seg["Source"])

    body = {
        "CorrelationId": str(uuid.uuid4()),
        "Client":        "IPTranslator.Client",
        "ClientVersion": "2.0.31.0",
        "Translate": {
            "Options": {
                "CompletionUnit":        1,
                "BeamWidth":             4,
                "TopK":                  1,
                "MaxWordLength":         8,
                "NoAlign":               False,
                "DecoderOptions":        [],
                "ContrastiveAlpha":      0.0,
                "BackTranslationLambda": 0.0,
                "Dictionary":            dictionary,
                "DictionaryReward":      1.0 if dictionary else 0.0,
                "ShortLength":           5,
            },
            "RequestId":   str(uuid.uuid4()),
            "DocumentRef": {
                "Id":        doc_id,
                "Key":       doc_key,
                "Name":      doc_name,
                "CustomRef": None,
                "Created":   datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.0000000+00:00"),
            },
            "Model":          MODEL,
            "SourceLanguage": SRC_LANG,
            "TargetLanguage": TGT_LANG,
            "Source":         seg["Source"],
            "Target":         "",
            "Job": {
                "SourceCurrent": seg["Source"],
                "TargetCurrent": "",
                "SourceSupport": [],
                "TargetSupport": [],
                "SourcePast":    [s["Source"] for s in past],
                "TargetPast":    [translations.get(s["ID"], "") for s in past],
                "SourceFuture":  [s["Source"] for s in future],
            },
        },
        "Align": None, "Evaluate": None, "Embed": None, "Replace": None,
        "Cancel": None, "GenAICheck": None, "Ping": None, "Usage": None,
        "Join": None, "Leave": None, "Update": None, "GetApiKey": None,
    }

    try:
        resp = requests.post(API_URL, headers=headers, data=json.dumps(body).encode("utf-8"), timeout=30, verify=False)
        resp.raise_for_status()
        data = resp.json()
        trans_list = [] if data is None else (data.get("Translate") or {}).get("Translations", [])
        if trans_list:
            consecutive_nulls = 0
            translated = trans_list[0].get("TargetText", "").strip()
            translations[seg["ID"]] = translated
            if seg["ID"] in row_map:
                ws.cell(row=row_map[seg["ID"]], column=3).value = translated
                try:
                    wb.save(out_path)
                except PermissionError:
                    print(f"    WARNING: Could not save — close {os.path.basename(out_path)} in Excel.")
            dict_info = f" [{len(dictionary)}d]" if dictionary else ""
            print(f"  [{i+1}/{len(segments)}] {seg['ID']}{dict_info}: {translated[:80]}")
        else:
            consecutive_nulls += 1
            label = "null" if data is None else "empty"
            print(f"  [{i+1}/{len(segments)}] {seg['ID']}: {label} response (#{consecutive_nulls})")
            errors.append(seg["ID"])
            if consecutive_nulls >= NULL_ABORT:
                print(f"\nAborted: {NULL_ABORT} consecutive empty responses (likely out of credits).")
                print(f"  {len(translations)} segments translated so far.")
                print("\n  [1] Continue workflow with translations we have")
                print("  [2] Stop here and do nothing")
                while True:
                    choice = input("  Choice (1/2): ").strip()
                    if choice == "1":
                        break
                    if choice == "2":
                        print("Stopping.")
                        exit(1)
                break

    except requests.exceptions.HTTPError as e:
        if resp.status_code == 401:
            print(f"\nERROR 401: API key rejected.")
            break
        print(f"  [{i+1}/{len(segments)}] {seg['ID']}: HTTP error — {e}")
        errors.append(seg["ID"])
    except Exception as e:
        print(f"  [{i+1}/{len(segments)}] {seg['ID']}: error — {e}")
        errors.append(seg["ID"])

    time.sleep(1)

print(f"\nDone. {len(translations)} translated, {len(errors)} errors.")
print(f'Output: "{out_path}".')

if errors:
    print(f"\nSegments with errors: {', '.join(errors)}")
