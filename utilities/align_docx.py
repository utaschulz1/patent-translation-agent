"""
align_docx.py — Align two translated Word files paragraph-by-paragraph into Excel.

Extracts non-empty paragraphs from an EN and DE docx, pairs them by position,
and writes a 3-column Excel: Segment | EN | DE.

If the paragraph counts differ, the shorter file gets blank cells for the extra
rows — review those manually in Excel.

Usage:
    python align_docx.py <en.docx> <de.docx>
    python align_docx.py <en.docx> <de.docx> <output.xlsx>   # explicit output path
"""

import sys
from pathlib import Path

from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def read_paragraphs(path: Path) -> list[str]:
    """Return all non-empty paragraphs from a docx file."""
    doc = Document(str(path))
    return [p.text.strip() for p in doc.paragraphs if p.text.strip()]


def write_excel(rows: list[tuple[int, str, str]], out_path: Path, en_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alignment"

    # ── 3-row header ──────────────────────────────────────────────────────────
    # The LLM comparison scripts (LLM_verb/noun/capability_comparison_xlsx.py)
    # always skip the first 3 rows via iloc[3:] because XTM-downloaded Excels
    # have 3 header rows (row 1: filename, row 2: column labels, row 3: language).
    # This output replicates that structure so the same scripts can be run on
    # aligned Excels via workflow_review.py without any special-casing.
    # Row 1: source filename
    ws.cell(row=1, column=1, value=en_path.name)
    # Row 2: column labels
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    for col, label in enumerate(["Id", "Source (EN)", "Target (DE)"], start=1):
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    # Row 3: language info (mirrors XTM's "German (Germany)" in col C)
    ws.cell(row=3, column=3, value="German (Germany)")

    # ── Data rows start at row 4 ─────────────────────────────────────────────
    mismatch_fill = PatternFill("solid", fgColor="FFE699")
    for seg, en, de in rows:
        row_num = seg + 3   # seg is 1-based → row 4 onward
        ws.cell(row=row_num, column=1, value=seg)
        ws.cell(row=row_num, column=2, value=en)
        ws.cell(row=row_num, column=3, value=de)
        if not en or not de:
            for col in range(1, 4):
                ws.cell(row=row_num, column=col).fill = mismatch_fill

    # Column widths
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70

    # Wrap text for data rows
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=4, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = wrap

    wb.save(out_path)


def main():
    if len(sys.argv) < 3:
        print("Usage: python align_docx.py <en.docx> <de.docx> [output.xlsx]")
        raise SystemExit(1)

    en_path = Path(sys.argv[1])
    de_path = Path(sys.argv[2])

    if len(sys.argv) >= 4:
        out_path = Path(sys.argv[3])
    else:
        stem = en_path.stem.replace("_EN", "").replace("_en", "")
        out_path = en_path.parent / f"{stem}_aligned.xlsx"

    en_paras = read_paragraphs(en_path)
    de_paras = read_paragraphs(de_path)

    print(f"EN paragraphs: {len(en_paras)}")
    print(f"DE paragraphs: {len(de_paras)}")

    if len(en_paras) != len(de_paras):
        print(f"WARNING: counts differ by {abs(len(en_paras) - len(de_paras))} — "
              f"mismatched rows will be highlighted in yellow.")

    n = max(len(en_paras), len(de_paras))
    rows = [
        (i + 1, en_paras[i] if i < len(en_paras) else "", de_paras[i] if i < len(de_paras) else "")
        for i in range(n)
    ]

    write_excel(rows, out_path, en_path)
    print(f"Written {n} rows to: {out_path}")


if __name__ == "__main__":
    main()
