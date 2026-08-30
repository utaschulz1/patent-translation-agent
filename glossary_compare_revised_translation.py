"""
glossary_compare_revised_translation.py — Verb and noun glossary compliance check.
No LLM used — only lookup-based lemmatization and truncation matching.

Thin wrapper since Phase 0 of PRD_glossary_agent.md §4: the matching logic
lives in glossary_lib/matching.py and is re-exported here, so every existing
importer (app.py's CAT UI check endpoint, review_agent, tests) keeps working
unchanged. Only the script entry point (main) remains here.

Usage: python glossary_compare_revised_translation.py [--pid <project_id>]
  --pid   project folder name under projects/; defaults to current project context

Reads the *_translated.xlsx for the active project, copies ID/EN/DE
into a new revised_translation_checks.xlsx, then annotates column D with
glossary mismatches found via lookup-based lemmatization (verbs) and
truncation matching (nouns).

Annotation format in column D:
  EN found, DE absent  → "EN: {term} ({n}), DE: missing, expected: {de_term}"
  EN and DE found,
  counts differ        → "EN: {term} ({en_n}), DE: {de_term} ({de_n})"

Constraint — source-triggered only: checks are initiated by finding a glossary
term in the EN source. A DE glossary term that appears in the target without a
corresponding EN term in the source is not detected here. Target-triggered
checks (e.g. "umfass*" without "compris*", "Vielzahl" without "plurality") are
handled by the linter instead.

Public API (importable):
  build_glossary_lookups(proj_dir) → (verb_lookup, verb_fallback, noun_lookup, all_de_noun_terms)
  check_segment_glossary(en_text, de_text, verb_lookup, noun_lookup, all_de_noun_terms, verb_fallback) → list[str]
"""

import argparse
import glob
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

import project_log
from glossary_lib.matching import (  # noqa: F401
    _DE_ADJ_SUFFIXES,
    _count_en_phrase,
    _count_lemmas,
    _count_noun_in_de,
    _mask_de_noun_phrases,
    build_glossary_lookups,
    check_segment_glossary,
    de_verb_lookup,
    en_verb_lookup,
    load_lemma_tables,
)

HERE = Path(__file__).parent
HEADER_ROWS = 3  # rows 1–3 are filename / column-name / language lines in the xlsx


# ── Script entry point ────────────────────────────────────────────────────────

def main() -> None:
    """Annotate the active project's *_translated.xlsx with glossary mismatches."""
    _parser = argparse.ArgumentParser()
    _parser.add_argument("--pid", default=None)
    args = _parser.parse_args()

    if args.pid:
        proj_dir = project_log.find_project_dir(args.pid)
    else:
        proj_dir = project_log.project_dir()

    verb_lookup, verb_fallback, noun_lookup, all_de_noun_terms = build_glossary_lookups(proj_dir)
    lemma_tables = load_lemma_tables(proj_dir)

    print(f"Glossary: {proj_dir}")
    print(f"Verb entries loaded: {len(verb_lookup)}")
    for en, de in verb_lookup.items():
        print(f"  {en} → {de}")
    print(f"Verb fallback entries: {len(verb_fallback)}")
    for en, de in verb_fallback.items():
        print(f"  {en} → {de} (word match)")
    print(f"Noun entries loaded: {len(noun_lookup)}")
    for en, de in noun_lookup.items():
        print(f"  {en} → {de}")

    for pattern in [
        str(proj_dir / "*_revised_translation_checks.xlsx"),
        str(proj_dir / "*_GERMAN_translated.xlsx"),
        str(proj_dir / "*_translated.xlsx"),
    ]:
        src_files = [f for f in glob.glob(pattern) if not Path(f).name.startswith("~$")]
        if src_files:
            break

    if not src_files:
        raise FileNotFoundError(f"No _translated.xlsx found in {proj_dir}")

    src_path = Path(src_files[0])
    print(f"\nSource: {src_path.name}")

    src_wb = openpyxl.load_workbook(src_path)
    src_ws = src_wb.active

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active

    for row_num in range(1, HEADER_ROWS + 1):
        for col in range(1, 4):
            out_ws.cell(row=row_num, column=col).value = src_ws.cell(row=row_num, column=col).value

    out_ws.cell(row=2, column=4).value = "Glossary Checks"
    out_ws.column_dimensions["A"].width = 8
    out_ws.column_dimensions["B"].width = 20
    out_ws.column_dimensions["C"].width = 20
    out_ws.column_dimensions["D"].width = 35

    annotated = 0

    for row_num in range(HEADER_ROWS + 1, src_ws.max_row + 1):
        seg_id  = src_ws.cell(row=row_num, column=1).value
        en_text = src_ws.cell(row=row_num, column=2).value
        de_text = src_ws.cell(row=row_num, column=3).value

        out_ws.cell(row=row_num, column=1).value = seg_id
        for col, val in ((2, en_text), (3, de_text)):
            cell = out_ws.cell(row=row_num, column=col)
            cell.value = val
            cell.alignment = Alignment(wrap_text=True)

        if not en_text or not de_text:
            continue

        notes = check_segment_glossary(
            str(en_text), str(de_text), verb_lookup, noun_lookup, all_de_noun_terms,
            verb_fallback=verb_fallback, lemma_tables=lemma_tables,
        )

        if notes:
            cell = out_ws.cell(row=row_num, column=4)
            cell.value = "\n".join(notes)
            cell.alignment = Alignment(wrap_text=True)
            annotated += 1

    out_name = src_path.name.replace("_translated.xlsx", "_revised_translation_checks.xlsx")
    if out_name == src_path.name:
        out_name = src_path.stem + "_re-checked.xlsx"
    out_path = proj_dir / out_name
    try:
        out_wb.save(out_path)
    except PermissionError:
        stamp = datetime.now().strftime("%H%M%S")
        out_path = proj_dir / out_name.replace(".xlsx", f"_{stamp}.xlsx")
        out_wb.save(out_path)

    print(f"\nAnnotated {annotated} segment(s).")
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
