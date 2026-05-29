"""
matecat_xtm_verify.py  —  Verify XTM upload by comparing translations

Downloads the current bilingual XLIFF from the XTM Workbench and compares
every segment's target text against the *_GERMAN.xlf downloaded from MateCat.

Reports segment IDs where translations differ, so silent upload failures
can be detected without manual inspection.

Usage:
    python matecat_xtm_verify.py <project_id>
    e.g.  python matecat_xtm_verify.py CATG_2605_P0222
"""

import glob
import re
import sys
import tempfile
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HERE = Path(__file__).parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT))

from xtm_initial_download import (
    _setup_session,
    _download_preview,
    _unpack_xbpkg,
    _find_pre_folder,
)
from matecat_xtm_upload import _read_xlf

_XLF_NS = "urn:oasis:names:tc:xliff:document:1.2"


def _normalise(text: str) -> str:
    """Normalise whitespace and unicode for comparison."""
    text = unicodedata.normalize("NFC", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def run(project_id: str) -> None:
    proj_dir = ROOT / "projects" / project_id

    # ── MateCat reference file ────────────────────────────────────────────────

    mc_files = list(proj_dir.glob("*_GERMAN.xlf"))
    if not mc_files:
        raise RuntimeError(f"No *_GERMAN.xlf found in {proj_dir}")
    mc_path = mc_files[-1]          # most recent if multiple
    print(f"MateCat reference: {mc_path.name}")

    mc_segments, _ = _read_xlf(mc_path)
    mc_translations: dict[int, str] = dict(mc_segments)
    print(f"  {len(mc_translations)} segments with translations")

    # ── Download current XTM XLIFF ────────────────────────────────────────────

    # ── Find or download XTM XLIFF ────────────────────────────────────────────

    from config import WORK_DIR as XTRF_BASE
    _XLIFF_EXTS = {".xlf", ".xliff", ".sdlxliff", ".mqxliff"}

    def _find_local_xlf() -> Path | None:
        """Look for an existing xbpkg or xlf in any XTRF job folder for this project."""
        for folder in XTRF_BASE.glob(f"*_{project_id}"):
            if not folder.is_dir():
                continue
            # Prefer already-extracted XLIFF
            for ext in _XLIFF_EXTS:
                hits = list(folder.glob(f"*{ext}"))
                if hits:
                    print(f"  Found local XLIFF: {hits[0]}")
                    return hits[0]
            # Fall back to xbpkg — unpack it
            for xbpkg in folder.glob("*.xbpkg"):
                print(f"  Found local xbpkg: {xbpkg.name} — unpacking...")
                xliffs = []
                with zipfile.ZipFile(xbpkg) as z:
                    for member in z.namelist():
                        if Path(member).suffix.lower() in _XLIFF_EXTS:
                            dest = xbpkg.parent / Path(member).name
                            dest.write_bytes(z.read(member))
                            xliffs.append(dest)
                            print(f"  Extracted: {dest.name}")
                if xliffs:
                    return xliffs[0]
        return None

    print()
    xtm_path = _find_local_xlf()

    if xtm_path is None:
        print("No local XLIFF found — downloading from XTM...")
        session, session_token, csrf_token = _setup_session(project_id)
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            xbpkg = _download_preview(
                session, session_token, csrf_token,
                "XBENCH_INTERACTIVE", tmp_path, project_id,
            )
            xliffs = _unpack_xbpkg(xbpkg)
            if not xliffs:
                raise RuntimeError("No XLIFF extracted from XTM xbpkg.")
            xtm_path = xliffs[0]
            print(f"  Extracted: {xtm_path.name}")
            xtm_segments, _ = _read_xlf(xtm_path)
            xtm_translations: dict[int, str] = dict(xtm_segments)
            print(f"  {len(xtm_translations)} segments with translations")
    else:
        xtm_segments, _ = _read_xlf(xtm_path)
        xtm_translations: dict[int, str] = dict(xtm_segments)
        print(f"  {len(xtm_translations)} segments with translations")

    # ── Compare ───────────────────────────────────────────────────────────────

    print()
    all_ids = sorted(set(mc_translations) | set(xtm_translations))
    diffs: list[tuple[int, str, str]] = []

    for seg_id in all_ids:
        mc_text  = _normalise(mc_translations.get(seg_id, ""))
        xtm_text = _normalise(xtm_translations.get(seg_id, ""))

        if mc_text and not xtm_text:
            diffs.append((seg_id, mc_text, "<missing in XTM>"))
        elif mc_text and mc_text != xtm_text:
            diffs.append((seg_id, mc_text, xtm_text))

    # ── Report ────────────────────────────────────────────────────────────────

    print(f"Compared {len(all_ids)} segments — {len(diffs)} mismatch(es).")

    if not diffs:
        print(f"OK — all {len(mc_translations)} segments match.")
        return

    # Write XLSX
    out_path = proj_dir / f"xtm_verify_{project_id}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mismatches"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    wrap        = Alignment(wrap_text=True, vertical="top")
    red_fill    = PatternFill("solid", fgColor="FCE4D6")

    headers = ["Segment ID", "Status", "MateCat (reference)", "XTM (actual)"]
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 60

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = wrap

    for row_num, (seg_id, mc_text, xtm_text) in enumerate(diffs, 2):
        status = "missing in XTM" if xtm_text == "<missing in XTM>" else "mismatch"
        for col, val in enumerate([seg_id, status, mc_text, xtm_text], 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = wrap
            if status == "missing in XTM":
                cell.fill = red_fill
        ws.row_dimensions[row_num].height = 60

    wb.save(out_path)
    print(f"\nSaved: {out_path.name}  ({len(diffs)} row(s))")
    print(f"Segment IDs: {[s for s, *_ in diffs]}")


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python matecat_xtm_verify.py <project_id>")
        raise SystemExit(1)
    try:
        run(sys.argv[1])
    except (RuntimeError, TimeoutError) as e:
        print(f"\nError: {e}")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
