# ============================================================
# matecat_xlf_to_excel.py
# ============================================================
# Converts a *_GERMAN.xlf (downloaded from MateCat) to the
# *_translated.xlsx format expected by glossary_compare_revised_translation.py
# and linter.py.
#
# Output format:
#   Row 1: filename
#   Row 2: ID | Source | Target
#   Row 3: (blank)
#   Row 4+: segment ID (int) | EN source | DE target
#
# USAGE   python matecat_xlf_to_excel.py [--pid <project_id>]
# ============================================================

import argparse
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

HERE = Path(__file__).parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT))
from project_log import project_dir as _pdir

_XLF_NS = "urn:oasis:names:tc:xliff:document:1.2"

def _ns(tag: str) -> str:
    return f"{{{_XLF_NS}}}{tag}"


def _element_text(elem: ET.Element) -> str:
    return "".join(elem.itertext()).strip()


def _segment_id(unit_id: str, xtm_url: str) -> int | None:
    if unit_id.startswith("t") and unit_id[1:].isdigit():
        return int(unit_id[1:])
    m = re.search(r"segmentId=(\d+)", xtm_url)
    return int(m.group(1)) if m else None


_ap = argparse.ArgumentParser()
_ap.add_argument("--pid", default=None)
_args = _ap.parse_args()

if _args.pid:
    proj_dir = ROOT / "projects" / _args.pid
    if not proj_dir.exists():
        print(f"ERROR: Project folder not found: {proj_dir}")
        sys.exit(1)
else:
    proj_dir = _pdir()

project_id = proj_dir.name

xlf_files = list(proj_dir.glob("*_GERMAN.xlf"))
if not xlf_files:
    print(f"ERROR: No *_GERMAN.xlf found in {proj_dir}")
    print("  Run matecat_download.py first.")
    sys.exit(1)
if len(xlf_files) > 1:
    print(f"  Multiple *_GERMAN.xlf files, using: {xlf_files[0].name}")
xlf_path = xlf_files[0]

print(f"Converting {xlf_path.name}...")

tree = ET.parse(xlf_path)
rows: list[tuple[int, str, str]] = []

for tu in tree.getroot().iter(_ns("trans-unit")):
    seg_id = _segment_id(
        tu.get("id", ""),
        tu.get("{urn:xliff-xtm-extensions}url", ""),
    )
    if seg_id is None:
        continue
    source = _element_text(tu.find(_ns("source"))) if tu.find(_ns("source")) is not None else ""
    target = _element_text(tu.find(_ns("target"))) if tu.find(_ns("target")) is not None else ""
    rows.append((seg_id, source, target))

rows.sort(key=lambda r: r[0])
print(f"  {len(rows)} segments  ({sum(1 for _, _, t in rows if t)} with translation)")

# ── Write Excel ───────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 80

wrap = Alignment(wrap_text=True, vertical="top")

# Header rows (same format as *_translated.xlsx from XTM)
ws.cell(row=1, column=1).value = xlf_path.stem   # filename in A1
ws.cell(row=2, column=1).value = "ID"
ws.cell(row=2, column=2).value = "Source"
ws.cell(row=2, column=3).value = "Target"
ws.cell(row=3, column=2).value = "EN"
ws.cell(row=3, column=3).value = "DE"

for i, (seg_id, source, target) in enumerate(rows, start=4):
    ws.cell(row=i, column=1).value = seg_id
    ws.cell(row=i, column=2).value = source
    ws.cell(row=i, column=3).value = target
    ws.cell(row=i, column=2).alignment = wrap
    ws.cell(row=i, column=3).alignment = wrap

out_path = proj_dir / f"{xlf_path.stem}_translated.xlsx"
wb.save(out_path)
print(f"  Saved: {out_path.name}")

# Remove stale *_revised_translation_checks.xlsx so glossary_compare picks up
# the fresh translated file above rather than an old checks file.
for stale in proj_dir.glob("*_revised_translation_checks*.xlsx"):
    stale.unlink()
    print(f"  Removed stale: {stale.name}")
print(f"\nNext steps:")
print(f"  python glossary_compare_revised_translation.py --pid {project_id}")
print(f"  python linter.py --pid {project_id}")
