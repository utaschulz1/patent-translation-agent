"""
xtm_workbench.py  —  XTM Workbench: login, find project, download Excel

Downloads the bilingual EXCEL_EXTENDED_TABLE from the XTM preview menu
for the given project ID.

When run in the workflow, the file is saved to:
    WORK_DIR\\pre-processing\\

where <project-folder> is the first subfolder whose name contains the project ID.

When run directly, the file is saved to C:\\Users\\utasc\\Downloads

Usage:
    python xtm_workbench.py <project_id>
    e.g.  python xtm_workbench.py RTC_2604_P0732
"""

import glob
import json
import os
import shutil
import zipfile
import random
import re
import string
import sys
import time
import uuid
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl
import requests
import websocket as _websocket
from dotenv import load_dotenv

import project_log
from config import WORK_DIR

DOWNLOADS = Path(r"C:\Users\utasc\Downloads")


BASE_URL = "https://word.welocalize.com/project-manager-gui"
WB_BASE  = "https://word.welocalize.com/workbench"

_ENV = Path(__file__).parent / ".env"


def _load_creds() -> tuple[str, str]:
    """Load XTM username and password from the .env file next to this script."""
    load_dotenv(_ENV)
    return (
        os.environ["XTM_WORKBENCH_USERNAME5"],
        os.environ["XTM_WORKBENCH_PASSWORD5"],
    )


def _login(session: requests.Session, username: str, password: str) -> str:
    """Authenticate against XTM and return the uust session token.

    The token is embedded as a hidden input in the post-login HTML page and
    must be sent as a request header on all subsequent API calls.
    """
    login_page = f"{BASE_URL}/login.jsp?client=IP"

    # Seed cookies that AngularJS sets via JavaScript (requests doesn't run JS)
    session.cookies.set("client", "IP", domain="word.welocalize.com")
    session.cookies.set("languageCode", "en_GB", domain="word.welocalize.com")
    session.cookies.set("xtm-user-app-uuid", str(uuid.uuid4()), domain="word.welocalize.com")

    # Fetch login page so server sets JSESSIONID
    session.get(login_page)

    r = session.post(
        f"{BASE_URL}/login.serv",
        json={"client": "IP", "username": username, "password": password},
        headers={
            "Referer": login_page,
            "Origin": "https://word.welocalize.com",
            "X-Requested-With": "XMLHttpRequest",
        },
    )
    print(f"  Login response: {r.status_code} — {r.text[:300]}")
    r.raise_for_status()
    data = r.json()
    if "error" in str(data).lower() and "redirectURL" not in data:
        raise RuntimeError(f"XTM login failed: {data}")
    print(f"  Logged in as {username}")

    # Fetch the post-login page and extract the uust token from the hidden input
    redirect = data.get("redirectURL", "/project-manager-gui/configuration-pages.action")
    config_html = session.get(f"https://word.welocalize.com{redirect}").text
    m = re.search(r'<input[^>]+id=["\']uust["\'][^>]+value=["\']([^"\']+)["\']', config_html)
    if not m:
        m = re.search(r'<input[^>]+value=["\']([^"\']+)["\'][^>]+id=["\']uust["\']', config_html)
    if not m:
        raise RuntimeError("Could not extract uust token from configuration page")
    uust = m.group(1)
    print(f"  uust token: {uust[:12]}...")
    return uust


def _get_tasks(session: requests.Session) -> list[dict]:
    """Fetch up to 100 in-progress tasks from the XTM inbox, ordered by due date."""
    r = session.get(
        f"{BASE_URL}/myinbox/getInProgressElements.serv",
        params={
            "draw": 1,
            "length": 100,
            "start": 0,
            "orderColumn": "DUE_DATE",
            "orderDir": "DESC",
            "searchValue": "",
            "assignToValue": "ALL",
        },
    )
    if not r.ok:
        print(f"  Task list error {r.status_code}: {r.text[:500]}")
    r.raise_for_status()
    data = r.json()
    print(f"  Raw response keys: {list(data.keys())}")
    print(f"  recordsTotal={data.get('recordsTotal')}  recordsFiltered={data.get('recordsFiltered')}")
    tasks = data.get("data", [])
    if tasks:
        print(f"  First task keys: {list(tasks[0].keys())}")
    return tasks


def _find_task(tasks: list[dict], project_id: str, file_filter: str | None = None) -> dict:
    """Return the task whose projectName contains project_id.

    file_filter: optional substring matched against the FILE field name —
    useful when multiple tasks share the same project name (e.g. drawings vs claims).
    If omitted and multiple matches exist, warns and returns the first.
    """
    matches = [t for t in tasks
               if project_id in t.get("additionalData", {}).get("projectName", "")]
    if not matches:
        raise RuntimeError(
            f"Project '{project_id}' not found in task list "
            f"({len(tasks)} tasks returned). Check the project ID."
        )
    if file_filter:
        filtered = [t for t in matches
                    if file_filter.lower() in _task_filename(t).lower()]
        if filtered:
            print(f"  File filter '{file_filter}' → matched: {_task_filename(filtered[0])}")
            return filtered[0]
        print(f"  Warning: --file '{file_filter}' matched none of:")
        for t in matches:
            print(f"    {_task_filename(t)}")
        print(f"  Ignoring filter, using first match.")
    if len(matches) > 1:
        print(f"  Warning: {len(matches)} tasks found for '{project_id}' — using first.")
        print(f"  Use --file <substring> to select a specific one:")
        for i, t in enumerate(matches):
            fname = _task_filename(t)
            words = t.get("WORDS", "?")
            fid   = t.get("additionalData", {}).get("fileId", "?")
            print(f"    [{i}] fileId={fid}  words={words}  file={fname}")
    return matches[0]


def _task_filename(task: dict) -> str:
    """Extract a human-readable filename from a task dict."""
    # FILE field can be a string or a dict with a 'name' key
    f = task.get("FILE", "")
    if isinstance(f, dict):
        return f.get("name", str(f))
    if f:
        return str(f)
    return task.get("additionalData", {}).get("fileName", "?")


def _open_editor(session: requests.Session, task: dict) -> tuple[str, str]:
    """POST to openEditor.serv, return (workbench_url, session_token)."""
    ad = task["additionalData"]
    uust = session.headers.get("uust", "")
    payload = {
        "actorType":                  ad.get("actorType", "USERGROUP"),
        "fileId":                     ad["fileId"],
        "groupTaskAction":            "undefined",
        "isGroupTaskWithPriorities":  "false",
        "readOnly":                   str(ad.get("readOnly", True)).lower(),
        "workflowStepName":           ad.get("stepName", ""),
        "stepReferenceId":            ad["stepReferenceId"],
        "taskType":                   "ACTIVE",
        "workflowReferenceStepName":  ad.get("stepReferenceName", ""),
        "uust":                       uust,
    }
    r = session.post(
        f"{BASE_URL}/openEditor.serv",
        data=payload,
        allow_redirects=False,
    )
    r.raise_for_status()

    # Response is XML — check for an application-level error first
    xml = r.text
    if "<result>error</result>" in xml:
        msg_m = re.search(r"<msg>(.*?)</msg>", xml)
        raise RuntimeError(f"openEditor error: {msg_m.group(1) if msg_m else xml[:300]}")

    url_m = re.search(r"<url>(.*?)</url>", xml, re.DOTALL) or \
            re.search(r"<msg>(https?://.*?)</msg>", xml, re.DOTALL)
    if not url_m:
        if r.headers.get("Location"):
            wb_url = r.headers["Location"]
        else:
            raise RuntimeError(f"Cannot parse openEditor response:\n{xml[:500]}")
    else:
        wb_url = url_m.group(1).strip().replace("&amp;", "&")

    session_token = parse_qs(urlparse(wb_url).query).get("_s", [None])[0]
    if not session_token:
        raise RuntimeError(f"No _s token in workbench URL: {wb_url}")
    return wb_url, session_token


def _init_workbench(session: requests.Session, wb_url: str, session_token: str) -> str:
    """Navigate to the workbench start URL, hit /web/init, and return the CSRF token."""
    session.get(wb_url)
    r = session.get(
        f"{WB_BASE}/web/init",
        params={"_s": session_token},
    )
    r.raise_for_status()
    csrf_token = r.json().get("csrfToken", {}).get("token", "")
    print("  Workbench initialised")
    return csrf_token


def _generate_preview(session: requests.Session, session_token: str, csrf_token: str, preview_type: str = "EXCEL_EXTENDED_TABLE") -> str:
    """Connect via WebSocket STOMP, request preview generation, return download ticket."""
    server_id = str(random.randint(0, 999)).zfill(3)
    session_id = "".join(random.choices(string.ascii_lowercase + string.digits, k=8))
    ws_url = (
        f"wss://word.welocalize.com/workbench/ws/{server_id}/{session_id}"
        f"/websocket?_s={session_token}"
    )
    cookie_str = "; ".join(f"{c.name}={c.value}" for c in session.cookies)

    ws = _websocket.WebSocket()
    ws.settimeout(15)
    ws.connect(ws_url, cookie=cookie_str)
    try:
        try:
            ws.recv()  # SockJS open frame 'o'
        except (TimeoutError, OSError, _websocket.WebSocketTimeoutException) as e:
            raise RuntimeError(f"WebSocket handshake timed out (SockJS open frame): {e}") from e

        connect_frame = (
            f"CONNECT\nX-CSRF-TOKEN:{csrf_token}\n"
            f"accept-version:1.0,1.1,1.2\nheart-beat:10000,10000\n\n\x00"
        )
        ws.send(json.dumps([connect_frame]))
        try:
            ws.recv()  # CONNECTED
        except (TimeoutError, OSError, _websocket.WebSocketTimeoutException) as e:
            raise RuntimeError(f"WebSocket handshake timed out (STOMP CONNECTED): {e}") from e

        ws.send(json.dumps(["SUBSCRIBE\nid:sub-0\ndestination:/user/queue/main\n\n\x00"]))

        request_id = str(int(time.time() * 1000))
        body = json.dumps({"requestId": request_id, "previewType": preview_type})
        send_frame = (
            f"SEND\ndestination:/workbench/document/preview/generate\n"
            f"_s:{session_token}\ncontent-length:{len(body)}\n\n{body}\x00"
        )
        ws.send(json.dumps([send_frame]))

        ws.settimeout(5)
        deadline = time.time() + 30   # 30s cap — normal generation takes ~5s
        received_types: list[str] = []  # collect message types for debug output on timeout
        try:
            while time.time() < deadline:
                try:
                    raw = ws.recv()
                except _websocket.WebSocketTimeoutException:
                    continue  # heartbeat gap — keep waiting until deadline
                if not raw or raw == "h":
                    continue
                if raw.startswith("a"):
                    for stomp_msg in json.loads(raw[1:]):
                        # Extract message type for debug logging
                        try:
                            hdr, _, body_part = stomp_msg.partition("\n\n")
                            msg_type = next(
                                (l.split(":", 1)[1].strip() for l in hdr.splitlines()
                                 if l.startswith("type:")), None
                            ) or (
                                json.loads(body_part.rstrip("\x00")).get("type", "?")
                                if body_part.strip().startswith("{") else "?"
                            )
                        except Exception:
                            msg_type = "?"
                        received_types.append(msg_type)
                        print(f"    [WS] {msg_type}")

                        if "PREVIEW_GENERATION_FINISHED" in stomp_msg:
                            body_str = stomp_msg[stomp_msg.rfind("\n\n") + 2:].rstrip("\x00")
                            payload = json.loads(body_str).get("payload", {})
                            if payload.get("resultType") == "SUCCESS":
                                return payload["downloadTicket"]
                            raise RuntimeError(f"Preview generation failed: {payload}")
            raise RuntimeError(
                f"Preview generation timed out after 30 s.\n"
                f"  Message types received: {received_types}\n"
                f"  If this list is empty the WebSocket was idle — likely a tag error in XTM blocking generation.\n"
                f"  Download the xbpkg manually or fix tag errors in XTM first."
            )
        except _websocket.WebSocketConnectionClosedException as e:
            raise RuntimeError("WebSocket closed before download ticket was received") from e
    finally:
        ws.close()


def _keepalive(session: requests.Session) -> None:
    """Ping the server to prevent session timeout; silently ignores any network error."""
    try:
        session.post(f"{BASE_URL}/sayHelloToServer.serv")
    except Exception:
        pass


def _download_excel(
    session: requests.Session,
    session_token: str,
    ticket: str,
    dest_folder: Path,
    project_id: str,
) -> Path:
    """Download the bilingual Excel preview file and save it to dest_folder, returning the path."""
    r = session.get(
        f"{WB_BASE}/web/preview/document",
        params={"_s": session_token, "downloadTicket": ticket},
        stream=True,
    )
    r.raise_for_status()

    # Derive filename from Content-Disposition or fall back to project_id
    cd = r.headers.get("content-disposition", "")
    m = re.search(r'filename[^;=\n]*=\s*["\']?([^"\';\n]+)', cd)
    fname = m.group(1).strip() if m else f"{project_id}_bilingual.xlsx"
    fname = f"{project_id}_{fname}"
    out = dest_folder / fname

    with open(out, "wb") as f:
        for chunk in r.iter_content(8192):
            f.write(chunk)

    return out


def _find_pre_folder(project_id: str) -> Path:
    """Locate the pre-processing subfolder for this project inside WORK_DIR."""
    for candidate in WORK_DIR.iterdir():
        if candidate.is_dir() and project_id in candidate.name:
            pre = candidate / "pre-processing"
            pre.mkdir(exist_ok=True)
            return pre
    raise RuntimeError(
        f"Project folder containing '{project_id}' not found in {WORK_DIR}"
    )


XLIFF_EXTENSIONS = {".xlf", ".xliff", ".sdlxliff", ".mqxliff"}


def _unpack_xbpkg(xbpkg: Path) -> list[Path]:
    """Extract XLIFF files from an xbpkg ZIP, delete the package, return extracted paths."""
    dest = xbpkg.parent
    xliffs = []
    with zipfile.ZipFile(xbpkg) as z:
        members = z.namelist()
        print(f"  Contents: {members}")
        for member in members:
            if Path(member).suffix.lower() in XLIFF_EXTENSIONS:
                target = dest / Path(member).name
                target.write_bytes(z.read(member))
                xliffs.append(target)
                print(f"  Extracted: {target.name}")
    if not xliffs:
        raise RuntimeError(f"No XLIFF files found inside {xbpkg.name}. Contents: {members}")
    xbpkg.unlink()
    return xliffs


def _setup_session(project_id: str) -> tuple[requests.Session, str, str]:
    """Login, find task, open workbench. Returns (session, session_token, csrf_token)."""
    username, password = _load_creds()

    session = requests.Session()
    session.headers.update({
        "Accept": "application/json, text/plain, */*",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:149.0) Gecko/20100101 Firefox/149.0",
    })

    print("Step 1 — Login to XTM Workbench...")
    uust = _login(session, username, password)
    session.headers.update({"uust": uust, "X-Requested-With": "XMLHttpRequest"})

    print("Step 2 — Fetching task list...")
    tasks = _get_tasks(session)
    print(f"  {len(tasks)} task(s) in progress")

    task = _find_task(tasks, project_id)
    ad = task["additionalData"]
    print(f"  Found: {ad.get('projectName', '?')}  (file {ad.get('fileId', '?')})")

    print("Step 3 — Opening workbench editor...")
    wb_url, session_token = _open_editor(session, task)
    print(f"  Session token: {session_token[:12]}...")

    csrf_token = _init_workbench(session, wb_url, session_token)
    time.sleep(3)
    _keepalive(session)

    return session, session_token, csrf_token


def _download_preview(
    session: requests.Session,
    session_token: str,
    csrf_token: str,
    preview_type: str,
    folder: Path,
    project_id: str,
) -> Path:
    print(f"  Generating preview ({preview_type})...")
    ticket = _generate_preview(session, session_token, csrf_token, preview_type)
    print(f"  Preview ticket: {ticket}")
    out_path = _download_excel(session, session_token, ticket, folder, project_id)
    print(f"  Saved: {out_path}")
    return out_path


def run(project_id: str, dest_folder: Path | None = None) -> dict[str, Path | list[Path]]:
    """Login once, download both the bilingual Excel and the XLIFF. Returns dict with both paths."""
    session, session_token, csrf_token = _setup_session(project_id)
    folder = dest_folder if dest_folder is not None else _find_pre_folder(project_id)

    print("Step 4 — Downloading bilingual Excel...")
    xlsx = _download_preview(session, session_token, csrf_token, "EXCEL_EXTENDED_TABLE", folder, project_id)

    print("Step 5 — Downloading XLIFF...")
    xbpkg = _download_preview(session, session_token, csrf_token, "XBENCH_INTERACTIVE", folder, project_id)
    xliffs = _unpack_xbpkg(xbpkg)

    return {"xlsx": xlsx, "xliff": xliffs}


def _find_xlsx(project_id: str) -> Path | None:
    """Return the most recently modified xlsx in Downloads matching project_id, or None."""
    matches = sorted(
        glob.glob(str(DOWNLOADS / f"{project_id}*.xlsx")),
        key=os.path.getmtime,
        reverse=True,
    )
    return Path(matches[0]) if matches else None


def run_workflow(project_id: str, msg_id: str | None = None) -> bool:
    """Download originals to pre-processing, copy trimmed xlsx + xliff to project folder.

    msg_id: Gmail message ID used to write log events; pass None for the manual path.
    Returns True on success, False if no files could be obtained.
    """
    proj_dir = project_log.project_dir()
    xlsx_path: Path | None = None

    def _log(state, detail=None):
        if msg_id:
            project_log.log_event(msg_id, state, detail=detail)

    # 1. Try XTM API: originals → pre-processing, trimmed xlsx + xliff → project folder
    try:
        print("Downloading XTM Excel + XLIFF via API...")
        result = run(project_id)  # dest_folder=None → _find_pre_folder → pre-processing
        xlsx_orig = result["xlsx"]
        xliff_paths = result["xliff"]

        xlsx_dest = proj_dir / xlsx_orig.name
        shutil.copy2(xlsx_orig, xlsx_dest)
        wb = openpyxl.load_workbook(xlsx_dest)
        ws = wb.active
        while ws.max_column > 3:
            ws.delete_cols(ws.max_column)
        wb.save(xlsx_dest)
        xlsx_path = xlsx_dest

        for xlf in xliff_paths:
            shutil.copy2(xlf, proj_dir / xlf.name)

        print(f"  xlsx original: {xlsx_orig.name} (pre-processing)")
        print(f"  xlsx trimmed copy: {xlsx_dest.name} (project folder)")
        for xlf in xliff_paths:
            print(f"  xliff copy: {xlf.name} (project folder)")
        _log("XLSX_DOWNLOADED", detail=str(xlsx_path))
    except Exception as e:
        print(f"  XTM download failed ({e}) — falling back to Downloads folder.")

    # 2. Fall back to pre-downloaded xlsx in Downloads (no xliff in this path)
    if xlsx_path is None:
        found = _find_xlsx(project_id)
        if found:
            try:
                pre_folder = _find_pre_folder(project_id)
                shutil.copy2(found, pre_folder / found.name)
                print(f"  Original saved: {found.name} (pre-processing)")
            except Exception:
                pass

            dest = proj_dir / found.name
            shutil.copy2(found, dest)
            wb = openpyxl.load_workbook(dest)
            ws = wb.active
            while ws.max_column > 3:
                ws.delete_cols(ws.max_column)
            wb.save(dest)
            xlsx_path = dest
            print(f"  Copied from Downloads: {found.name}")
            _log("XLSX_FOUND", detail=str(xlsx_path))

    if xlsx_path is None:
        _log("XLSX_NOT_FOUND")
        print(
            f"WARNING: Could not obtain xlsx/xlf for {project_id}.\n"
            f"  Place xlsx and xlf in {proj_dir} and re-run."
        )
        return False

    _log("JOB_FINISHED_SUCCESSFULLY", detail=project_id)
    return True


def main():
    """CLI entry point: read project_id from argv and call run()."""
    if len(sys.argv) < 2:
        print("Usage: python xtm_initial_download.py <project_id>")
        raise SystemExit(1)
    result = run(sys.argv[1], dest_folder=Path(r"C:\Users\utasc\Downloads"))
    print(f"Excel:  {result['xlsx']}")
    for p in result["xliff"]:
        print(f"XLIFF:  {p}")


if __name__ == "__main__":
    main()
