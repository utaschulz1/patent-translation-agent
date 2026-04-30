"""
get_XTRF_link.py — Step 3: Gmail intake → XTRF job setup → XTM Excel download

Reads the ComunicaDK/TODO label in Gmail, selects the unprocessed email with
the closest deadline, and runs the full intake pipeline:

  1. Extract XTRF job URL from the "Open Job Manager" anchor
  2. Run xtrf_job_setup: login to XTRF, create folders, download source files,
     extract EPO title and write glossary CSV
  3. Download the bilingual XTM Excel and XLF (Xbench paket) via the XTM Workbench API directly into
     the project folder; falls back to copying from the Downloads folder if the
     API download fails; logs XLSX_NOT_FOUND and returns None if neither works
     (workflow will retry on the next run once the file is available)

State logged to project_log.json per Gmail message ID:
  LINK_EXTRACTED          XTRF URL found in email
  JOB_SETUP_OK / FAILED   XTRF job setup result
  XLSX_DOWNLOADED         XTM API download succeeded
  XLSX_FOUND              fallback copy from Downloads succeeded
  XLSX_NOT_FOUND          no xlsx available — needs retry
  JOB_FINISHED_SUCCESSFULLY  all steps completed, email will be skipped next run

FIRST-TIME SETUP
  1. Go to https://console.cloud.google.com/
  2. Create a project, enable the Gmail API.
  3. Create OAuth 2.0 credentials (Desktop app), add your mail address as test
     user, download as gmail_credentials.json and place it next to this file.
  4. Run this script once — a browser window opens for consent.
     The token is saved to gmail_token.json and reused from then on.
"""

import base64
import glob
import os
import re
import shutil
from datetime import datetime

import openpyxl
from email import message_from_bytes
from html.parser import HTMLParser
from pathlib import Path

import requests as _requests
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

import project_log
import xtrf_job_setup
import xtm_initial_download as xtm_download

HERE        = Path(__file__).parent
DOWNLOADS   = Path(r"C:\Users\utasc\Downloads")
GMAIL_LABEL = "ComunicaDK/TODO"
SCOPES      = ["https://www.googleapis.com/auth/gmail.readonly"]
CREDS_FILE  = HERE / "gmail_credentials.json"
TOKEN_FILE  = HERE / "gmail_token.json"


# ── Gmail auth ────────────────────────────────────────────────────────────────

def _get_service():
    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDS_FILE.exists():
                raise FileNotFoundError(
                    f"gmail_credentials.json not found at {CREDS_FILE}\n"
                    "See FIRST-TIME SETUP in the docstring above."
                )
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDS_FILE), SCOPES)
            creds = flow.run_local_server(port=0)
        TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")
    return build("gmail", "v1", credentials=creds)


# ── Label lookup ──────────────────────────────────────────────────────────────

def _get_label_id(service, label_name: str) -> str:
    labels = service.users().labels().list(userId="me").execute().get("labels", [])
    for label in labels:
        if label["name"] == label_name:
            return label["id"]
    available = [l["name"] for l in labels]
    raise RuntimeError(
        f"Gmail label '{label_name}' not found.\nAvailable labels: {available}"
    )


# ── Message parsing ───────────────────────────────────────────────────────────

def _get_header(msg_data: dict, name: str) -> str:
    for h in msg_data["payload"]["headers"]:
        if h["name"].lower() == name.lower():
            return h["value"]
    return ""


def _decode_part(part: dict) -> str:
    data = part.get("body", {}).get("data", "")
    if not data:
        return ""
    return base64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace")


def _get_html_body(msg_data: dict) -> str:
    payload = msg_data["payload"]

    def _find_html(parts):
        for part in parts:
            if part.get("mimeType") == "text/html":
                return _decode_part(part)
            if "parts" in part:
                found = _find_html(part["parts"])
                if found:
                    return found
        return ""

    if payload.get("mimeType") == "text/html":
        return _decode_part(payload)
    return _find_html(payload.get("parts", []))


# ── HTML anchor extraction ────────────────────────────────────────────────────

class _AnchorFinder(HTMLParser):
    """Finds the href of the first anchor whose visible text contains target."""

    def __init__(self, target: str):
        super().__init__()
        self._target  = target
        self._href    = None
        self._in_a    = False
        self._text    = ""
        self.result   = None

    def handle_starttag(self, tag, attrs):
        if tag == "a":
            self._href = dict(attrs).get("href", "")
            self._in_a = True
            self._text = ""

    def handle_endtag(self, tag):
        if tag == "a" and self._in_a:
            if self._target in self._text and self._href:
                self.result = self._href
            self._in_a = False

    def handle_data(self, data):
        if self._in_a:
            self._text += data


def _extract_xtrf_url(html_body: str) -> str | None:
    finder = _AnchorFinder("Open Job Manager")
    finder.feed(html_body)
    href = finder.result
    if not href:
        return None

    # Direct XTRF URL — use as-is
    if "comunicadk.s.xtrf.eu" in href:
        return href

    # Tracking/redirect URL — follow to get the real XTRF URL
    try:
        resp = _requests.head(href, allow_redirects=True, timeout=10)
        if "comunicadk.s.xtrf.eu" in resp.url:
            return resp.url
        # HEAD might be blocked; try GET
        resp = _requests.get(href, allow_redirects=True, timeout=10)
        if "comunicadk.s.xtrf.eu" in resp.url:
            return resp.url
    except Exception as e:
        print(f"  Warning: could not follow redirect ({e}), using raw href.")

    return href  # pass to xtrf_job_setup and let it fail with a clear error


def _extract_project_id(subject: str) -> str | None:
    # Subject: "You can start with job: ... | SYICTL_2604_P0069 (...)"
    m = re.search(r'\b([A-Z]{2,}_\d{4}_[A-Z0-9]+)\b', subject)
    return m.group(1) if m else None


def _parse_deadline(html_body: str) -> datetime | None:
    """Extract deadline from XTRF job email body (format: DD-MM-YYYY HH:MM)."""
    text = re.sub(r"<[^>]+>", " ", html_body)
    text = re.sub(r"\s+", " ", text)
    m = re.search(r"Deadline[^:]*:\s*(\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2})", text)
    if m:
        try:
            return datetime.strptime(m.group(1).strip(), "%d-%m-%Y %H:%M")
        except ValueError:
            return None
    return None


# ── Downloads lookup ──────────────────────────────────────────────────────────

def _find_xlsx(project_id: str) -> Path | None:
    matches = sorted(
        glob.glob(str(DOWNLOADS / f"{project_id}*.xlsx")),
        key=os.path.getmtime,
        reverse=True,
    )
    return Path(matches[0]) if matches else None


# ── Main ──────────────────────────────────────────────────────────────────────

def run(target_project_id: str | None = None) -> str | None:
    """
    Process the oldest unhandled email in ComunicaDK/TODO.
    If target_project_id is given, only the email for that project is processed.
    Returns project_id on success, None if nothing to do or on failure.
    """
    service = _get_service()
    label_id = _get_label_id(service, GMAIL_LABEL)

    result = service.users().messages().list(
        userId="me", labelIds=[label_id], maxResults=50
    ).execute()
    messages = result.get("messages", [])

    if not messages:
        print("ComunicaDK/TODO is empty.")
        return None

    full_log = project_log.get_all_logs()

    # Collect all unprocessed emails
    unprocessed = [
        m for m in messages
        if not any(
            e.get("state") == "JOB_FINISHED_SUCCESSFULLY"
            for e in full_log.get(m["id"], {}).get("events", [])
        )
    ]

    if not unprocessed:
        print(f"No new emails in {GMAIL_LABEL} ({len(messages)} already processed).")
        return None

    # Fetch full data for each and sort by deadline (earliest first)
    msg_cache: dict[str, dict] = {}
    deadlines: dict[str, datetime | None] = {}
    for m in unprocessed:
        mid = m["id"]
        data = service.users().messages().get(userId="me", id=mid, format="full").execute()
        msg_cache[mid] = data
        deadlines[mid] = _parse_deadline(_get_html_body(data))

    if target_project_id:
        unprocessed = [
            m for m in unprocessed
            if _extract_project_id(_get_header(msg_cache[m["id"]], "Subject")) == target_project_id
        ]
        if not unprocessed:
            print(f"ERROR: No unprocessed email found for project {target_project_id}.")
            return None

    unprocessed.sort(key=lambda m: (
        1 if deadlines[m["id"]] is None else 0,
        deadlines[m["id"]] or datetime.max,
    ))

    msg_id = unprocessed[0]["id"]
    msg_data = msg_cache[msg_id]
    dl = deadlines[msg_id]
    print(f"Selected email {msg_id} — deadline: {dl.strftime('%d-%m-%Y %H:%M') if dl else 'unknown'}")

    # Always read subject and XTRF link from the oldest message in the thread.
    # The labeled message may be a reply or forward; the original job notification
    # (with the correct subject and "Open Job Manager" link) is always the oldest.
    thread_id = msg_data.get("threadId")
    if thread_id:
        thread = service.users().threads().get(userId="me", id=thread_id, format="full").execute()
        thread_msgs = sorted(thread.get("messages", []), key=lambda m: int(m.get("internalDate", 0)))
    else:
        thread_msgs = [msg_data]

    oldest_msg = thread_msgs[0] if thread_msgs else msg_data
    subject = _get_header(oldest_msg, "Subject")

    # ── Extract project ID from subject ───────────────────────────────────────
    project_id = _extract_project_id(subject)
    if not project_id:
        project_log.log_event(msg_id, "PARSE_FAILED", detail="no project ID in subject")
        print(f"ERROR: Could not extract project ID from: {subject!r}")
        return None

    print(f"Project ID: {project_id}")

    # ── Extract XTRF URL — search thread oldest-first ─────────────────────────
    xtrf_url = None
    for thread_msg in thread_msgs:
        xtrf_url = _extract_xtrf_url(_get_html_body(thread_msg))
        if xtrf_url:
            if thread_msg["id"] != msg_id:
                print(f"  Found 'Open Job Manager' link in thread message {thread_msg['id']}")
            break

    if not xtrf_url:
        project_log.log_event(msg_id, "PARSE_FAILED", detail="no 'Open Job Manager' link found in thread")
        print("ERROR: Could not find 'Open Job Manager' link in thread.")
        return None

    project_log.log_event(msg_id, "LINK_EXTRACTED", detail=xtrf_url)
    print(f"XTRF URL: {xtrf_url}")

    # ── Step 4: XTRF job setup ────────────────────────────────────────────────
    try:
        setup_result = xtrf_job_setup.run(xtrf_url, project_id_override=project_id)
        project_log.log_event(msg_id, "JOB_SETUP_OK", detail=setup_result["project_folder"])
        print(f"Job setup complete → {setup_result['project_folder']}")
    except Exception as e:
        project_log.log_event(msg_id, "JOB_SETUP_FAILED", detail=str(e))
        print(f"ERROR: Job setup failed — {e}")
        return None

    # ── Download XTM Excel + XLIFF ────────────────────────────────────────────
    proj_dir = project_log.project_dir()
    xlsx_path: Path | None = None

    # 1. Try XTM API: originals → pre-processing, trimmed xlsx + xliff → project folder
    try:
        print("Downloading XTM Excel + XLIFF via API...")
        result = xtm_download.run(project_id)  # dest_folder=None → pre-processing
        xlsx_orig = result["xlsx"]
        xliff_paths = result["xliff"]

        xlsx_dest = proj_dir / xlsx_orig.name
        shutil.copy2(xlsx_orig, xlsx_dest)
        wb = openpyxl.load_workbook(xlsx_dest)
        ws = wb.active
        while ws.max_column > 3:
            ws.delete_cols(ws.max_column)
        wb.save(xlsx_dest)
        xlsx_path = xlsx_dest

        for xlf in xliff_paths:
            shutil.copy2(xlf, proj_dir / xlf.name)

        print(f"  xlsx original: {xlsx_orig.name} (pre-processing)")
        print(f"  xlsx trimmed copy: {xlsx_dest.name} (project folder)")
        for xlf in xliff_paths:
            print(f"  xliff copy: {xlf.name} (project folder)")
        project_log.log_event(msg_id, "XLSX_DOWNLOADED", detail=str(xlsx_path))
    except Exception as e:
        print(f"  XTM download failed ({e}) — falling back to Downloads folder.")

    # 2. Fall back to pre-downloaded file in Downloads
    if xlsx_path is None:
        found = _find_xlsx(project_id)
        if found:
            try:
                pre_folder = xtm_download._find_pre_folder(project_id)
                shutil.copy2(found, pre_folder / found.name)
                print(f"  Original saved: {found.name} (pre-processing)")
            except Exception:
                pass

            dest = proj_dir / found.name
            shutil.copy2(found, dest)
            wb = openpyxl.load_workbook(dest)
            ws = wb.active
            while ws.max_column > 3:
                ws.delete_cols(ws.max_column)
            wb.save(dest)
            xlsx_path = dest
            print(f"  Copied from Downloads: {found.name}")
            project_log.log_event(msg_id, "XLSX_FOUND", detail=str(xlsx_path))

    if xlsx_path is None:
        project_log.log_event(msg_id, "XLSX_NOT_FOUND")
        print(
            f"WARNING: Could not obtain xlsx/xlf for {project_id}.\n"
            f"  Place xlsx and xlf in {proj_dir} and re-run."
        )
        return None

    project_log.log_event(msg_id, "JOB_FINISHED_SUCCESSFULLY", detail=project_id)
    return project_id



if __name__ == "__main__":
    run()
