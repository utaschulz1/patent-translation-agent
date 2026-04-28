# ============================================================
# glossary_compare.py
# ============================================================
# For rule-based pre-processing without LLM
# doesn't look at the translation itself, just the verb/noun extraction output.
# Compares the verb and noun extraction output against the project
# glossary. For each segment pair where the EN term appears in the
# glossary, checks if the DE term used matches the required DE term.
#
# Comparison uses exact (case-insensitive) matching after lemmatisation by
# the upstream verb/noun scripts.
#
# REQUIRES (run these first):
#   LLM_verb_comparison_xlsx.py   → projects/<id>/verb_segment_pairs.csv
#   LLM_noun_comparison_xlsx.py   → projects/<id>/noun_segment_pairs.csv
#
# INPUT   projects/<id>/glossary_*.csv
#         projects/<id>/verb_segment_pairs.csv
#         projects/<id>/noun_segment_pairs.csv
#         projects/<id>/*_checks.xlsx  (or *_translated.xlsx)
# OUTPUT  projects/<id>/glossary_compare_flags.csv
#         projects/<id>/<name>_checks.xlsx  (Glossary Mismatch column)
# ============================================================

import os
import glob
import pandas as pd
import openpyxl
from collections import defaultdict
from dotenv import load_dotenv
from datetime import datetime

from project_log import project_dir as _pdir

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))




proj_dir = _pdir()

# ============================================================
# Load project glossary
# ============================================================

glossary_files = [f for f in glob.glob(str(proj_dir / "glossary_*.csv"))
                  if not any(x in os.path.basename(f) for x in ("results", "flags"))]
if not glossary_files:
    print(f"ERROR: No glossary_*.csv found in '{proj_dir}'.")
    exit()
if len(glossary_files) > 1:
    print(f"Multiple glossary files found, using: {glossary_files[0]}")
glossary_path = glossary_files[0]

try:
    # names= gives pandas enough columns so rows with extra fields (canonical/count/total)
    # are not treated as bad lines and skipped.
    gloss_df = pd.read_csv(glossary_path, encoding="utf-8-sig", comment="#",
                           header=0, usecols=[0, 1])
    gloss_df.columns = ["EN", "DE"]
    en_col, de_col = "EN", "DE"
    glossary_lookup = {}
    for _, r in gloss_df.iterrows():
        en_key = str(r[en_col]).strip().lower()
        de_val = str(r[de_col]).strip()
        if en_key and de_val and en_key != "en":
            glossary_lookup.setdefault(en_key, de_val)  # first occurrence = canonical
except Exception as e:
    print(f"ERROR: Could not load glossary '{glossary_path}': {e}")
    exit()

print(f"Loaded {len(glossary_lookup)} glossary terms from '{glossary_path}'.")

# ============================================================
# Load segment pairs CSVs
# ============================================================

pairs_sources = [
    (str(proj_dir / "verb_segment_pairs.csv"), "en_verb",   "de_verb",   "verb"),
    (str(proj_dir / "noun_segment_pairs.csv"), "en_phrase", "de_phrase", "noun"),
]

all_pairs = []
for path, en_col_name, de_col_name, kind in pairs_sources:
    if not os.path.exists(path):
        print(f"WARNING: '{os.path.basename(path)}' not found — run the {kind} comparison script first.")
        continue
    df = pd.read_csv(path, encoding="utf-8-sig")
    df = df.dropna(subset=[en_col_name, de_col_name])
    for _, row in df.iterrows():
        all_pairs.append({
            "segment_id": str(row["segment_id"]),
            "en_term":    str(row[en_col_name]).strip(),
            "de_used":    str(row[de_col_name]).strip(),
            "kind":       kind,
        })

print(f"Loaded {len(all_pairs)} segment pairs total.")

if not all_pairs:
    print("No segment pairs to check.")
    exit()

# ============================================================
# Compare against glossary
# ============================================================

flags = []
seen = set()

for pair in all_pairs:
    en_key = pair["en_term"].lower()
    expected_de = glossary_lookup.get(en_key)
    if expected_de is None:
        continue

    de_used = pair["de_used"]
    if not de_used:
        continue

    if expected_de.lower().strip() == de_used.lower().strip():
        continue

    dedup_key = (pair["segment_id"], en_key, de_used.lower())
    if dedup_key in seen:
        continue
    seen.add(dedup_key)

    flags.append({
        "segment_id":  pair["segment_id"],
        "kind":        pair["kind"],
        "en_term":     pair["en_term"],
        "expected_de": expected_de,
        "de_used":     de_used,
    })

print(f"\nFlags: {len(flags)} "
      f"({sum(1 for f in flags if f['kind']=='verb')} verb, "
      f"{sum(1 for f in flags if f['kind']=='noun')} noun).")

# ============================================================
# Save flags CSV
# ============================================================

if flags:
    flags_df = pd.DataFrame(flags)
    print(flags_df[["segment_id", "kind", "en_term", "expected_de", "de_used"]].to_string(index=False))
    flags_df.to_csv(str(proj_dir / "compare_flags.csv"), index=False, encoding="utf-8-sig")
    print('\nSaved "compare_flags.csv".')
else:
    print("No glossary mismatches found.")

# ============================================================
# Annotate _checks.xlsx
# ============================================================

patterns = [
    str(proj_dir / "*_translated_checks.xlsx"),
    str(proj_dir / "*_checks.xlsx"),
    str(proj_dir / "*_translated.xlsx"),
    str(proj_dir / "*.xlsx"),
]
xlsx_files = []
for pat in patterns:
    candidates = [
        f for f in glob.glob(pat)
        if not os.path.basename(f).startswith("~$")
    ]
    if candidates:
        xlsx_files = candidates
        break

if not xlsx_files:
    print(f"WARNING: No .xlsx file found in '{proj_dir}' to annotate.")
    exit()
if len(xlsx_files) > 1:
    print(f"Multiple files found, annotating: {xlsx_files[0]}")
wb_path = xlsx_files[0]

seg_notes = defaultdict(list)
for f in flags:
    note = f'GLOSSARY: "{f["en_term"]}" — expected "{f["expected_de"]}", found "{f["de_used"]}"'
    seg_notes[f["segment_id"]].append(note)

wb = openpyxl.load_workbook(wb_path)
ws = wb.active

# Reuse existing "Glossary Mismatch" column if present, else append.
insert_col = None
for c in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=c).value == "Glossary Mismatch":
        insert_col = c
        break
if insert_col is None:
    insert_col = ws.max_column + 1
    ws.cell(row=1, column=insert_col).value = "Glossary Mismatch"

# Clear any previous values in that column.
for row_num in range(2, ws.max_row + 1):
    ws.cell(row=row_num, column=insert_col).value = None

for row_num in range(4, ws.max_row + 1):
    raw = ws.cell(row=row_num, column=1).value
    # Normalise: drop ".0" suffix so "17.0" matches "17"
    seg_id = str(int(float(raw))) if raw is not None else ""
    if seg_id in seg_notes:
        ws.cell(row=row_num, column=insert_col).value = "\n".join(seg_notes[seg_id])
        ws.cell(row=row_num, column=insert_col).alignment = openpyxl.styles.Alignment(wrap_text=True)

checks_path = wb_path if "_checks" in wb_path else wb_path.replace(".xlsx", "_checks.xlsx")
try:
    wb.save(checks_path)
    print(f'Saved annotated Excel: "{checks_path}".')
except PermissionError:
    stamp = datetime.now().strftime("%H%M%S")
    fallback = checks_path.replace("_checks.xlsx", f"_checks_{stamp}.xlsx")
    wb.save(fallback)
    print(f'File was open in Excel — saved as "{fallback}" instead.')
